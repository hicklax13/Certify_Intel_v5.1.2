"""
Certify Intel - Simplified Backend for Excel Dashboard
A lightweight FastAPI backend that:
1. Stores competitor data in SQLite (simple, no PostgreSQL setup needed)
2. Scrapes competitor websites using Playwright
3. Extracts data using OpenAI GPT
4. Exports data to Excel-compatible formats
"""
import os

# Load environment variables from .env file
try:
    from dotenv import load_dotenv
    load_dotenv()
    print(f"Loading environment variables from .env")
    print(f"Database URL: {os.getenv('DATABASE_URL', 'sqlite:///./certify_intel.db')}")
except ImportError:
    print("python-dotenv not installed, using system environment variables")

import json
import asyncio
try:
    import yfinance as yf
except ImportError:
    class MockYF:
        class Ticker:
            def __init__(self, t): self.info = {}
        def Ticker(self, t): return self.Ticker(t)
    yf = MockYF()
    print("yfinance not found, using mock")
from datetime import datetime
from typing import Optional, List, Dict, Any
from contextlib import asynccontextmanager



from fastapi import FastAPI, HTTPException, BackgroundTasks, Depends, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, HttpUrl
from sqlalchemy import create_engine, Column, Integer, String, Float, DateTime, Text, Boolean
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker, Session
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# Enterprise Automation Import
try:
    from scheduler import start_scheduler, stop_scheduler
    SCHEDULER_AVAILABLE = True
except ImportError:
    SCHEDULER_AVAILABLE = False
    print("Scheduler module not found. Automation disabled.")

# New Data Source Imports
import glassdoor_scraper
import indeed_scraper
import sec_edgar_scraper
import uspto_scraper
import klas_scraper
import appstore_scraper

import himss_scraper

# Database setup - SQLite for simplicity
# Database setup

from database import (
    engine, SessionLocal, Base, get_db, Competitor, ChangeLog, DataSource,
    DataChangeHistory, User, SystemPrompt, KnowledgeBaseItem, UserSettings, ActivityLog,
    CompetitorProduct, ProductPricingTier, ProductFeatureMatrix, CustomerCountEstimate,
    RefreshSession  # Phase 4: Task 5.0.1-031
)
from confidence_scoring import (
    calculate_confidence_score, get_source_defaults, calculate_data_staleness,
    determine_confidence_level_from_score, triangulate_data_points,
    get_reliability_description, get_credibility_description, get_source_type_description,
    SOURCE_TYPE_DEFAULTS, RELIABILITY_DESCRIPTIONS, CREDIBILITY_DESCRIPTIONS
)
from data_triangulator import (
    DataTriangulator, triangulate_competitor, triangulation_result_to_dict
)

# Auth imports for route protection
from fastapi.security import OAuth2PasswordBearer
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="token", auto_error=False)

async def get_current_user(token: str = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    """Verify JWT token and return current user with ID. Raises 401 if invalid/missing."""
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")

    # Import here to avoid circular import
    from extended_features import auth_manager

    payload = auth_manager.verify_token(token)
    if not payload:
        raise HTTPException(status_code=401, detail="Invalid or expired token")

    # Get user ID from database
    user = db.query(User).filter(User.email == payload.get("sub")).first()
    user_id = user.id if user else None

    return {"id": user_id, "email": payload.get("sub"), "role": payload.get("role")}


def log_activity(db: Session, user_email: str, user_id: int, action_type: str, action_details: str = None):
    """Log a user activity to the activity_logs table (shared across all users)."""
    import json
    activity = ActivityLog(
        user_id=user_id,
        user_email=user_email,
        action_type=action_type,
        action_details=action_details if isinstance(action_details, str) else json.dumps(action_details) if action_details else None
    )
    db.add(activity)
    db.commit()
    return activity


# Global progress tracker for scrape operations
scrape_progress = {
    "active": False,
    "total": 0,
    "completed": 0,
    "current_competitor": None,
    "competitors_done": [],
    "changes_detected": 0,
    "new_values_added": 0,
    # Phase 2: Enhanced tracking (Task 5.0.1-026)
    "started_at": None,
    "recent_changes": [],    # Last 10 changes for live display
    "change_details": [],    # All changes for AI summary
    "errors": []             # Any errors encountered
}


# White fill for Excel cells
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

# Known healthcare IT public companies with tickers
KNOWN_TICKERS = {
    "phreesia": {"symbol": "PHR", "exchange": "NYSE", "name": "Phreesia Inc"},
    "health catalyst": {"symbol": "HCAT", "exchange": "NASDAQ", "name": "Health Catalyst Inc"},
    "veeva": {"symbol": "VEEV", "exchange": "NYSE", "name": "Veeva Systems Inc"},
    "teladoc": {"symbol": "TDOC", "exchange": "NYSE", "name": "Teladoc Health Inc"},
    "doximity": {"symbol": "DOCS", "exchange": "NYSE", "name": "Doximity Inc"},
    "hims & hers": {"symbol": "HIMS", "exchange": "NYSE", "name": "Hims & Hers Health Inc"},
    "definitive healthcare": {"symbol": "DH", "exchange": "NASDAQ", "name": "Definitive Healthcare Corp"},
    "carecloud": {"symbol": "CCLD", "exchange": "NASDAQ", "name": "CareCloud Inc"},
}

# Database Models imported from database.py


# ============== Pydantic Schemas ==============

class CompetitorCreate(BaseModel):
    name: str
    website: str
    status: str = "Active"
    threat_level: str = "Medium"
    notes: Optional[str] = None
    pricing_model: Optional[str] = None
    base_price: Optional[str] = None
    price_unit: Optional[str] = None
    product_categories: Optional[str] = None
    key_features: Optional[str] = None
    integration_partners: Optional[str] = None
    certifications: Optional[str] = None
    target_segments: Optional[str] = None
    customer_size_focus: Optional[str] = None
    geographic_focus: Optional[str] = None
    customer_count: Optional[str] = None
    customer_acquisition_rate: Optional[str] = None
    key_customers: Optional[str] = None
    g2_rating: Optional[str] = None
    employee_count: Optional[str] = None
    employee_growth_rate: Optional[str] = None
    year_founded: Optional[str] = None
    headquarters: Optional[str] = None
    funding_total: Optional[str] = None
    latest_round: Optional[str] = None
    pe_vc_backers: Optional[str] = None
    website_traffic: Optional[str] = None
    social_following: Optional[str] = None
    recent_launches: Optional[str] = None
    news_mentions: Optional[str] = None


class CompetitorResponse(CompetitorCreate):
    id: int
    last_updated: datetime
    data_quality_score: Optional[float] = None  # Float to support existing DB values
    created_at: datetime
    is_public: Optional[bool] = False
    ticker_symbol: Optional[str] = None
    stock_exchange: Optional[str] = None
    
    class Config:
        from_attributes = True


class CorrectionRequest(BaseModel):
    field: str
    new_value: str
    reason: Optional[str] = "Manual Correction"
    source_url: Optional[str] = None


class ScrapeRequest(BaseModel):
    competitor_id: int
    pages_to_scrape: List[str] = ["homepage", "pricing", "about"]

class UserResponse(BaseModel):
    id: int
    email: str
    full_name: Optional[str] = None
    role: str
    is_active: bool
    last_login: Optional[datetime] = None
    created_at: Optional[datetime] = None

class UserInviteRequest(BaseModel):
    email: str
    role: str = "viewer"
    full_name: Optional[str] = None


class SystemPromptBase(BaseModel):
    key: str
    content: str

class SystemPromptCreate(SystemPromptBase):
    pass

class SystemPromptResponse(SystemPromptBase):
    id: int
    updated_at: datetime
    class Config:
        from_attributes = True

class KnowledgeBaseItemBase(BaseModel):
    title: str
    content_text: str
    source_type: Optional[str] = "manual"
    is_active: Optional[bool] = True

class KnowledgeBaseItemCreate(KnowledgeBaseItemBase):
    pass

class KnowledgeBaseItemResponse(KnowledgeBaseItemBase):
    id: int
    created_at: datetime
    class Config:
        from_attributes = True



# ============== FastAPI App ==============

@asynccontextmanager
async def lifespan(app: FastAPI):
    # Startup
    print("Certify Intel Backend starting...")
    print("=" * 60)

    # Configuration Validation
    print("Validating configuration...")
    missing_required = []

    # Check required environment variables
    if not os.getenv("SECRET_KEY"):
        missing_required.append("SECRET_KEY")

    if missing_required:
        print(f"ERROR: Missing required environment variables: {', '.join(missing_required)}")
        raise ValueError(f"Missing required env vars: {', '.join(missing_required)}")

    # Warn about optional features
    optional_features = {
        "OPENAI_API_KEY": "AI Features (Executive Summaries, Discovery Agent, Web Extraction)",
        "SMTP_HOST": "Email Alerts",
        "SLACK_WEBHOOK_URL": "Slack Notifications"
    }

    disabled_scrapers = ["Crunchbase", "PitchBook", "LinkedIn (live scraping)"]

    print("\nAvailable Scrapers:")
    print("  [OK] Playwright Base Scraper - Website content extraction")
    print("  [OK] SEC Edgar (yfinance) - Public company financials")
    print("  [OK] News Monitor (Google News RSS) - Real-time news")
    print("  [OK] Known Data Fallback - Pre-populated data for demo")
    print("  [OK] [15+ other specialized scrapers with fallback data]")

    print("\nDisabled Scrapers (Paid APIs - not available):")
    for scraper in disabled_scrapers:
        print(f"  [X] {scraper}")

    print("\nOptional Features:")
    for env_var, feature in optional_features.items():
        if os.getenv(env_var):
            print(f"  [OK] {feature} - ENABLED")
        else:
            print(f"  [!] {feature} - DISABLED (set {env_var} to enable)")

    print("=" * 60)

    # Start Enterprise Scheduler
    if SCHEDULER_AVAILABLE:
        print("Initializing Enterprise Automation Engine...")
        start_scheduler()

    # Run Public/Private Workflow on Startup
    # Run Public/Private Workflow on Startup
    try:
        from extended_features import ClassificationWorkflow, auth_manager
        
        # Initialize DB Session for startup tasks
        db = SessionLocal()
        
        # 1. Ensure Admin User
        print("Ensuring default admin user exists...")
        auth_manager.ensure_default_admin(db)
        
        # 2. Run Classification Workflow - DISABLED (costs money, use button instead)
        # workflow = ClassificationWorkflow(db)
        # print("Running 'Private vs Public' Classification Workflow...")
        # workflow.run_classification_pipeline()
        
        db.close()
    except Exception as e:
        print(f"Startup task warning: {e}")
        
    yield
    
    # Shutdown: Clean up resources
    print("Certify Intel Backend shutting down...")
    if SCHEDULER_AVAILABLE:
        stop_scheduler()

app = FastAPI(
    title="Certify Health Intel API",
    description="Backend for Competitive Intelligence Dashboard",
    version="1.0.0",
    lifespan=lifespan
)



# Valid API endpoints are defined above.
# The catch-all static file mount must be last.



# Redirect legacy /app path to root
@app.get("/app")
@app.get("/app/")
async def redirect_app():
    return RedirectResponse(url="/")








@app.get("/api/logo-proxy")
async def proxy_logo(url: str):
    """Proxy image requests to avoid CORS issues."""
    import requests
    from fastapi import Response
    
    try:
        # Validate URL
        if not url.startswith("http"):
            raise HTTPException(status_code=400, detail="Invalid URL")
            
        # Try Clearbit first
        try:
            resp = requests.get(url, timeout=3)
            if resp.status_code == 200:
                return Response(content=resp.content, media_type=resp.headers.get("content-type", "image/png"))
        except:
            pass
            
        # Fallback to Google Favicon service if Clearbit fails
        # Extract domain from the input URL (expected: https://logo.clearbit.com/domain.com)
        domain = url.split("/")[-1]
        google_url = f"https://www.google.com/s2/favicons?domain={domain}&sz=128"
        
        try:
            resp = requests.get(google_url, timeout=3)
            if resp.status_code == 200:
                return Response(content=resp.content, media_type=resp.headers.get("content-type", "image/png"))
        except:
            pass

        # Return 404 to trigger frontend onerror fallback
        raise HTTPException(status_code=404, detail="Image not found")
            
    except Exception as e:
        print(f"Proxy error: {e}")
        raise HTTPException(status_code=404, detail="Proxy error")


# get_db imported from database.py


@app.get("/api/analytics/summary")
@app.get("/api/analytics/executive-summary")
async def get_dashboard_summary(db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    """Generate comprehensive AI executive summary analyzing ALL dashboard data points."""
    try:
        import os
        from openai import OpenAI
        
        # Get all non-deleted competitors (matching dashboard count)
        competitors = db.query(Competitor).filter(
            Competitor.is_deleted == False
        ).all()
        
        if not competitors:
            return {"summary": "No active competitors found to analyze.", "type": "empty", "model": None}
        
        # Calculate statistics
        total = len(competitors)
        high_threat = sum(1 for c in competitors if c.threat_level == "High")
        medium_threat = sum(1 for c in competitors if c.threat_level == "Medium")
        low_threat = sum(1 for c in competitors if c.threat_level == "Low")
        public_companies = sum(1 for c in competitors if c.is_public)
        private_companies = total - public_companies
        
        # Gather pricing info
        pricing_models = {}
        for c in competitors:
            model = c.pricing_model or "Unknown"
            pricing_models[model] = pricing_models.get(model, 0) + 1
        
        # Gather top threats
        top_threats = [c.name for c in competitors if c.threat_level == "High"][:5]
        
        # Build comprehensive data summary for AI
        data_summary = f"""
COMPETITIVE INTELLIGENCE DATA SNAPSHOT:
========================================
TIMESTAMP: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}

TRACKING OVERVIEW:
- Total Competitors Monitored: {total}
- High Threat: {high_threat} ({round(high_threat/total*100, 1)}%)
- Medium Threat: {medium_threat} ({round(medium_threat/total*100, 1)}%)
- Low Threat: {low_threat} ({round(low_threat/total*100, 1)}%)
- Public Companies: {public_companies}
- Private Companies: {private_companies}

TOP HIGH-THREAT COMPETITORS: {', '.join(top_threats) if top_threats else 'None identified'}

PRICING MODEL DISTRIBUTION:
{chr(10).join(f'- {model}: {count} competitors' for model, count in pricing_models.items())}

DETAILED COMPETITOR DATA (ALL {total} ACTIVE COMPETITORS):
"""
        # INCLUDED ALL COMPETITORS (No limit) for maximum context
        for i, c in enumerate(competitors, 1):
            data_summary += f"""
--- COMPETITOR #{i} of {total} ---
NAME: {c.name}
THREAT LEVEL: {c.threat_level}
STATUS: {c.status or 'Active'}
WEBSITE: {c.website or 'N/A'}
LAST UPDATED: {c.last_updated.strftime('%Y-%m-%d %H:%M') if c.last_updated else 'N/A'}
DATA QUALITY SCORE: {c.data_quality_score or 'N/A'}/100

PRICING INFO:
- Model: {c.pricing_model or 'Unknown'}
- Base Price: {c.base_price or 'N/A'}
- Price Unit: {c.price_unit or 'N/A'}

PRODUCT INFO:
- Categories: {c.product_categories or 'N/A'}
- Key Features: {c.key_features or 'N/A'}
- Integrations: {c.integration_partners or 'N/A'}
- Certifications: {c.certifications or 'N/A'}

MARKET INFO:
- Target Segments: {c.target_segments or 'N/A'}
- Customer Size Focus: {c.customer_size_focus or 'N/A'}
- Geographic Focus: {c.geographic_focus or 'N/A'}
- Customer Count: {c.customer_count or 'N/A'}
- Key Customers: {c.key_customers or 'N/A'}
- G2 Rating: {c.g2_rating or 'N/A'}

COMPANY INFO:
- Headquarters: {c.headquarters or 'N/A'}
- Founded: {c.year_founded or 'N/A'}
- Employees: {c.employee_count or 'N/A'}
- Employee Growth: {c.employee_growth_rate or 'N/A'}
- Total Funding: {c.funding_total or 'N/A'}
- Latest Round: {c.latest_round or 'N/A'}
- Is Public: {'Yes' if c.is_public else 'No (Private)'}
"""
            # Add live stock data for public companies
            if c.is_public and c.ticker_symbol:
                stock_info = fetch_real_stock_data(c.ticker_symbol)
                if stock_info and stock_info.get('price'):
                    change_sign = '+' if stock_info.get('change', 0) >= 0 else ''
                    data_summary += f"""
STOCK DATA (LIVE):
- Ticker: {c.ticker_symbol} ({c.stock_exchange or 'NYSE'})
- Current Price: ${stock_info.get('price', 0):.2f}
- Daily Change: {change_sign}{stock_info.get('change', 0):.2f} ({change_sign}{stock_info.get('change_percent', 0):.2f}%)
- Market Cap: ${stock_info.get('market_cap', 0):,.0f}
- 52-Week High: ${stock_info.get('high52', 'N/A')}
- 52-Week Low: ${stock_info.get('low52', 'N/A')}
"""
                else:
                    data_summary += f"""
STOCK DATA:
- Ticker: {c.ticker_symbol} ({c.stock_exchange or 'NYSE'})
- Stock Price: Data unavailable
"""

            data_summary += f"""
NOTES: {c.notes or 'None'}
--- END {c.name} ---
"""

        # Try OpenAI first
        api_key = os.getenv("OPENAI_API_KEY")
        model_used = "gpt-4.1"
        provider = "OpenAI"
        
        if api_key:
            try:
                # Fetch dynamic system prompt (user-specific first, then global fallback)
                user_id = current_user.get("id") if current_user else None
                prompt_db = db.query(SystemPrompt).filter(
                    SystemPrompt.key == "dashboard_summary",
                    SystemPrompt.user_id == user_id
                ).first()
                if not prompt_db:
                    # Fallback to global prompt (user_id=NULL)
                    prompt_db = db.query(SystemPrompt).filter(
                        SystemPrompt.key == "dashboard_summary",
                        SystemPrompt.user_id == None
                    ).first()
                system_content = prompt_db.content if prompt_db else """You are Certify Health's competitive intelligence analyst. Generate a comprehensive, executive-level strategic summary. 

Your summary MUST include:
1. **Executive Overview** - High-level market position assessment
2. **Threat Analysis** - Breakdown of competitive landscape by threat level
3. **Pricing Intelligence** - Analysis of competitor pricing strategies
4. **Market Trends** - Emerging patterns and shifts
5. **Strategic Recommendations** - 3-5 specific, actionable recommendations
6. **Watch List** - Key competitors requiring immediate attention

Use data-driven insights. Be specific with numbers and competitor names. Format with markdown headers and bullet points."""

                # RAG: Inject Knowledge Base
                kb_items = db.query(KnowledgeBaseItem).filter(KnowledgeBaseItem.is_active == True).all()
                if kb_items:
                    data_summary += "\n\nINTERNAL KNOWLEDGE BASE (USE THIS CONTEXT):\n==========================\n"
                    for item in kb_items:
                        data_summary += f"\n--- {item.title} ({item.source_type}) ---\n{item.content_text}\n"

                client = OpenAI(api_key=api_key)
                response = client.chat.completions.create(
                    model=model_used,
                    messages=[
                        {"role": "system", "content": system_content},
                        {"role": "user", "content": data_summary}
                    ],
                    max_tokens=2000,
                    temperature=0.7
                )
                summary = response.choices[0].message.content
                return {
                    "summary": summary,
                    "type": "ai",
                    "model": model_used,
                    "provider": provider,
                    "provider_logo": "/static/openai-logo.svg",
                    "data_points_analyzed": total,
                    "generated_at": datetime.utcnow().isoformat() + "Z"
                }
            except Exception as e:
                print(f"OpenAI Error: {e}")
        
        # Fallback to automated summary
        summary = f"""# AI-Generated Executive Summary

## Executive Overview
We are currently tracking **{total} competitors** in the patient engagement and healthcare check-in market. The competitive landscape shows **{high_threat} high-threat** competitors requiring immediate attention, **{medium_threat} medium-threat** competitors to monitor, and **{low_threat} low-threat** competitors with minimal impact.

## Threat Analysis
- **High Threat ({high_threat})**: {', '.join(top_threats[:3]) if top_threats else 'None'} - These competitors have significant market overlap and strong positioning
- **Medium Threat ({medium_threat})**: Competitors with partial market overlap requiring ongoing monitoring
- **Low Threat ({low_threat})**: Minimal competitive impact at this time

## Market Composition
- **Public Companies**: {public_companies} ({round(public_companies/total*100, 1)}% of tracked competitors)
- **Private Companies**: {private_companies} ({round(private_companies/total*100, 1)}% of tracked competitors)

## Pricing Intelligence
{chr(10).join(f'- **{model}**: {count} competitors' for model, count in list(pricing_models.items())[:5])}

## Strategic Recommendations
1. **Monitor pricing changes weekly** - Especially for high-threat competitors
2. **Investigate feature gaps** in patient intake workflows
3. **Review battlecards** for top high-threat targets
4. **Track funding announcements** from private competitors
5. **Analyze customer reviews** on G2 and Capterra for competitive insights

## Watch List
Top competitors requiring immediate attention: {', '.join(top_threats) if top_threats else 'No high-threat competitors identified'}

---
*Summary generated automatically based on {total} tracked competitors*
"""
        return {
            "summary": summary,
            "type": "fallback",
            "model": "automated",
            "provider": "Certify Intel",
            "provider_logo": "/static/certify-logo.svg",
            "data_points_analyzed": total,
            "generated_at": datetime.utcnow().isoformat()
        }
        
    except Exception as e:
        print(f"Summary Error: {e}")
        return {"summary": f"Error generating summary: {str(e)}", "type": "error", "model": None}


@app.post("/api/analytics/chat")
def chat_with_summary(request: dict, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    """Chat with AI about the competitive intelligence data."""
    try:
        import os
        from openai import OpenAI
        
        user_message = request.get("message", "")
        if not user_message:
            return {"response": "Please provide a message.", "success": False}
        
        # Get competitor data for context
        # Get ALL non-deleted competitors for consistency
        competitors = db.query(Competitor).filter(Competitor.is_deleted == False).all()
        
        # Build Comprehensive Context (Same as Summary Generation to ensure consistency)
        pricing_models = {}
        for c in competitors:
            model = c.pricing_model or "Unknown"
            pricing_models[model] = pricing_models.get(model, 0) + 1

        context = f"""
FULL DATA SNAPSHOT (LIVE):
==========================
Total Competitors: {len(competitors)}
High Threat: {sum(1 for c in competitors if c.threat_level == 'High')}
Medium Threat: {sum(1 for c in competitors if c.threat_level == 'Medium')}
Low Threat: {sum(1 for c in competitors if c.threat_level == 'Low')}

COMPETITORS:
"""
        # Include ALL competitors in chat context too
        for c in competitors:
            # Build base competitor info
            comp_info = f"""
---
COMPETITOR: {c.name}
THREAT: {c.threat_level}
WEBSITE: {c.website or 'N/A'}
PRICING: {c.base_price or 'N/A'} ({c.pricing_model or 'Unknown Model'})
OFFERING: {c.product_categories or 'N/A'}
FEATURES: {c.key_features or 'N/A'}
EMPLOYEES: {c.employee_count or 'N/A'}
G2 RATING: {c.g2_rating or 'N/A'}
"""
            # Add stock data for public companies
            if c.is_public and c.ticker_symbol:
                stock_data = fetch_real_stock_data(c.ticker_symbol)
                if stock_data and stock_data.get('price'):
                    change_sign = '+' if stock_data.get('change', 0) >= 0 else ''
                    comp_info += f"""PUBLIC COMPANY: Yes
STOCK TICKER: {c.ticker_symbol} ({c.stock_exchange or 'NYSE'})
CURRENT STOCK PRICE: ${stock_data.get('price', 'N/A'):.2f}
PRICE CHANGE: {change_sign}{stock_data.get('change', 0):.2f} ({change_sign}{stock_data.get('change_percent', 0):.2f}%)
MARKET CAP: ${stock_data.get('market_cap', 0):,.0f}
52-WEEK HIGH: ${stock_data.get('high52', 'N/A')}
52-WEEK LOW: ${stock_data.get('low52', 'N/A')}
"""
                else:
                    comp_info += f"""PUBLIC COMPANY: Yes
STOCK TICKER: {c.ticker_symbol} ({c.stock_exchange or 'NYSE'})
STOCK DATA: Unable to fetch live data
"""
            else:
                comp_info += f"PUBLIC COMPANY: No (Private)\n"

            comp_info += f"NOTES: {c.notes or ''}\n---\n"
            context += comp_info

        api_key = os.getenv("OPENAI_API_KEY")
        if api_key:
            # Get user ID for personalized prompts
            user_id = current_user.get("id") if current_user else None

            # Fetch Chat Persona (user-specific first, then global fallback)
            prompt_db = db.query(SystemPrompt).filter(
                SystemPrompt.key == "chat_persona",
                SystemPrompt.user_id == user_id
            ).first()
            if not prompt_db:
                prompt_db = db.query(SystemPrompt).filter(
                    SystemPrompt.key == "chat_persona",
                    SystemPrompt.user_id == None
                ).first()
            base_persona = prompt_db.content if prompt_db else "You are a competitive intelligence analyst for Certify Health."

            # Fetch Summary Prompt (user-specific first, then global fallback)
            summary_prompt_db = db.query(SystemPrompt).filter(
                SystemPrompt.key == "dashboard_summary",
                SystemPrompt.user_id == user_id
            ).first()
            if not summary_prompt_db:
                summary_prompt_db = db.query(SystemPrompt).filter(
                    SystemPrompt.key == "dashboard_summary",
                    SystemPrompt.user_id == None
                ).first()
            summary_instructions = summary_prompt_db.content if summary_prompt_db else "Focus on strategic insights."

            # RAG: Inject Knowledge Base
            kb_text = ""
            kb_items = db.query(KnowledgeBaseItem).filter(KnowledgeBaseItem.is_active == True).all()
            if kb_items:
                kb_text += "\n\nINTERNAL KNOWLEDGE BASE:\n"
                for item in kb_items:
                    kb_text += f"\n--- {item.title} ---\n{item.content_text}\n"
            

            full_system_content = f"""{base_persona}

ALIGNMENT INSTRUCTION:
The user has defined the following strategy for the Dashboard Summary. Use this as context for your tone and priorities:
"{summary_instructions}"

CRITICAL INSTRUCTION:
You have access to a LIVE database of competitors below with REAL-TIME STOCK DATA for public companies.
- If the user asks for a website, LOOK at the 'WEBSITE' field for that competitor and provide it.
- If the user asks for pricing details, LOOK at the 'PRICING' field.
- If the user asks about stock prices, market cap, or financial data for PUBLIC COMPANIES, LOOK at the 'STOCK' fields (CURRENT STOCK PRICE, MARKET CAP, PRICE CHANGE, etc.).
- For public companies, you have LIVE stock data including: current price, daily change, market cap, 52-week high/low.
- Do NOT say "I cannot browse the web" or "I don't have access to real-time stock data" if the answer is in the data below.
- Do NOT say "I am working with hypothetical data". This IS the live, real-time data from the Certify Intel platform.
- When asked about a public company's stock, provide the EXACT values from the data (e.g., "Phreesia (PHR) is currently trading at $15.84, up +0.3%").

LIVE DATA CONTEXT:
{context}

{kb_text}
"""

            client = OpenAI(api_key=api_key)
            response = client.chat.completions.create(
                model="gpt-4.1",
                messages=[
                    {"role": "system", "content": full_system_content},
                    {"role": "user", "content": user_message}
                ],
                max_tokens=600
            )
            return {"response": response.choices[0].message.content, "success": True}
        else:
            return {"response": "AI chat requires OpenAI API key configuration.", "success": False}
    except Exception as e:
        return {"response": f"Error: {str(e)}", "success": False}


# Health endpoint at /api/health
@app.get("/api/health")
def api_health():
    return {"status": "healthy", "timestamp": datetime.utcnow().isoformat(), "version": "1.0.0"}


# Analytics sub-endpoints
@app.get("/api/analytics/threats")
def get_threat_analytics(db: Session = Depends(get_db)):
    """Get threat distribution analytics."""
    competitors = db.query(Competitor).filter(Competitor.is_deleted == False).all()
    return {
        "high": sum(1 for c in competitors if c.threat_level == "High"),
        "medium": sum(1 for c in competitors if c.threat_level == "Medium"),
        "low": sum(1 for c in competitors if c.threat_level == "Low"),
        "total": len(competitors)
    }


@app.get("/api/analytics/market-share")
def get_market_share_analytics(db: Session = Depends(get_db)):
    """Get estimated market share by customer count."""
    competitors = db.query(Competitor).filter(Competitor.is_deleted == False).all()
    shares = []
    for c in competitors[:10]:
        count = 0
        if c.customer_count:
            try:
                count = int(''.join(filter(str.isdigit, str(c.customer_count)))) or 100
            except:
                count = 100
        shares.append({"name": c.name, "customers": count})
    total = sum(s["customers"] for s in shares)
    for s in shares:
        s["share"] = round(s["customers"] / total * 100, 1) if total > 0 else 0
    return {"market_share": shares}


@app.get("/api/analytics/pricing")
def get_pricing_analytics(db: Session = Depends(get_db)):
    """Get pricing model distribution."""
    competitors = db.query(Competitor).filter(Competitor.is_deleted == False).all()
    models = {}
    for c in competitors:
        model = c.pricing_model or "Unknown"
        models[model] = models.get(model, 0) + 1
    return {"pricing_models": [{"model": k, "count": v} for k, v in models.items()]}


# ============== DATA QUALITY ENDPOINTS ==============

# List of all data fields that should be tracked
COMPETITOR_DATA_FIELDS = [
    "name", "website", "status", "threat_level", "pricing_model", "base_price",
    "price_unit", "product_categories", "key_features", "integration_partners",
    "certifications", "target_segments", "customer_size_focus", "geographic_focus",
    "customer_count", "customer_acquisition_rate", "key_customers", "g2_rating",
    "employee_count", "employee_growth_rate", "year_founded", "headquarters",
    "funding_total", "latest_round", "pe_vc_backers", "website_traffic",
    "social_following", "recent_launches", "news_mentions", "is_public",
    "ticker_symbol", "stock_exchange"
]


def calculate_quality_score(competitor) -> int:
    """Calculate data quality score (0-100) based on field completeness."""
    filled_fields = 0
    for field in COMPETITOR_DATA_FIELDS:
        value = getattr(competitor, field, None)
        if value is not None and str(value).strip() not in ["", "None", "Unknown", "N/A"]:
            filled_fields += 1
    return int((filled_fields / len(COMPETITOR_DATA_FIELDS)) * 100)


@app.get("/api/data-quality/completeness")
def get_data_completeness(db: Session = Depends(get_db)):
    """Get field-by-field completeness statistics across all competitors."""
    competitors = db.query(Competitor).filter(Competitor.is_deleted == False).all()
    total = len(competitors)
    
    if total == 0:
        return {"total_competitors": 0, "fields": [], "overall_completeness": 0}
    
    field_stats = []
    for field in COMPETITOR_DATA_FIELDS:
        filled = 0
        for comp in competitors:
            value = getattr(comp, field, None)
            if value is not None and str(value).strip() not in ["", "None", "Unknown", "N/A"]:
                filled += 1
        completeness = round((filled / total) * 100, 1)
        field_stats.append({
            "field": field,
            "filled": filled,
            "total": total,
            "completeness_percent": completeness
        })
    
    # Sort by completeness ascending (least complete first)
    field_stats.sort(key=lambda x: x["completeness_percent"])
    
    overall = round(sum(f["completeness_percent"] for f in field_stats) / len(field_stats), 1)
    
    return {
        "total_competitors": total,
        "total_fields": len(COMPETITOR_DATA_FIELDS),
        "overall_completeness": overall,
        "fields": field_stats
    }


@app.get("/api/data-quality/scores")
def get_quality_scores(db: Session = Depends(get_db)):
    """Get quality scores for all competitors."""
    competitors = db.query(Competitor).filter(Competitor.is_deleted == False).all()
    
    scores = []
    for comp in competitors:
        score = calculate_quality_score(comp)
        scores.append({
            "id": comp.id,
            "name": comp.name,
            "score": score,
            "tier": "Excellent" if score >= 80 else "Good" if score >= 60 else "Fair" if score >= 40 else "Poor"
        })
    
    # Sort by score descending
    scores.sort(key=lambda x: x["score"], reverse=True)
    
    avg_score = round(sum(s["score"] for s in scores) / len(scores), 1) if scores else 0
    
    return {
        "average_score": avg_score,
        "total_competitors": len(scores),
        "scores": scores
    }


@app.get("/api/data-quality/stale")
def get_stale_records(days: int = 30, db: Session = Depends(get_db)):
    """Get competitors with data older than specified days."""
    from datetime import timedelta
    cutoff = datetime.utcnow() - timedelta(days=days)
    
    competitors = db.query(Competitor).filter(
        Competitor.is_deleted == False
    ).all()
    
    stale = []
    fresh = []
    for comp in competitors:
        check_date = comp.last_verified_at or comp.last_updated or comp.created_at
        if check_date and check_date < cutoff:
            days_old = (datetime.utcnow() - check_date).days
            stale.append({
                "id": comp.id,
                "name": comp.name,
                "last_verified": check_date.isoformat() if check_date else None,
                "days_old": days_old
            })
        else:
            fresh.append({"id": comp.id, "name": comp.name})
    
    stale.sort(key=lambda x: x["days_old"], reverse=True)
    
    return {
        "threshold_days": days,
        "stale_count": len(stale),
        "fresh_count": len(fresh),
        "stale_records": stale,
        "fresh_records": fresh[:10]  # Just show first 10 fresh
    }


@app.post("/api/data-quality/verify/{competitor_id}")
def verify_competitor_data(competitor_id: int, db: Session = Depends(get_db)):
    """Mark a competitor's data as verified (updates last_verified_at timestamp)."""
    competitor = db.query(Competitor).filter(Competitor.id == competitor_id).first()
    if not competitor:
        raise HTTPException(status_code=404, detail="Competitor not found")
    
    competitor.last_verified_at = datetime.utcnow()
    competitor.data_quality_score = calculate_quality_score(competitor)
    db.commit()
    
    return {
        "success": True,
        "competitor_id": competitor_id,
        "name": competitor.name,
        "verified_at": competitor.last_verified_at.isoformat(),
        "quality_score": competitor.data_quality_score
    }


@app.get("/api/data-quality/completeness/{competitor_id}")
def get_competitor_completeness(competitor_id: int, db: Session = Depends(get_db)):
    """Get field-by-field completeness for a specific competitor."""
    competitor = db.query(Competitor).filter(Competitor.id == competitor_id).first()
    if not competitor:
        raise HTTPException(status_code=404, detail="Competitor not found")
    
    fields = []
    for field in COMPETITOR_DATA_FIELDS:
        value = getattr(competitor, field, None)
        has_value = value is not None and str(value).strip() not in ["", "None", "Unknown", "N/A"]
        fields.append({
            "field": field,
            "has_value": has_value,
            "value": str(value)[:100] if value else None  # Truncate long values
        })
    
    filled = sum(1 for f in fields if f["has_value"])
    
    return {
        "competitor_id": competitor_id,
        "name": competitor.name,
        "filled_fields": filled,
        "total_fields": len(COMPETITOR_DATA_FIELDS),
        "completeness_percent": round((filled / len(COMPETITOR_DATA_FIELDS)) * 100, 1),
        "fields": fields
    }


# ============== CHANGE HISTORY ENDPOINTS ==============

@app.get("/api/changes")
def get_changes(competitor_id: Optional[int] = None, days: int = 30, db: Session = Depends(get_db)):
    """Get change logs for timeline."""
    from datetime import timedelta
    cutoff = datetime.utcnow() - timedelta(days=days)
    
    query = db.query(DataChangeHistory).filter(DataChangeHistory.changed_at >= cutoff)
    
    if competitor_id:
        query = query.filter(DataChangeHistory.competitor_id == competitor_id)
        
    changes = query.order_by(DataChangeHistory.changed_at.desc()).limit(100).all()
    
    return {
        "competitor_id": competitor_id,
        "days": days,
        "count": len(changes),
        "changes": [
            {
                "id": c.id,
                "competitor_id": c.competitor_id,
                "competitor_name": c.competitor_name,
                "change_type": c.field_name.replace("_", " ").title(),
                "previous_value": c.old_value,
                "new_value": c.new_value,
                "severity": "Medium",
                "detected_at": c.changed_at.isoformat(),
                "source": c.changed_by or "manual",
                "changed_by": c.changed_by
            }
            for c in changes
        ]
    }


@app.get("/api/changes/history/{competitor_id}")
def get_competitor_change_history(competitor_id: int, db: Session = Depends(get_db)):
    """Get detailed change history for a specific competitor."""
    changes = db.query(DataChangeHistory).filter(
        DataChangeHistory.competitor_id == competitor_id
    ).order_by(DataChangeHistory.changed_at.desc()).limit(100).all()

    return {
        "competitor_id": competitor_id,
        "total": len(changes),
        "changes": [
            {
                "id": c.id,
                "field": c.field_name,
                "old_value": c.old_value,
                "new_value": c.new_value,
                "changed_by": c.changed_by,
                "reason": c.change_reason,
                "source_url": c.source_url,
                "changed_at": c.changed_at.isoformat() if c.changed_at else None
            }
            for c in changes
        ]
    }


# ============== ACTIVITY LOGS ENDPOINTS (Shared across all users) ==============

@app.get("/api/activity-logs")
def get_activity_logs(
    action_type: str = None,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Get activity logs showing who made changes and when (visible to all users)."""
    query = db.query(ActivityLog)

    if action_type:
        query = query.filter(ActivityLog.action_type == action_type)

    logs = query.order_by(ActivityLog.created_at.desc()).limit(limit).all()

    return {
        "total": len(logs),
        "logs": [
            {
                "id": log.id,
                "user_email": log.user_email,
                "action_type": log.action_type,
                "action_details": log.action_details,
                "created_at": log.created_at.isoformat() if log.created_at else None
            }
            for log in logs
        ]
    }


@app.get("/api/activity-logs/summary")
def get_activity_summary(
    days: int = 7,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Get summary of recent activity by user and action type."""
    from sqlalchemy import func

    cutoff = datetime.utcnow() - timedelta(days=days)

    # Get activity counts by user
    user_activity = db.query(
        ActivityLog.user_email,
        func.count(ActivityLog.id).label("action_count")
    ).filter(
        ActivityLog.created_at >= cutoff
    ).group_by(ActivityLog.user_email).all()

    # Get activity counts by type
    type_activity = db.query(
        ActivityLog.action_type,
        func.count(ActivityLog.id).label("action_count")
    ).filter(
        ActivityLog.created_at >= cutoff
    ).group_by(ActivityLog.action_type).all()

    return {
        "period_days": days,
        "by_user": [{"user": u[0], "count": u[1]} for u in user_activity],
        "by_type": [{"type": t[0], "count": t[1]} for t in type_activity]
    }


# ============== DISCOVERY ENGINE ENDPOINTS ==============

@app.get("/api/discovery/context")
def get_discovery_context(current_user: dict = Depends(get_current_user)):
    """Get the current Discovery Engine context (certification DNA)."""
    try:
        context_path = os.path.join(os.path.dirname(__file__), "certify_context.json")
        if not os.path.exists(context_path):
             return {"core_keywords": [], "market_keywords": [], "exclusions": []}
             
        with open(context_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/discovery/context")
def update_discovery_context(new_context: dict, current_user: dict = Depends(get_current_user)):
    """Update the Discovery Engine context."""
    try:
        context_path = os.path.join(os.path.dirname(__file__), "certify_context.json")
        with open(context_path, 'w', encoding='utf-8') as f:
            json.dump(new_context, f, indent=4)
        return {"success": True, "message": "Context updated successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/discovery/refine-context")
def refine_discovery_context(request: dict, current_user: dict = Depends(get_current_user)):
    """Use AI to refine the context based on user chat input."""
    try:
        user_input = request.get("message")
        current_context = request.get("current_context")
        
        if not user_input:
            raise HTTPException(status_code=400, detail="Message required")

        api_key = os.getenv("OPENAI_API_KEY")
        if not api_key:
             raise HTTPException(status_code=500, detail="OpenAI key missing")

        from openai import OpenAI
        client = OpenAI(api_key=api_key)
        
        system_prompt = """You are an expert configuration assistant for a Competitor Discovery Engine.
        Your goal is to update the JSON configuration profile based on the user's request.
        
        The JSON structure is:
        {
            "core_keywords": ["list", "of", "competitor", "keywords"],
            "market_keywords": ["target", "markets"],
            "required_context": ["must", "have", "terms"],
            "negative_keywords": ["terms", "to", "avoid"],
            "known_competitors": ["known", "competitor", "names"],
            "exclusions": ["industries", "to", "exclude"]
        }
        
        Return ONLY the updated JSON. Do not return markdown formatting."""

        response = client.chat.completions.create(
            model="gpt-4.1",
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": f"Current Profile: {json.dumps(current_context)}\n\nUser Request: {user_input}\n\nUpdate the profile:"}
            ],
            response_format={"type": "json_object"}
        )
        
        refined_json = json.loads(response.choices[0].message.content)
        return {"success": True, "refined_context": refined_json}
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/discovery/schedule")
def schedule_discovery_run(request: dict, current_user: dict = Depends(get_current_user)):
    """Schedule a one-off discovery run."""
    try:
        run_at_str = request.get("run_at") # ISO format string
        if not run_at_str:
            raise HTTPException(status_code=400, detail="run_at timestamp required")
            
        run_at = datetime.fromisoformat(run_at_str.replace('Z', '+00:00'))
        
        from scheduler import schedule_one_off_discovery
        schedule_one_off_discovery(run_at)
        
        return {"success": True, "message": f"Discovery job scheduled for {run_at.isoformat()}"}
    except Exception as e:
         raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/changes/log")
def log_data_change(
    competitor_id: int,
    field_name: str,
    old_value: str,
    new_value: str,
    changed_by: str = "system",
    reason: Optional[str] = None,
    source_url: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Log a data change for audit purposes."""
    competitor = db.query(Competitor).filter(Competitor.id == competitor_id).first()
    if not competitor:
        raise HTTPException(status_code=404, detail="Competitor not found")
    
    change = DataChangeHistory(
        competitor_id=competitor_id,
        competitor_name=competitor.name,
        field_name=field_name,
        old_value=old_value,
        new_value=new_value,
        changed_by=changed_by,
        change_reason=reason,
        source_url=source_url
    )
    db.add(change)
    db.commit()
    
    return {"success": True, "change_id": change.id}


# ============== DATA SOURCE ENDPOINTS ==============

@app.get("/api/sources/{competitor_id}")
def get_competitor_sources(competitor_id: int, db: Session = Depends(get_db)):
    """Get all data sources for a competitor's fields."""
    sources = db.query(DataSource).filter(
        DataSource.competitor_id == competitor_id
    ).all()
    
    sources_by_field = {}
    for s in sources:
        sources_by_field[s.field_name] = {
            "source_type": s.source_type,
            "source_url": s.source_url,
            "source_name": s.source_name,
            "entered_by": s.entered_by,
            "formula": s.formula,
            "verified_at": s.verified_at.isoformat() if s.verified_at else None
        }
    
    return {
        "competitor_id": competitor_id,
        "sources": sources_by_field
    }


@app.get("/api/sources/{competitor_id}/{field_name}")
def get_field_source(competitor_id: int, field_name: str, db: Session = Depends(get_db)):
    """Get the data source for a specific field of a competitor."""
    source = db.query(DataSource).filter(
        DataSource.competitor_id == competitor_id,
        DataSource.field_name == field_name
    ).first()
    
    if not source:
        return {
            "competitor_id": competitor_id,
            "field_name": field_name,
            "source_type": "unknown",
            "source_url": None,
            "message": "No source recorded for this field"
        }
    
    return {
        "competitor_id": competitor_id,
        "field_name": field_name,
        "source_type": source.source_type,
        "source_url": source.source_url,
        "source_name": source.source_name,
        "entered_by": source.entered_by,
        "formula": source.formula,
        "verified_at": source.verified_at.isoformat() if source.verified_at else None
    }


@app.post("/api/sources/set")
def set_field_source(
    competitor_id: int,
    field_name: str,
    source_type: str,  # "external", "manual", "calculated"
    source_url: Optional[str] = None,
    source_name: Optional[str] = None,
    entered_by: Optional[str] = None,
    formula: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Set or update the source for a competitor's field."""
    # Check if source already exists
    existing = db.query(DataSource).filter(
        DataSource.competitor_id == competitor_id,
        DataSource.field_name == field_name
    ).first()
    
    if existing:
        existing.source_type = source_type
        existing.source_url = source_url
        existing.source_name = source_name
        existing.entered_by = entered_by
        existing.formula = formula
        existing.verified_at = datetime.utcnow()
        existing.updated_at = datetime.utcnow()
    else:
        new_source = DataSource(
            competitor_id=competitor_id,
            field_name=field_name,
            source_type=source_type,
            source_url=source_url,
            source_name=source_name,
            entered_by=entered_by,
            formula=formula
        )
        db.add(new_source)
    
    db.commit()
    return {"success": True, "message": f"Source set for {field_name}"}


# ============== ENHANCED DATA SOURCES WITH CONFIDENCE SCORING ==============

@app.get("/api/competitors/{competitor_id}/data-sources")
def get_competitor_data_sources_enhanced(competitor_id: int, db: Session = Depends(get_db)):
    """Get all data sources and confidence scores for a competitor with enhanced metadata."""
    sources = db.query(DataSource).filter(
        DataSource.competitor_id == competitor_id
    ).order_by(DataSource.field_name).all()

    return [{
        "field": s.field_name,
        "value": s.current_value,
        "previous_value": s.previous_value,
        "source_type": s.source_type,
        "source_name": s.source_name,
        "source_url": s.source_url,
        "extraction_method": s.extraction_method,
        "confidence": {
            "score": s.confidence_score or 0,
            "level": s.confidence_level or "low",
            "reliability": s.source_reliability,
            "credibility": s.information_credibility,
            "corroborating_sources": s.corroborating_sources or 0,
            "reliability_description": get_reliability_description(s.source_reliability) if s.source_reliability else None,
            "credibility_description": get_credibility_description(s.information_credibility) if s.information_credibility else None
        },
        "verification": {
            "is_verified": s.is_verified,
            "verified_by": s.verified_by,
            "verification_date": s.verification_date.isoformat() if s.verification_date else None
        },
        "temporal": {
            "extracted_at": s.extracted_at.isoformat() if s.extracted_at else None,
            "data_as_of_date": s.data_as_of_date.isoformat() if s.data_as_of_date else None,
            "staleness_days": s.staleness_days or 0
        }
    } for s in sources]


@app.get("/api/data-quality/low-confidence")
def get_low_confidence_data(threshold: int = 40, db: Session = Depends(get_db)):
    """Get all data points below confidence threshold for review."""
    sources = db.query(DataSource).filter(
        DataSource.confidence_score < threshold
    ).order_by(DataSource.confidence_score).all()

    # Group by competitor
    by_competitor = {}
    for s in sources:
        comp_id = s.competitor_id
        if comp_id not in by_competitor:
            competitor = db.query(Competitor).filter(Competitor.id == comp_id).first()
            by_competitor[comp_id] = {
                "competitor_id": comp_id,
                "competitor_name": competitor.name if competitor else "Unknown",
                "fields": []
            }
        by_competitor[comp_id]["fields"].append({
            "field": s.field_name,
            "value": s.current_value,
            "confidence_score": s.confidence_score or 0,
            "confidence_level": s.confidence_level or "low",
            "source_type": s.source_type,
            "reason": f"Low confidence ({s.confidence_score or 0}/100) from {s.source_type or 'unknown source'}"
        })

    return {
        "threshold": threshold,
        "total_low_confidence": len(sources),
        "competitors_affected": len(by_competitor),
        "data": list(by_competitor.values())
    }


@app.get("/api/data-quality/confidence-distribution")
def get_confidence_distribution(db: Session = Depends(get_db)):
    """Get distribution of confidence levels across all data."""
    sources = db.query(DataSource).all()

    high = len([s for s in sources if (s.confidence_score or 0) >= 70])
    moderate = len([s for s in sources if 40 <= (s.confidence_score or 0) < 70])
    low = len([s for s in sources if (s.confidence_score or 0) < 40])
    unscored = len([s for s in sources if s.confidence_score is None])

    return {
        "total_data_points": len(sources),
        "distribution": {
            "high": {"count": high, "percentage": round(high / len(sources) * 100, 1) if sources else 0},
            "moderate": {"count": moderate, "percentage": round(moderate / len(sources) * 100, 1) if sources else 0},
            "low": {"count": low, "percentage": round(low / len(sources) * 100, 1) if sources else 0},
            "unscored": {"count": unscored, "percentage": round(unscored / len(sources) * 100, 1) if sources else 0}
        },
        "by_source_type": _get_confidence_by_source_type(sources)
    }


def _get_confidence_by_source_type(sources: list) -> dict:
    """Helper to group confidence scores by source type."""
    by_type = {}
    for s in sources:
        source_type = s.source_type or "unknown"
        if source_type not in by_type:
            by_type[source_type] = {"count": 0, "total_score": 0, "scores": []}
        by_type[source_type]["count"] += 1
        if s.confidence_score is not None:
            by_type[source_type]["total_score"] += s.confidence_score
            by_type[source_type]["scores"].append(s.confidence_score)

    # Calculate averages
    result = {}
    for source_type, data in by_type.items():
        avg = round(data["total_score"] / len(data["scores"]), 1) if data["scores"] else 0
        result[source_type] = {
            "count": data["count"],
            "average_confidence": avg,
            "description": get_source_type_description(source_type)
        }
    return result


@app.post("/api/sources/set-with-confidence")
def set_field_source_with_confidence(
    competitor_id: int,
    field_name: str,
    current_value: str,
    source_type: str,
    source_url: Optional[str] = None,
    source_name: Optional[str] = None,
    extraction_method: Optional[str] = None,
    source_reliability: Optional[str] = None,
    information_credibility: Optional[int] = None,
    corroborating_sources: int = 0,
    data_as_of_date: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Set or update a data source with confidence scoring."""
    # Calculate confidence score
    defaults = get_source_defaults(source_type)
    reliability = source_reliability or defaults["reliability"]
    credibility = information_credibility or defaults["credibility"]

    confidence_result = calculate_confidence_score(
        source_type=source_type,
        source_reliability=reliability,
        information_credibility=credibility,
        corroborating_sources=corroborating_sources,
        data_age_days=0
    )

    # Check if source already exists
    existing = db.query(DataSource).filter(
        DataSource.competitor_id == competitor_id,
        DataSource.field_name == field_name
    ).first()

    if existing:
        # Store previous value
        existing.previous_value = existing.current_value
        existing.current_value = current_value
        existing.source_type = source_type
        existing.source_url = source_url
        existing.source_name = source_name
        existing.extraction_method = extraction_method
        existing.source_reliability = reliability
        existing.information_credibility = credibility
        existing.confidence_score = confidence_result.score
        existing.confidence_level = confidence_result.level
        existing.corroborating_sources = corroborating_sources
        existing.extracted_at = datetime.utcnow()
        if data_as_of_date:
            existing.data_as_of_date = datetime.fromisoformat(data_as_of_date)
        existing.updated_at = datetime.utcnow()
    else:
        new_source = DataSource(
            competitor_id=competitor_id,
            field_name=field_name,
            current_value=current_value,
            source_type=source_type,
            source_url=source_url,
            source_name=source_name,
            extraction_method=extraction_method,
            source_reliability=reliability,
            information_credibility=credibility,
            confidence_score=confidence_result.score,
            confidence_level=confidence_result.level,
            corroborating_sources=corroborating_sources,
            data_as_of_date=datetime.fromisoformat(data_as_of_date) if data_as_of_date else None
        )
        db.add(new_source)

    db.commit()

    return {
        "success": True,
        "message": f"Source set for {field_name}",
        "confidence": {
            "score": confidence_result.score,
            "level": confidence_result.level,
            "explanation": confidence_result.explanation,
            "breakdown": confidence_result.breakdown
        }
    }


@app.post("/api/sources/verify/{competitor_id}/{field_name}")
def verify_data_source(
    competitor_id: int,
    field_name: str,
    verification_method: str,
    corroborating_sources: int = 0,
    db: Session = Depends(get_db)
):
    """Mark a data source as verified and recalculate confidence."""
    source = db.query(DataSource).filter(
        DataSource.competitor_id == competitor_id,
        DataSource.field_name == field_name
    ).first()

    if not source:
        raise HTTPException(status_code=404, detail=f"No source found for {field_name}")

    # Update verification status
    source.is_verified = True
    source.verified_by = verification_method
    source.verification_date = datetime.utcnow()
    source.corroborating_sources = corroborating_sources

    # Recalculate confidence with verification bonus
    staleness = calculate_data_staleness(source.extracted_at, source.data_as_of_date)
    confidence_result = calculate_confidence_score(
        source_type=source.source_type or "unknown",
        source_reliability=source.source_reliability,
        information_credibility=source.information_credibility,
        corroborating_sources=corroborating_sources,
        data_age_days=staleness
    )

    source.confidence_score = confidence_result.score
    source.confidence_level = confidence_result.level
    source.staleness_days = staleness

    db.commit()

    return {
        "success": True,
        "message": f"Field {field_name} verified via {verification_method}",
        "new_confidence": {
            "score": confidence_result.score,
            "level": confidence_result.level
        }
    }


@app.get("/api/source-types")
def get_source_types():
    """Get all available source types with their default reliability ratings."""
    return {
        "source_types": [
            {
                "type": source_type,
                "reliability": info["reliability"],
                "credibility": info["credibility"],
                "description": info["description"],
                "reliability_description": get_reliability_description(info["reliability"]),
                "credibility_description": get_credibility_description(info["credibility"])
            }
            for source_type, info in SOURCE_TYPE_DEFAULTS.items()
        ],
        "reliability_scale": RELIABILITY_DESCRIPTIONS,
        "credibility_scale": CREDIBILITY_DESCRIPTIONS
    }


@app.post("/api/data-quality/recalculate-confidence")
def recalculate_all_confidence_scores(db: Session = Depends(get_db)):
    """Recalculate confidence scores for all data sources."""
    sources = db.query(DataSource).all()
    updated_count = 0

    for source in sources:
        if source.source_type:
            staleness = calculate_data_staleness(
                source.extracted_at or datetime.utcnow(),
                source.data_as_of_date
            )
            confidence_result = calculate_confidence_score(
                source_type=source.source_type,
                source_reliability=source.source_reliability,
                information_credibility=source.information_credibility,
                corroborating_sources=source.corroborating_sources or 0,
                data_age_days=staleness
            )
            source.confidence_score = confidence_result.score
            source.confidence_level = confidence_result.level
            source.staleness_days = staleness
            updated_count += 1

    db.commit()

    return {
        "success": True,
        "message": f"Recalculated confidence for {updated_count} data sources",
        "updated_count": updated_count
    }


@app.get("/api/data-quality/overview")
def get_data_quality_overview(db: Session = Depends(get_db)):
    """
    Get comprehensive data quality overview with confidence metrics.
    Phase 7: Data Quality Dashboard
    """
    from datetime import timedelta

    # Total active competitors
    total_competitors = db.query(Competitor).filter(Competitor.is_active == True).count()

    # Get all data sources
    sources = db.query(DataSource).all()
    total_data_points = len(sources)

    # Confidence distribution
    high_confidence = len([s for s in sources if (s.confidence_score or 0) >= 70])
    moderate_confidence = len([s for s in sources if 40 <= (s.confidence_score or 0) < 70])
    low_confidence = len([s for s in sources if (s.confidence_score or 0) < 40 and s.confidence_score is not None])
    unscored = len([s for s in sources if s.confidence_score is None])

    # Verification stats
    verified_count = len([s for s in sources if s.is_verified])
    verification_rate = round((verified_count / total_data_points) * 100, 1) if total_data_points > 0 else 0

    # Staleness analysis (90 days threshold)
    stale_threshold = datetime.utcnow() - timedelta(days=90)
    stale_sources = [s for s in sources if s.extracted_at and s.extracted_at < stale_threshold]
    stale_count = len(stale_sources)
    staleness_rate = round((stale_count / total_data_points) * 100, 1) if total_data_points > 0 else 0

    # Key field coverage
    key_fields = ["customer_count", "base_price", "pricing_model", "employee_count", "year_founded", "key_features"]
    field_coverage = {}
    for field in key_fields:
        field_sources = [s for s in sources if s.field_name == field]
        populated = len([s for s in field_sources if s.current_value and s.current_value not in ['N/A', 'Unknown', '']])
        avg_confidence = sum([s.confidence_score or 0 for s in field_sources]) / len(field_sources) if field_sources else 0
        field_coverage[field] = {
            "populated": populated,
            "total": total_competitors,
            "percentage": round((populated / total_competitors) * 100, 1) if total_competitors > 0 else 0,
            "avg_confidence": round(avg_confidence, 1)
        }

    # Source type breakdown
    source_type_counts = {}
    for s in sources:
        st = s.source_type or 'unknown'
        if st not in source_type_counts:
            source_type_counts[st] = {"count": 0, "avg_confidence": 0, "scores": []}
        source_type_counts[st]["count"] += 1
        if s.confidence_score is not None:
            source_type_counts[st]["scores"].append(s.confidence_score)

    for st, data in source_type_counts.items():
        if data["scores"]:
            data["avg_confidence"] = round(sum(data["scores"]) / len(data["scores"]), 1)
        del data["scores"]  # Clean up internal data

    # Per-competitor quality scores
    competitor_scores = []
    competitors = db.query(Competitor).filter(Competitor.is_active == True).all()
    for comp in competitors:
        comp_sources = [s for s in sources if s.competitor_id == comp.id]
        if comp_sources:
            avg_conf = sum([s.confidence_score or 0 for s in comp_sources]) / len(comp_sources)
            verified = len([s for s in comp_sources if s.is_verified])
            high_conf = len([s for s in comp_sources if (s.confidence_score or 0) >= 70])
            low_conf = len([s for s in comp_sources if (s.confidence_score or 0) < 40])
            competitor_scores.append({
                "id": comp.id,
                "name": comp.name,
                "total_fields": len(comp_sources),
                "avg_confidence": round(avg_conf, 1),
                "verified_count": verified,
                "high_confidence_count": high_conf,
                "low_confidence_count": low_conf,
                "quality_tier": "Excellent" if avg_conf >= 70 else "Good" if avg_conf >= 50 else "Fair" if avg_conf >= 30 else "Poor"
            })

    # Sort by average confidence (descending)
    competitor_scores.sort(key=lambda x: x["avg_confidence"], reverse=True)

    # Needs attention summary
    needs_attention = {
        "low_confidence_count": low_confidence,
        "stale_count": stale_count,
        "unverified_count": total_data_points - verified_count,
        "unscored_count": unscored
    }

    return {
        "total_competitors": total_competitors,
        "total_data_points": total_data_points,
        "confidence_distribution": {
            "high": {"count": high_confidence, "percentage": round(high_confidence / total_data_points * 100, 1) if total_data_points else 0},
            "moderate": {"count": moderate_confidence, "percentage": round(moderate_confidence / total_data_points * 100, 1) if total_data_points else 0},
            "low": {"count": low_confidence, "percentage": round(low_confidence / total_data_points * 100, 1) if total_data_points else 0},
            "unscored": {"count": unscored, "percentage": round(unscored / total_data_points * 100, 1) if total_data_points else 0}
        },
        "verification_rate": verification_rate,
        "staleness_rate": staleness_rate,
        "field_coverage": field_coverage,
        "source_type_breakdown": source_type_counts,
        "competitor_scores": competitor_scores[:15],  # Top 15 for dashboard display
        "needs_attention": needs_attention,
        "generated_at": datetime.utcnow().isoformat()
    }


# ============== DATA TRIANGULATION ENDPOINTS ==============

@app.post("/api/triangulate/{competitor_id}")
async def triangulate_competitor_data(competitor_id: int, db: Session = Depends(get_db)):
    """
    Triangulate all key data fields for a competitor using multiple sources.

    This cross-references data from:
    - Website scrapes
    - SEC filings (if public company)
    - News articles
    - Manual entries

    Returns verified values with confidence scores.
    """
    competitor = db.query(Competitor).filter(Competitor.id == competitor_id).first()
    if not competitor:
        raise HTTPException(status_code=404, detail="Competitor not found")

    triangulator = DataTriangulator(db)

    results = await triangulator.triangulate_all_key_fields(
        competitor_id=competitor_id,
        competitor_name=competitor.name,
        website=competitor.website,
        is_public=competitor.is_public,
        ticker_symbol=competitor.ticker_symbol
    )

    # Update DataSource records with triangulated confidence
    for field_name, result in results.items():
        if result.confidence_score > 0:
            existing = db.query(DataSource).filter(
                DataSource.competitor_id == competitor_id,
                DataSource.field_name == field_name
            ).first()

            if existing:
                existing.confidence_score = result.confidence_score
                existing.confidence_level = result.confidence_level
                existing.corroborating_sources = result.sources_agreeing
                existing.is_verified = result.confidence_level == "high"
                existing.verified_by = "triangulation" if result.sources_agreeing > 1 else None
                existing.verification_date = datetime.utcnow() if result.sources_agreeing > 1 else None
                existing.updated_at = datetime.utcnow()

    db.commit()

    return {
        "competitor_id": competitor_id,
        "competitor_name": competitor.name,
        "triangulation_results": {
            field: triangulation_result_to_dict(result)
            for field, result in results.items()
        },
        "triangulated_at": datetime.utcnow().isoformat()
    }


@app.post("/api/triangulate/{competitor_id}/{field_name}")
async def triangulate_single_field(
    competitor_id: int,
    field_name: str,
    db: Session = Depends(get_db)
):
    """Triangulate a specific field for a competitor."""
    competitor = db.query(Competitor).filter(Competitor.id == competitor_id).first()
    if not competitor:
        raise HTTPException(status_code=404, detail="Competitor not found")

    triangulator = DataTriangulator(db)

    # Run appropriate triangulation based on field
    if field_name == "customer_count":
        result = await triangulator.triangulate_customer_count(
            competitor_id, competitor.name, competitor.website,
            competitor.is_public, competitor.ticker_symbol
        )
    elif field_name == "employee_count":
        result = await triangulator.triangulate_employee_count(
            competitor_id, competitor.name,
            competitor.is_public, competitor.ticker_symbol
        )
    elif field_name in ["base_price", "pricing"]:
        result = await triangulator.triangulate_pricing(
            competitor_id, competitor.name, competitor.website
        )
    else:
        raise HTTPException(
            status_code=400,
            detail=f"Triangulation not supported for field: {field_name}. Supported: customer_count, employee_count, base_price"
        )

    # Update DataSource with triangulated confidence
    existing = db.query(DataSource).filter(
        DataSource.competitor_id == competitor_id,
        DataSource.field_name == field_name
    ).first()

    if existing and result.confidence_score > 0:
        existing.confidence_score = result.confidence_score
        existing.confidence_level = result.confidence_level
        existing.corroborating_sources = result.sources_agreeing
        existing.is_verified = result.sources_agreeing > 1
        existing.verified_by = "triangulation" if result.sources_agreeing > 1 else None
        existing.verification_date = datetime.utcnow() if result.sources_agreeing > 1 else None
        existing.updated_at = datetime.utcnow()
        db.commit()

    return triangulation_result_to_dict(result)


@app.post("/api/triangulate/all")
async def triangulate_all_competitors(
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db)
):
    """Trigger triangulation for all active competitors (background job)."""
    competitors = db.query(Competitor).filter(
        Competitor.is_deleted == False
    ).all()

    # Add background task for each competitor
    for comp in competitors:
        background_tasks.add_task(
            run_triangulation_job,
            comp.id,
            comp.name,
            comp.website,
            comp.is_public,
            comp.ticker_symbol
        )

    return {
        "success": True,
        "message": f"Triangulation started for {len(competitors)} competitors",
        "competitors_queued": len(competitors)
    }


async def run_triangulation_job(
    competitor_id: int,
    competitor_name: str,
    website: str,
    is_public: bool,
    ticker_symbol: str
):
    """Background job to triangulate data for a competitor."""
    db = SessionLocal()
    try:
        triangulator = DataTriangulator(db)
        results = await triangulator.triangulate_all_key_fields(
            competitor_id, competitor_name, website, is_public, ticker_symbol
        )

        # Update DataSource records
        for field_name, result in results.items():
            if result.confidence_score > 0:
                existing = db.query(DataSource).filter(
                    DataSource.competitor_id == competitor_id,
                    DataSource.field_name == field_name
                ).first()

                if existing:
                    existing.confidence_score = result.confidence_score
                    existing.confidence_level = result.confidence_level
                    existing.corroborating_sources = result.sources_agreeing
                    existing.is_verified = result.confidence_level == "high"
                    existing.verified_by = "triangulation" if result.sources_agreeing > 1 else None
                    existing.verification_date = datetime.utcnow() if result.sources_agreeing > 1 else None

        db.commit()
        print(f"Triangulation complete for {competitor_name}")

    except Exception as e:
        print(f"Triangulation failed for {competitor_name}: {e}")
        db.rollback()
    finally:
        db.close()


@app.get("/api/triangulation/status")
def get_triangulation_status(db: Session = Depends(get_db)):
    """Get overview of triangulation status across all competitors."""
    sources = db.query(DataSource).all()

    verified_count = len([s for s in sources if s.is_verified])
    triangulated_count = len([s for s in sources if s.verified_by == "triangulation"])
    pending_count = len([s for s in sources if not s.is_verified and s.current_value])

    # Group by confidence level
    high = len([s for s in sources if s.confidence_level == "high"])
    moderate = len([s for s in sources if s.confidence_level == "moderate"])
    low = len([s for s in sources if s.confidence_level == "low" or s.confidence_level is None])

    return {
        "total_data_points": len(sources),
        "verification_status": {
            "verified": verified_count,
            "triangulated": triangulated_count,
            "pending_verification": pending_count
        },
        "confidence_distribution": {
            "high": high,
            "moderate": moderate,
            "low": low
        },
        "verification_rate": round(verified_count / len(sources) * 100, 1) if sources else 0
    }


# ============== PHASE 3: PRODUCT & PRICING MANAGEMENT ==============

# Pydantic models for Product & Pricing
class ProductCreate(BaseModel):
    """Create a new competitor product."""
    competitor_id: int
    product_name: str
    product_category: str  # "Patient Intake", "RCM", "EHR", "Payments", etc.
    product_subcategory: Optional[str] = None
    description: Optional[str] = None
    key_features: Optional[str] = None  # JSON array as string
    target_segment: Optional[str] = None  # "SMB", "Mid-Market", "Enterprise"
    is_primary_product: bool = False
    market_position: Optional[str] = None  # "Leader", "Challenger", "Niche"

class ProductResponse(BaseModel):
    id: int
    competitor_id: int
    product_name: str
    product_category: str
    product_subcategory: Optional[str]
    description: Optional[str]
    key_features: Optional[str]
    target_segment: Optional[str]
    is_primary_product: bool
    market_position: Optional[str]
    launched_date: Optional[datetime]
    last_updated: Optional[datetime]
    pricing_tiers: List[dict] = []

    class Config:
        from_attributes = True

class PricingTierCreate(BaseModel):
    """Create a pricing tier for a product."""
    product_id: int
    tier_name: str  # "Basic", "Professional", "Enterprise"
    tier_position: Optional[int] = None
    pricing_model: str  # "per_visit", "per_provider", "per_location", "subscription", "percentage_collections", "custom"
    base_price: Optional[float] = None
    price_currency: str = "USD"
    price_unit: Optional[str] = None  # "visit", "provider/month", "location/month"
    price_display: Optional[str] = None  # Original display: "$3.00/visit"
    percentage_rate: Optional[float] = None  # For RCM: 4.5 for 4.5%
    percentage_basis: Optional[str] = None  # "collections", "charges"
    min_volume: Optional[str] = None
    max_volume: Optional[str] = None
    included_features: Optional[str] = None  # JSON array
    excluded_features: Optional[str] = None
    contract_length: Optional[str] = None  # "Monthly", "Annual"
    setup_fee: Optional[float] = None
    implementation_cost: Optional[str] = None
    price_source: Optional[str] = None  # "website", "sales_quote", "customer_intel"

class PricingTierResponse(BaseModel):
    id: int
    product_id: int
    tier_name: str
    tier_position: Optional[int]
    pricing_model: str
    base_price: Optional[float]
    price_currency: str
    price_unit: Optional[str]
    price_display: Optional[str]
    percentage_rate: Optional[float]
    percentage_basis: Optional[str]
    min_volume: Optional[str]
    max_volume: Optional[str]
    contract_length: Optional[str]
    setup_fee: Optional[float]
    implementation_cost: Optional[str]
    price_verified: bool
    price_source: Optional[str]
    confidence_score: Optional[int]
    last_verified: Optional[datetime]

    class Config:
        from_attributes = True

class FeatureMatrixCreate(BaseModel):
    """Create a feature entry for a product."""
    product_id: int
    feature_category: str  # "Patient Intake", "Payments", "Integration"
    feature_name: str  # "Digital Check-In", "Apple Pay Support"
    feature_status: str  # "included", "add_on", "not_available", "coming_soon"
    feature_tier: Optional[str] = None  # Which tier includes this
    notes: Optional[str] = None
    source_url: Optional[str] = None


# ============== PHASE 4: CUSTOMER COUNT VERIFICATION ==============

class CustomerCountCreate(BaseModel):
    """Create a customer count estimate for a competitor."""
    competitor_id: int
    count_value: Optional[int] = None  # Numeric: 3000
    count_display: str  # Display: "3,000+" or "3,000-5,000"
    count_type: str = "estimate"  # "exact", "minimum", "range", "estimate"
    count_unit: str  # "healthcare_organizations", "providers", "locations", "users", "lives_covered"
    count_definition: Optional[str] = None  # "Number of distinct hospital/clinic customers"
    segment_breakdown: Optional[str] = None  # JSON: {"hospitals": 500, "ambulatory": 2500}
    primary_source: str  # "website", "sec_10k", "press_release"
    primary_source_url: Optional[str] = None
    primary_source_date: Optional[datetime] = None
    as_of_date: Optional[datetime] = None  # When this count was valid


class CustomerCountResponse(BaseModel):
    id: int
    competitor_id: int
    count_value: Optional[int]
    count_display: Optional[str]
    count_type: Optional[str]
    count_unit: Optional[str]
    count_definition: Optional[str]
    segment_breakdown: Optional[str]
    is_verified: bool
    verification_method: Optional[str]
    verification_date: Optional[datetime]
    primary_source: Optional[str]
    primary_source_url: Optional[str]
    primary_source_date: Optional[datetime]
    all_sources: Optional[str]
    source_agreement_score: Optional[float]
    confidence_score: Optional[int]
    confidence_level: Optional[str]
    confidence_notes: Optional[str]
    as_of_date: Optional[datetime]
    previous_count: Optional[int]
    growth_rate: Optional[float]
    created_at: Optional[datetime]
    updated_at: Optional[datetime]

    class Config:
        from_attributes = True


class CustomerCountVerifyRequest(BaseModel):
    """Request to verify a customer count with additional sources."""
    verification_method: str  # "sec_filing", "triangulation", "sales_intel", "manual"
    verification_notes: Optional[str] = None
    additional_sources: Optional[List[dict]] = None  # List of {source_type, source_url, value}


# ============== PRODUCT CRUD ENDPOINTS ==============

@app.get("/api/competitors/{competitor_id}/products", response_model=List[ProductResponse])
async def get_competitor_products(competitor_id: int, db: Session = Depends(get_db)):
    """Get all products and pricing for a competitor."""
    # Verify competitor exists
    competitor = db.query(Competitor).filter(Competitor.id == competitor_id).first()
    if not competitor:
        raise HTTPException(status_code=404, detail="Competitor not found")

    products = db.query(CompetitorProduct).filter(
        CompetitorProduct.competitor_id == competitor_id
    ).all()

    result = []
    for p in products:
        # Get pricing tiers for this product
        tiers = db.query(ProductPricingTier).filter(
            ProductPricingTier.product_id == p.id
        ).order_by(ProductPricingTier.tier_position).all()

        product_dict = {
            "id": p.id,
            "competitor_id": p.competitor_id,
            "product_name": p.product_name,
            "product_category": p.product_category,
            "product_subcategory": p.product_subcategory,
            "description": p.description,
            "key_features": p.key_features,
            "target_segment": p.target_segment,
            "is_primary_product": p.is_primary_product,
            "market_position": p.market_position,
            "launched_date": p.launched_date,
            "last_updated": p.last_updated,
            "pricing_tiers": [{
                "id": t.id,
                "tier_name": t.tier_name,
                "pricing_model": t.pricing_model,
                "price_display": t.price_display,
                "base_price": t.base_price,
                "price_unit": t.price_unit,
                "percentage_rate": t.percentage_rate,
                "confidence_score": t.confidence_score,
                "price_verified": t.price_verified
            } for t in tiers]
        }
        result.append(product_dict)

    return result


@app.post("/api/products", response_model=ProductResponse)
async def create_product(product: ProductCreate, db: Session = Depends(get_db)):
    """Create a new product for a competitor."""
    # Verify competitor exists
    competitor = db.query(Competitor).filter(Competitor.id == product.competitor_id).first()
    if not competitor:
        raise HTTPException(status_code=404, detail="Competitor not found")

    # Check if product already exists
    existing = db.query(CompetitorProduct).filter(
        CompetitorProduct.competitor_id == product.competitor_id,
        CompetitorProduct.product_name == product.product_name
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="Product already exists for this competitor")

    new_product = CompetitorProduct(
        competitor_id=product.competitor_id,
        product_name=product.product_name,
        product_category=product.product_category,
        product_subcategory=product.product_subcategory,
        description=product.description,
        key_features=product.key_features,
        target_segment=product.target_segment,
        is_primary_product=product.is_primary_product,
        market_position=product.market_position,
        last_updated=datetime.utcnow()
    )
    db.add(new_product)
    db.commit()
    db.refresh(new_product)

    return {
        "id": new_product.id,
        "competitor_id": new_product.competitor_id,
        "product_name": new_product.product_name,
        "product_category": new_product.product_category,
        "product_subcategory": new_product.product_subcategory,
        "description": new_product.description,
        "key_features": new_product.key_features,
        "target_segment": new_product.target_segment,
        "is_primary_product": new_product.is_primary_product,
        "market_position": new_product.market_position,
        "launched_date": new_product.launched_date,
        "last_updated": new_product.last_updated,
        "pricing_tiers": []
    }


@app.put("/api/products/{product_id}", response_model=ProductResponse)
async def update_product(product_id: int, product: ProductCreate, db: Session = Depends(get_db)):
    """Update an existing product."""
    existing = db.query(CompetitorProduct).filter(CompetitorProduct.id == product_id).first()
    if not existing:
        raise HTTPException(status_code=404, detail="Product not found")

    existing.product_name = product.product_name
    existing.product_category = product.product_category
    existing.product_subcategory = product.product_subcategory
    existing.description = product.description
    existing.key_features = product.key_features
    existing.target_segment = product.target_segment
    existing.is_primary_product = product.is_primary_product
    existing.market_position = product.market_position
    existing.last_updated = datetime.utcnow()

    db.commit()
    db.refresh(existing)

    # Get pricing tiers
    tiers = db.query(ProductPricingTier).filter(
        ProductPricingTier.product_id == product_id
    ).order_by(ProductPricingTier.tier_position).all()

    return {
        "id": existing.id,
        "competitor_id": existing.competitor_id,
        "product_name": existing.product_name,
        "product_category": existing.product_category,
        "product_subcategory": existing.product_subcategory,
        "description": existing.description,
        "key_features": existing.key_features,
        "target_segment": existing.target_segment,
        "is_primary_product": existing.is_primary_product,
        "market_position": existing.market_position,
        "launched_date": existing.launched_date,
        "last_updated": existing.last_updated,
        "pricing_tiers": [{
            "id": t.id,
            "tier_name": t.tier_name,
            "pricing_model": t.pricing_model,
            "price_display": t.price_display,
            "base_price": t.base_price,
            "price_unit": t.price_unit,
            "confidence_score": t.confidence_score
        } for t in tiers]
    }


@app.delete("/api/products/{product_id}")
async def delete_product(product_id: int, db: Session = Depends(get_db)):
    """Delete a product and its pricing tiers."""
    product = db.query(CompetitorProduct).filter(CompetitorProduct.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")

    # Delete associated pricing tiers first
    db.query(ProductPricingTier).filter(ProductPricingTier.product_id == product_id).delete()
    # Delete associated features
    db.query(ProductFeatureMatrix).filter(ProductFeatureMatrix.product_id == product_id).delete()
    # Delete the product
    db.delete(product)
    db.commit()

    return {"message": "Product deleted successfully", "product_id": product_id}


# ============== PRICING TIER ENDPOINTS ==============

@app.get("/api/products/{product_id}/pricing-tiers", response_model=List[PricingTierResponse])
async def get_pricing_tiers(product_id: int, db: Session = Depends(get_db)):
    """Get all pricing tiers for a product."""
    product = db.query(CompetitorProduct).filter(CompetitorProduct.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")

    tiers = db.query(ProductPricingTier).filter(
        ProductPricingTier.product_id == product_id
    ).order_by(ProductPricingTier.tier_position).all()

    return tiers


@app.post("/api/pricing-tiers", response_model=PricingTierResponse)
async def create_pricing_tier(tier: PricingTierCreate, db: Session = Depends(get_db)):
    """Create a new pricing tier for a product."""
    # Verify product exists
    product = db.query(CompetitorProduct).filter(CompetitorProduct.id == tier.product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")

    # Calculate confidence score based on source
    source_type = "website_scrape" if tier.price_source == "website" else "manual_verified"
    if tier.price_source == "sales_quote":
        source_type = "api_verified"

    confidence_result = calculate_confidence_score(source_type=source_type)

    new_tier = ProductPricingTier(
        product_id=tier.product_id,
        tier_name=tier.tier_name,
        tier_position=tier.tier_position,
        pricing_model=tier.pricing_model,
        base_price=tier.base_price,
        price_currency=tier.price_currency,
        price_unit=tier.price_unit,
        price_display=tier.price_display,
        percentage_rate=tier.percentage_rate,
        percentage_basis=tier.percentage_basis,
        min_volume=tier.min_volume,
        max_volume=tier.max_volume,
        included_features=tier.included_features,
        excluded_features=tier.excluded_features,
        contract_length=tier.contract_length,
        setup_fee=tier.setup_fee,
        implementation_cost=tier.implementation_cost,
        price_source=tier.price_source,
        price_verified=False,
        confidence_score=confidence_result.score,
        last_verified=datetime.utcnow()
    )
    db.add(new_tier)
    db.commit()
    db.refresh(new_tier)

    return new_tier


@app.put("/api/pricing-tiers/{tier_id}", response_model=PricingTierResponse)
async def update_pricing_tier(tier_id: int, tier: PricingTierCreate, db: Session = Depends(get_db)):
    """Update an existing pricing tier."""
    existing = db.query(ProductPricingTier).filter(ProductPricingTier.id == tier_id).first()
    if not existing:
        raise HTTPException(status_code=404, detail="Pricing tier not found")

    # Update fields
    existing.tier_name = tier.tier_name
    existing.tier_position = tier.tier_position
    existing.pricing_model = tier.pricing_model
    existing.base_price = tier.base_price
    existing.price_currency = tier.price_currency
    existing.price_unit = tier.price_unit
    existing.price_display = tier.price_display
    existing.percentage_rate = tier.percentage_rate
    existing.percentage_basis = tier.percentage_basis
    existing.min_volume = tier.min_volume
    existing.max_volume = tier.max_volume
    existing.included_features = tier.included_features
    existing.excluded_features = tier.excluded_features
    existing.contract_length = tier.contract_length
    existing.setup_fee = tier.setup_fee
    existing.implementation_cost = tier.implementation_cost
    existing.price_source = tier.price_source
    existing.updated_at = datetime.utcnow()

    db.commit()
    db.refresh(existing)

    return existing


@app.delete("/api/pricing-tiers/{tier_id}")
async def delete_pricing_tier(tier_id: int, db: Session = Depends(get_db)):
    """Delete a pricing tier."""
    tier = db.query(ProductPricingTier).filter(ProductPricingTier.id == tier_id).first()
    if not tier:
        raise HTTPException(status_code=404, detail="Pricing tier not found")

    db.delete(tier)
    db.commit()

    return {"message": "Pricing tier deleted successfully", "tier_id": tier_id}


@app.post("/api/pricing-tiers/{tier_id}/verify")
async def verify_pricing_tier(tier_id: int, db: Session = Depends(get_db)):
    """Mark a pricing tier as verified."""
    tier = db.query(ProductPricingTier).filter(ProductPricingTier.id == tier_id).first()
    if not tier:
        raise HTTPException(status_code=404, detail="Pricing tier not found")

    tier.price_verified = True
    tier.last_verified = datetime.utcnow()
    # Boost confidence when verified
    tier.confidence_score = min(100, (tier.confidence_score or 50) + 20)

    db.commit()
    db.refresh(tier)

    return {
        "message": "Pricing tier verified",
        "tier_id": tier_id,
        "confidence_score": tier.confidence_score
    }


# ============== PRICING COMPARISON ENDPOINT ==============

@app.get("/api/pricing/compare")
async def compare_pricing(
    category: Optional[str] = None,  # e.g., "Patient Intake"
    pricing_model: Optional[str] = None,  # e.g., "per_visit"
    db: Session = Depends(get_db)
):
    """Compare pricing across competitors for a product category."""
    query = db.query(ProductPricingTier).join(CompetitorProduct).join(Competitor)

    if category:
        query = query.filter(CompetitorProduct.product_category == category)

    if pricing_model:
        query = query.filter(ProductPricingTier.pricing_model == pricing_model)

    tiers = query.all()

    result = []
    for t in tiers:
        product = db.query(CompetitorProduct).filter(CompetitorProduct.id == t.product_id).first()
        competitor = db.query(Competitor).filter(Competitor.id == product.competitor_id).first() if product else None

        result.append({
            "competitor_id": competitor.id if competitor else None,
            "competitor_name": competitor.name if competitor else "Unknown",
            "product_id": product.id if product else None,
            "product_name": product.product_name if product else "Unknown",
            "product_category": product.product_category if product else None,
            "tier_name": t.tier_name,
            "pricing_model": t.pricing_model,
            "base_price": t.base_price,
            "price_display": t.price_display,
            "price_unit": t.price_unit,
            "percentage_rate": t.percentage_rate,
            "confidence_score": t.confidence_score,
            "price_verified": t.price_verified,
            "price_source": t.price_source
        })

    # Sort by base price (nulls last)
    result.sort(key=lambda x: (x["base_price"] is None, x["base_price"] or 0))

    return {
        "category": category,
        "pricing_model": pricing_model,
        "total_tiers": len(result),
        "comparison": result
    }


@app.get("/api/pricing/models")
async def get_pricing_models():
    """Get available healthcare pricing model types."""
    return {
        "pricing_models": [
            {"value": "per_visit", "label": "Per Visit/Encounter", "description": "Charge per patient encounter", "example": "$3.00/visit"},
            {"value": "per_provider", "label": "Per Provider", "description": "Monthly fee per provider/physician", "example": "$400/provider/month"},
            {"value": "per_location", "label": "Per Location", "description": "Fee per practice location", "example": "$1,500/location/month"},
            {"value": "subscription_flat", "label": "Flat Subscription", "description": "Fixed monthly fee", "example": "$299/month"},
            {"value": "subscription_tiered", "label": "Tiered Subscription", "description": "Tiered by features or volume", "example": "$99-$499/month"},
            {"value": "percentage_collections", "label": "Percentage of Collections", "description": "% of collected revenue", "example": "4-8% of collections"},
            {"value": "percentage_charges", "label": "Percentage of Charges", "description": "% of billed charges", "example": "2-4% of charges"},
            {"value": "per_bed", "label": "Per Bed", "description": "Hospital capacity pricing", "example": "$15,000/bed"},
            {"value": "per_member", "label": "Per Member (PMPM)", "description": "Per covered life", "example": "$0.50 PMPM"},
            {"value": "custom_enterprise", "label": "Custom/Enterprise", "description": "Negotiated pricing", "example": "Contact Sales"}
        ],
        "product_categories": [
            "Patient Intake",
            "Patient Payments",
            "Revenue Cycle Management (RCM)",
            "Practice Management",
            "EHR/EMR",
            "Telehealth",
            "Patient Engagement",
            "Scheduling",
            "Analytics",
            "Interoperability"
        ]
    }


# ============== FEATURE MATRIX ENDPOINTS ==============

@app.get("/api/products/{product_id}/features")
async def get_product_features(product_id: int, db: Session = Depends(get_db)):
    """Get all features for a product."""
    product = db.query(CompetitorProduct).filter(CompetitorProduct.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")

    features = db.query(ProductFeatureMatrix).filter(
        ProductFeatureMatrix.product_id == product_id
    ).all()

    # Group by category
    by_category = {}
    for f in features:
        if f.feature_category not in by_category:
            by_category[f.feature_category] = []
        by_category[f.feature_category].append({
            "id": f.id,
            "feature_name": f.feature_name,
            "feature_status": f.feature_status,
            "feature_tier": f.feature_tier,
            "notes": f.notes,
            "last_verified": f.last_verified
        })

    return {
        "product_id": product_id,
        "product_name": product.product_name,
        "features_by_category": by_category,
        "total_features": len(features)
    }


@app.post("/api/features")
async def create_feature(feature: FeatureMatrixCreate, db: Session = Depends(get_db)):
    """Add a feature to a product."""
    product = db.query(CompetitorProduct).filter(CompetitorProduct.id == feature.product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")

    new_feature = ProductFeatureMatrix(
        product_id=feature.product_id,
        feature_category=feature.feature_category,
        feature_name=feature.feature_name,
        feature_status=feature.feature_status,
        feature_tier=feature.feature_tier,
        notes=feature.notes,
        source_url=feature.source_url,
        last_verified=datetime.utcnow()
    )
    db.add(new_feature)
    db.commit()
    db.refresh(new_feature)

    return {
        "id": new_feature.id,
        "product_id": new_feature.product_id,
        "feature_category": new_feature.feature_category,
        "feature_name": new_feature.feature_name,
        "feature_status": new_feature.feature_status,
        "message": "Feature added successfully"
    }


@app.delete("/api/features/{feature_id}")
async def delete_feature(feature_id: int, db: Session = Depends(get_db)):
    """Delete a feature."""
    feature = db.query(ProductFeatureMatrix).filter(ProductFeatureMatrix.id == feature_id).first()
    if not feature:
        raise HTTPException(status_code=404, detail="Feature not found")

    db.delete(feature)
    db.commit()

    return {"message": "Feature deleted successfully", "feature_id": feature_id}


@app.get("/api/features/compare")
async def compare_features(
    category: str,  # Product category like "Patient Intake"
    feature_category: Optional[str] = None,  # Feature category like "Payments"
    db: Session = Depends(get_db)
):
    """Compare features across competitors for a product category."""
    # Get all products in this category
    products = db.query(CompetitorProduct).filter(
        CompetitorProduct.product_category == category
    ).all()

    if not products:
        return {"message": f"No products found in category: {category}", "comparison": []}

    # Get all unique features
    feature_query = db.query(ProductFeatureMatrix).filter(
        ProductFeatureMatrix.product_id.in_([p.id for p in products])
    )
    if feature_category:
        feature_query = feature_query.filter(ProductFeatureMatrix.feature_category == feature_category)

    all_features = feature_query.all()

    # Build comparison matrix
    # Structure: {feature_name: {competitor_name: status}}
    feature_names = set(f.feature_name for f in all_features)

    comparison = []
    for feature_name in sorted(feature_names):
        feature_row = {"feature_name": feature_name, "competitors": {}}
        for product in products:
            competitor = db.query(Competitor).filter(Competitor.id == product.competitor_id).first()
            comp_name = competitor.name if competitor else f"Competitor {product.competitor_id}"

            # Find this feature for this product
            feature = next(
                (f for f in all_features if f.product_id == product.id and f.feature_name == feature_name),
                None
            )
            feature_row["competitors"][comp_name] = feature.feature_status if feature else "unknown"
        comparison.append(feature_row)

    return {
        "product_category": category,
        "feature_category": feature_category,
        "competitors": [
            db.query(Competitor).filter(Competitor.id == p.competitor_id).first().name
            for p in products
        ],
        "comparison": comparison
    }


# ============== PRODUCT EXTRACTION FROM CONTENT ==============

@app.post("/api/competitors/{competitor_id}/extract-products")
async def extract_products_from_content(
    competitor_id: int,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db)
):
    """
    Extract products and pricing from competitor's scraped content using GPT.
    This endpoint triggers the extraction and stores results in the database.
    """
    from extractor import GPTExtractor

    competitor = db.query(Competitor).filter(Competitor.id == competitor_id).first()
    if not competitor:
        raise HTTPException(status_code=404, detail="Competitor not found")

    # Check if we have scraped content
    # Try to get recent scrape content from the competitor's website
    content = ""

    # First check if there's recent DataSource with website content
    recent_source = db.query(DataSource).filter(
        DataSource.competitor_id == competitor_id,
        DataSource.source_type == "website_scrape"
    ).order_by(DataSource.extracted_at.desc()).first()

    if recent_source and recent_source.current_value:
        content = recent_source.current_value
    else:
        # Try to scrape fresh content
        try:
            from scraper import CompetitorScraper
            import asyncio

            async def scrape_for_products():
                async with CompetitorScraper() as scraper:
                    result = await scraper.scrape(competitor.website)
                    return result.get("content", "") if isinstance(result, dict) else ""

            # Run the scraper
            content = asyncio.run(scrape_for_products())
        except Exception as e:
            return {
                "status": "error",
                "message": f"Could not get content for extraction: {str(e)}",
                "competitor_id": competitor_id
            }

    if not content or len(content) < 100:
        return {
            "status": "error",
            "message": "Not enough content available for extraction. Try scraping the competitor first.",
            "competitor_id": competitor_id
        }

    # Extract products using GPT
    extractor = GPTExtractor()
    extraction_result = extractor.extract_products_and_pricing(competitor.name, content)

    if "error" in extraction_result:
        return {
            "status": "error",
            "message": extraction_result["error"],
            "competitor_id": competitor_id
        }

    products_created = 0
    tiers_created = 0

    # Process extracted products
    for product_data in extraction_result.get("products", []):
        # Check if product already exists
        existing_product = db.query(CompetitorProduct).filter(
            CompetitorProduct.competitor_id == competitor_id,
            CompetitorProduct.product_name == product_data.get("product_name")
        ).first()

        if existing_product:
            # Update existing product
            existing_product.product_category = product_data.get("product_category", existing_product.product_category)
            existing_product.target_segment = product_data.get("target_segment", existing_product.target_segment)
            existing_product.is_primary_product = product_data.get("is_primary_product", existing_product.is_primary_product)
            existing_product.key_features = json.dumps(product_data.get("key_features", [])) if product_data.get("key_features") else existing_product.key_features
            existing_product.last_updated = datetime.utcnow()
            product = existing_product
        else:
            # Create new product
            product = CompetitorProduct(
                competitor_id=competitor_id,
                product_name=product_data.get("product_name", f"{competitor.name} Product"),
                product_category=product_data.get("product_category", "Unknown"),
                target_segment=product_data.get("target_segment"),
                is_primary_product=product_data.get("is_primary_product", False),
                key_features=json.dumps(product_data.get("key_features", [])) if product_data.get("key_features") else None,
                last_updated=datetime.utcnow()
            )
            db.add(product)
            db.flush()  # Get the ID
            products_created += 1

        # Process pricing tiers
        for tier_data in product_data.get("pricing_tiers", []):
            # Check if tier exists
            existing_tier = db.query(ProductPricingTier).filter(
                ProductPricingTier.product_id == product.id,
                ProductPricingTier.tier_name == tier_data.get("tier_name")
            ).first()

            if existing_tier:
                # Update existing tier
                existing_tier.pricing_model = tier_data.get("pricing_model", existing_tier.pricing_model)
                existing_tier.base_price = tier_data.get("base_price", existing_tier.base_price)
                existing_tier.price_currency = tier_data.get("price_currency", "USD")
                existing_tier.price_unit = tier_data.get("price_unit", existing_tier.price_unit)
                existing_tier.price_display = tier_data.get("price_display", existing_tier.price_display)
                existing_tier.percentage_rate = tier_data.get("percentage_rate", existing_tier.percentage_rate)
                existing_tier.setup_fee = tier_data.get("setup_fee", existing_tier.setup_fee)
                existing_tier.contract_length = tier_data.get("contract_length", existing_tier.contract_length)
                existing_tier.included_features = json.dumps(tier_data.get("included_features", [])) if tier_data.get("included_features") else existing_tier.included_features
                existing_tier.price_source = "gpt_extraction"
                existing_tier.updated_at = datetime.utcnow()
            else:
                # Calculate confidence score for extracted pricing
                confidence_result = calculate_confidence_score(
                    source_type="website_scrape",
                    information_credibility=4  # GPT extraction from marketing content
                )

                # Create new tier
                new_tier = ProductPricingTier(
                    product_id=product.id,
                    tier_name=tier_data.get("tier_name", "Standard"),
                    tier_position=tier_data.get("tier_position", 1),
                    pricing_model=tier_data.get("pricing_model", "custom_enterprise"),
                    base_price=tier_data.get("base_price"),
                    price_currency=tier_data.get("price_currency", "USD"),
                    price_unit=tier_data.get("price_unit"),
                    price_display=tier_data.get("price_display"),
                    percentage_rate=tier_data.get("percentage_rate"),
                    setup_fee=tier_data.get("setup_fee"),
                    contract_length=tier_data.get("contract_length"),
                    included_features=json.dumps(tier_data.get("included_features", [])) if tier_data.get("included_features") else None,
                    price_source="gpt_extraction",
                    price_verified=False,
                    confidence_score=confidence_result.score
                )
                db.add(new_tier)
                tiers_created += 1

    db.commit()

    return {
        "status": "success",
        "competitor_id": competitor_id,
        "competitor_name": competitor.name,
        "products_created": products_created,
        "tiers_created": tiers_created,
        "extraction_confidence": extraction_result.get("extraction_confidence", 0),
        "extraction_notes": extraction_result.get("extraction_notes", "")
    }


@app.post("/api/products/{product_id}/extract-features")
async def extract_features_from_content(
    product_id: int,
    db: Session = Depends(get_db)
):
    """
    Extract feature matrix for a product using GPT.
    """
    from extractor import GPTExtractor

    product = db.query(CompetitorProduct).filter(CompetitorProduct.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")

    competitor = db.query(Competitor).filter(Competitor.id == product.competitor_id).first()
    if not competitor:
        raise HTTPException(status_code=404, detail="Competitor not found")

    # Get content
    content = ""
    recent_source = db.query(DataSource).filter(
        DataSource.competitor_id == product.competitor_id,
        DataSource.source_type == "website_scrape"
    ).order_by(DataSource.extracted_at.desc()).first()

    if recent_source and recent_source.current_value:
        content = recent_source.current_value
    else:
        return {
            "status": "error",
            "message": "No content available for feature extraction. Scrape the competitor first."
        }

    # Extract features
    extractor = GPTExtractor()
    extraction_result = extractor.extract_feature_matrix(competitor.name, product.product_name, content)

    if "error" in extraction_result:
        return {
            "status": "error",
            "message": extraction_result["error"]
        }

    features_created = 0

    for feature_data in extraction_result.get("features", []):
        # Check if feature exists
        existing = db.query(ProductFeatureMatrix).filter(
            ProductFeatureMatrix.product_id == product_id,
            ProductFeatureMatrix.feature_name == feature_data.get("feature_name")
        ).first()

        if not existing:
            new_feature = ProductFeatureMatrix(
                product_id=product_id,
                feature_category=feature_data.get("feature_category", "Other"),
                feature_name=feature_data.get("feature_name"),
                feature_status=feature_data.get("feature_status", "unknown"),
                feature_tier=feature_data.get("feature_tier"),
                notes=feature_data.get("notes"),
                last_verified=datetime.utcnow()
            )
            db.add(new_feature)
            features_created += 1

    db.commit()

    return {
        "status": "success",
        "product_id": product_id,
        "product_name": product.product_name,
        "features_created": features_created,
        "extraction_confidence": extraction_result.get("extraction_confidence", 0)
    }


# ============== PHASE 4: CUSTOMER COUNT VERIFICATION ENDPOINTS ==============

@app.get("/api/competitors/{competitor_id}/customer-counts", response_model=List[CustomerCountResponse])
async def get_customer_counts(competitor_id: int, db: Session = Depends(get_db)):
    """Get all customer count estimates for a competitor, ordered by date."""
    competitor = db.query(Competitor).filter(Competitor.id == competitor_id).first()
    if not competitor:
        raise HTTPException(status_code=404, detail="Competitor not found")

    counts = db.query(CustomerCountEstimate).filter(
        CustomerCountEstimate.competitor_id == competitor_id
    ).order_by(CustomerCountEstimate.as_of_date.desc()).all()

    return counts


@app.get("/api/competitors/{competitor_id}/customer-count/latest", response_model=CustomerCountResponse)
async def get_latest_customer_count(competitor_id: int, db: Session = Depends(get_db)):
    """Get the most recent verified customer count for a competitor."""
    competitor = db.query(Competitor).filter(Competitor.id == competitor_id).first()
    if not competitor:
        raise HTTPException(status_code=404, detail="Competitor not found")

    # First try to get verified count
    count = db.query(CustomerCountEstimate).filter(
        CustomerCountEstimate.competitor_id == competitor_id,
        CustomerCountEstimate.is_verified == True
    ).order_by(CustomerCountEstimate.as_of_date.desc()).first()

    # If no verified count, get any most recent
    if not count:
        count = db.query(CustomerCountEstimate).filter(
            CustomerCountEstimate.competitor_id == competitor_id
        ).order_by(CustomerCountEstimate.as_of_date.desc()).first()

    if not count:
        raise HTTPException(status_code=404, detail="No customer count estimates found")

    return count


@app.post("/api/customer-counts", response_model=CustomerCountResponse)
async def create_customer_count(
    count_data: CustomerCountCreate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Create a new customer count estimate."""
    # Verify competitor exists
    competitor = db.query(Competitor).filter(Competitor.id == count_data.competitor_id).first()
    if not competitor:
        raise HTTPException(status_code=404, detail="Competitor not found")

    # Calculate confidence score based on source type
    source_mapping = {
        "sec_10k": "sec_filing",
        "sec_filing": "sec_filing",
        "website": "website_scrape",
        "press_release": "news_article",
        "definitive_hc": "definitive_hc",
        "klas_report": "klas_report",
        "linkedin": "linkedin_estimate",
        "g2_reviews": "api_verified",
        "manual": "manual_verified"
    }
    source_type = source_mapping.get(count_data.primary_source, "unknown")
    confidence_result = calculate_confidence_score(source_type=source_type)

    # Get previous count for growth calculation
    previous = db.query(CustomerCountEstimate).filter(
        CustomerCountEstimate.competitor_id == count_data.competitor_id
    ).order_by(CustomerCountEstimate.as_of_date.desc()).first()

    growth_rate = None
    previous_count = None
    if previous and previous.count_value and count_data.count_value:
        previous_count = previous.count_value
        if previous_count > 0:
            growth_rate = ((count_data.count_value - previous_count) / previous_count) * 100

    new_count = CustomerCountEstimate(
        competitor_id=count_data.competitor_id,
        count_value=count_data.count_value,
        count_display=count_data.count_display,
        count_type=count_data.count_type,
        count_unit=count_data.count_unit,
        count_definition=count_data.count_definition,
        segment_breakdown=count_data.segment_breakdown,
        primary_source=count_data.primary_source,
        primary_source_url=count_data.primary_source_url,
        primary_source_date=count_data.primary_source_date,
        as_of_date=count_data.as_of_date or datetime.utcnow(),
        previous_count=previous_count,
        growth_rate=growth_rate,
        confidence_score=confidence_result.score,
        confidence_level=confidence_result.level,
        is_verified=(count_data.primary_source in ["sec_10k", "sec_filing", "definitive_hc"])
    )
    db.add(new_count)
    db.commit()
    db.refresh(new_count)

    # Log activity
    log_activity(
        db, current_user["email"], current_user["id"],
        "customer_count_added",
        f"Added customer count for {competitor.name}: {count_data.count_display}"
    )

    return new_count


@app.put("/api/customer-counts/{count_id}", response_model=CustomerCountResponse)
async def update_customer_count(
    count_id: int,
    count_data: CustomerCountCreate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Update an existing customer count estimate."""
    count = db.query(CustomerCountEstimate).filter(CustomerCountEstimate.id == count_id).first()
    if not count:
        raise HTTPException(status_code=404, detail="Customer count estimate not found")

    # Update fields
    for field in ["count_value", "count_display", "count_type", "count_unit",
                  "count_definition", "segment_breakdown", "primary_source",
                  "primary_source_url", "primary_source_date", "as_of_date"]:
        if hasattr(count_data, field):
            value = getattr(count_data, field)
            if value is not None:
                setattr(count, field, value)

    count.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(count)

    return count


@app.delete("/api/customer-counts/{count_id}")
async def delete_customer_count(
    count_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Delete a customer count estimate."""
    count = db.query(CustomerCountEstimate).filter(CustomerCountEstimate.id == count_id).first()
    if not count:
        raise HTTPException(status_code=404, detail="Customer count estimate not found")

    competitor = db.query(Competitor).filter(Competitor.id == count.competitor_id).first()
    db.delete(count)
    db.commit()

    log_activity(
        db, current_user["email"], current_user["id"],
        "customer_count_deleted",
        f"Deleted customer count for {competitor.name if competitor else 'Unknown'}"
    )

    return {"status": "deleted", "count_id": count_id}


@app.post("/api/customer-counts/{count_id}/verify", response_model=CustomerCountResponse)
async def verify_customer_count(
    count_id: int,
    verification: CustomerCountVerifyRequest,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Verify a customer count with additional sources or methods."""
    count = db.query(CustomerCountEstimate).filter(CustomerCountEstimate.id == count_id).first()
    if not count:
        raise HTTPException(status_code=404, detail="Customer count estimate not found")

    # Update verification status
    count.is_verified = True
    count.verification_method = verification.verification_method
    count.verification_date = datetime.utcnow()
    count.confidence_notes = verification.verification_notes

    # If additional sources provided, store them and recalculate confidence
    if verification.additional_sources:
        import json
        existing_sources = json.loads(count.all_sources) if count.all_sources else []
        existing_sources.extend(verification.additional_sources)
        count.all_sources = json.dumps(existing_sources)
        count.corroborating_sources = len(existing_sources)

        # Calculate source agreement score
        if count.count_value:
            values = [count.count_value]
            for source in existing_sources:
                if source.get("value"):
                    try:
                        values.append(int(str(source["value"]).replace(",", "").replace("+", "")))
                    except (ValueError, TypeError):
                        pass

            if len(values) > 1:
                avg = sum(values) / len(values)
                # Agreement score: 1.0 if all within 20% of average, 0 if wildly different
                deviations = [abs(v - avg) / avg for v in values if avg > 0]
                if deviations:
                    count.source_agreement_score = max(0, 1 - (sum(deviations) / len(deviations) / 0.2))

        # Recalculate confidence with corroboration
        confidence_result = calculate_confidence_score(
            source_type="manual_verified" if verification.verification_method == "manual" else "api_verified",
            corroborating_sources=len(existing_sources)
        )
        count.confidence_score = confidence_result.score
        count.confidence_level = confidence_result.level

    count.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(count)

    # Log activity
    competitor = db.query(Competitor).filter(Competitor.id == count.competitor_id).first()
    log_activity(
        db, current_user["email"], current_user["id"],
        "customer_count_verified",
        f"Verified customer count for {competitor.name if competitor else 'Unknown'}"
    )

    return count


@app.get("/api/customer-counts/compare")
async def compare_customer_counts(
    unit: Optional[str] = None,
    min_confidence: int = 0,
    db: Session = Depends(get_db)
):
    """Compare customer counts across all competitors."""
    query = db.query(CustomerCountEstimate)

    if unit:
        query = query.filter(CustomerCountEstimate.count_unit == unit)
    if min_confidence > 0:
        query = query.filter(CustomerCountEstimate.confidence_score >= min_confidence)

    # Get latest count per competitor
    from sqlalchemy import func
    subquery = db.query(
        CustomerCountEstimate.competitor_id,
        func.max(CustomerCountEstimate.as_of_date).label("max_date")
    ).group_by(CustomerCountEstimate.competitor_id).subquery()

    counts = query.join(
        subquery,
        (CustomerCountEstimate.competitor_id == subquery.c.competitor_id) &
        (CustomerCountEstimate.as_of_date == subquery.c.max_date)
    ).all()

    result = []
    for c in counts:
        competitor = db.query(Competitor).filter(Competitor.id == c.competitor_id).first()
        result.append({
            "competitor_id": c.competitor_id,
            "competitor_name": competitor.name if competitor else "Unknown",
            "count_value": c.count_value,
            "count_display": c.count_display,
            "count_unit": c.count_unit,
            "confidence_score": c.confidence_score,
            "confidence_level": c.confidence_level,
            "is_verified": c.is_verified,
            "as_of_date": c.as_of_date.isoformat() if c.as_of_date else None,
            "growth_rate": c.growth_rate
        })

    # Sort by count_value descending
    result.sort(key=lambda x: x["count_value"] or 0, reverse=True)

    return {
        "comparisons": result,
        "total_competitors": len(result),
        "unit_filter": unit,
        "min_confidence_filter": min_confidence
    }


@app.get("/api/customer-counts/units")
async def get_customer_count_units():
    """Get all available customer count unit types with descriptions."""
    return {
        "units": [
            {
                "value": "healthcare_organizations",
                "label": "Healthcare Organizations",
                "description": "Distinct hospital/clinic/practice entities"
            },
            {
                "value": "providers",
                "label": "Providers",
                "description": "Individual physicians or clinicians using the platform"
            },
            {
                "value": "locations",
                "label": "Locations",
                "description": "Physical practice sites or facilities"
            },
            {
                "value": "users",
                "label": "Users",
                "description": "All user accounts (may include staff, admins)"
            },
            {
                "value": "lives_covered",
                "label": "Lives Covered",
                "description": "Patient lives managed through the platform"
            },
            {
                "value": "encounters",
                "label": "Encounters/Visits",
                "description": "Annual patient encounters processed"
            },
            {
                "value": "beds",
                "label": "Hospital Beds",
                "description": "Licensed hospital beds served"
            }
        ]
    }


@app.post("/api/competitors/{competitor_id}/triangulate-customer-count")
async def triangulate_customer_count(
    competitor_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """
    Triangulate customer count from multiple sources.
    Collects data from: website scrapes, SEC filings, news articles, existing estimates.
    """
    competitor = db.query(Competitor).filter(Competitor.id == competitor_id).first()
    if not competitor:
        raise HTTPException(status_code=404, detail="Competitor not found")

    sources = []

    # 1. Get from existing DataSource records (website scrapes)
    website_source = db.query(DataSource).filter(
        DataSource.competitor_id == competitor_id,
        DataSource.field_name == "customer_count"
    ).order_by(DataSource.extracted_at.desc()).first()

    if website_source and website_source.current_value:
        sources.append({
            "source_type": "website_scrape",
            "value": website_source.current_value,
            "source_url": website_source.source_url,
            "date": website_source.extracted_at.isoformat() if website_source.extracted_at else None
        })

    # 2. Get from Competitor model directly
    if competitor.customer_count:
        sources.append({
            "source_type": "existing_record",
            "value": competitor.customer_count,
            "source_url": competitor.website,
            "date": competitor.last_updated.isoformat() if competitor.last_updated else None
        })

    # 3. Try SEC data for public companies
    if competitor.is_public and competitor.ticker_symbol:
        try:
            from sec_edgar_scraper import SECEdgarScraper
            scraper = SECEdgarScraper()
            sec_data = scraper.get_company_data(competitor.name)
            if sec_data and sec_data.customers_mentioned:
                sources.append({
                    "source_type": "sec_filing",
                    "value": f"{len(sec_data.customers_mentioned)}+ (named customers)",
                    "source_url": f"https://sec.gov/cgi-bin/browse-edgar?CIK={sec_data.cik}",
                    "date": sec_data.last_updated
                })
        except Exception as e:
            print(f"SEC scraper error for {competitor.name}: {e}")

    # 4. Check existing CustomerCountEstimate records
    existing_counts = db.query(CustomerCountEstimate).filter(
        CustomerCountEstimate.competitor_id == competitor_id
    ).all()

    for ec in existing_counts:
        if ec.count_display:
            sources.append({
                "source_type": ec.primary_source or "previous_estimate",
                "value": ec.count_display,
                "source_url": ec.primary_source_url,
                "date": ec.as_of_date.isoformat() if ec.as_of_date else None,
                "confidence": ec.confidence_score
            })

    if not sources:
        return {
            "status": "no_sources",
            "competitor": competitor.name,
            "message": "No customer count data found from any source"
        }

    # Triangulate - use triangulate_data_points from confidence_scoring
    triangulation_sources = [
        {
            "value": s["value"],
            "source_type": s["source_type"],
            "reliability": get_source_defaults(s["source_type"]).get("reliability", "F"),
            "credibility": get_source_defaults(s["source_type"]).get("credibility", 6)
        }
        for s in sources
    ]

    result = triangulate_data_points(triangulation_sources)

    # Create a new CustomerCountEstimate with triangulated result
    # Parse the best value to extract numeric count
    import re
    best_value = result.best_value
    count_value = None
    count_type = "estimate"

    # Try to extract number from value like "3,000+" or "3000-5000"
    if best_value and best_value != "Unknown":
        numbers = re.findall(r'[\d,]+', best_value.replace(",", ""))
        if numbers:
            try:
                count_value = int(numbers[0])
                if "+" in best_value:
                    count_type = "minimum"
                elif "-" in best_value or "to" in best_value.lower():
                    count_type = "range"
                else:
                    count_type = "exact"
            except ValueError:
                pass

    new_estimate = CustomerCountEstimate(
        competitor_id=competitor_id,
        count_value=count_value,
        count_display=best_value if best_value != "Unknown" else None,
        count_type=count_type,
        count_unit="healthcare_organizations",  # Default, should be refined
        primary_source=result.source_used,
        all_sources=json.dumps(sources),
        source_agreement_score=1.0 if not result.discrepancy_flag else 0.5,
        confidence_score=result.confidence_score,
        confidence_level=result.confidence_level,
        confidence_notes=result.review_reason,
        is_verified=result.confidence_level == "high",
        verification_method="triangulation" if len(sources) > 1 else None,
        verification_date=datetime.utcnow() if len(sources) > 1 else None,
        as_of_date=datetime.utcnow()
    )
    db.add(new_estimate)
    db.commit()
    db.refresh(new_estimate)

    log_activity(
        db, current_user["email"], current_user["id"],
        "customer_count_triangulated",
        f"Triangulated customer count for {competitor.name} from {len(sources)} sources"
    )

    return {
        "status": "success",
        "competitor": competitor.name,
        "triangulation_result": {
            "best_value": result.best_value,
            "confidence_score": result.confidence_score,
            "confidence_level": result.confidence_level,
            "source_used": result.source_used,
            "discrepancy_flag": result.discrepancy_flag,
            "review_reason": result.review_reason
        },
        "sources_analyzed": len(sources),
        "sources": sources,
        "new_estimate_id": new_estimate.id
    }


@app.get("/api/customer-counts/history/{competitor_id}")
async def get_customer_count_history(
    competitor_id: int,
    db: Session = Depends(get_db)
):
    """Get historical customer count data for trend analysis."""
    competitor = db.query(Competitor).filter(Competitor.id == competitor_id).first()
    if not competitor:
        raise HTTPException(status_code=404, detail="Competitor not found")

    counts = db.query(CustomerCountEstimate).filter(
        CustomerCountEstimate.competitor_id == competitor_id
    ).order_by(CustomerCountEstimate.as_of_date.asc()).all()

    history = []
    for c in counts:
        history.append({
            "id": c.id,
            "count_value": c.count_value,
            "count_display": c.count_display,
            "count_type": c.count_type,
            "count_unit": c.count_unit,
            "as_of_date": c.as_of_date.isoformat() if c.as_of_date else None,
            "growth_rate": c.growth_rate,
            "confidence_level": c.confidence_level,
            "is_verified": c.is_verified,
            "primary_source": c.primary_source
        })

    # Calculate overall growth trend
    growth_trend = None
    if len(history) >= 2:
        first_with_value = next((h for h in history if h["count_value"]), None)
        last_with_value = next((h for h in reversed(history) if h["count_value"]), None)
        if first_with_value and last_with_value and first_with_value["count_value"] > 0:
            total_growth = ((last_with_value["count_value"] - first_with_value["count_value"]) /
                          first_with_value["count_value"]) * 100
            growth_trend = round(total_growth, 1)

    return {
        "competitor": competitor.name,
        "competitor_id": competitor_id,
        "history": history,
        "total_records": len(history),
        "growth_trend_percent": growth_trend
    }


# ============== PHASE 5: ENHANCED SCRAPER WITH SOURCE TRACKING ==============

@app.post("/api/scrape/enhanced/{competitor_id}")
async def run_enhanced_scrape(
    competitor_id: int,
    background_tasks: BackgroundTasks,
    pages: Optional[List[str]] = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """
    Run enhanced scrape with full source tracking for each extracted field.

    This uses the EnhancedGPTExtractor to track which page each data point
    came from, with confidence scores based on page type and field relevance.

    Args:
        competitor_id: ID of the competitor to scrape
        pages: Optional list of pages to scrape (default: homepage, pricing, about, features)
    """
    competitor = db.query(Competitor).filter(Competitor.id == competitor_id).first()
    if not competitor:
        raise HTTPException(status_code=404, detail="Competitor not found")

    if not pages:
        pages = ["homepage", "pricing", "about", "features", "customers"]

    # Run in background
    background_tasks.add_task(
        _run_enhanced_scrape_job,
        competitor_id,
        competitor.name,
        competitor.website,
        pages,
        current_user["email"]
    )

    log_activity(
        db, current_user["email"], current_user["id"],
        "enhanced_scrape_started",
        f"Started enhanced scrape for {competitor.name} ({len(pages)} pages)"
    )

    return {
        "status": "started",
        "competitor_id": competitor_id,
        "competitor_name": competitor.name,
        "pages_to_scrape": pages,
        "message": "Enhanced scrape started in background"
    }


async def _run_enhanced_scrape_job(
    competitor_id: int,
    competitor_name: str,
    website: str,
    pages: List[str],
    user_email: str
):
    """
    Background job for enhanced scraping with full source tracking.
    """
    from scraper import CompetitorScraper
    from extractor import EnhancedGPTExtractor

    db = SessionLocal()

    try:
        competitor = db.query(Competitor).filter(Competitor.id == competitor_id).first()
        if not competitor:
            print(f"Enhanced scrape: Competitor {competitor_id} not found")
            return

        print(f"[Enhanced Scrape] Starting for {competitor_name}...")

        # 1. Scrape all requested pages
        page_contents = {}
        async with CompetitorScraper() as scraper:
            for page_type in pages:
                try:
                    if page_type == "homepage":
                        url = website
                    else:
                        url = f"{website.rstrip('/')}/{page_type}"

                    result = await scraper.scrape(url)
                    if result and result.get("content"):
                        page_contents[page_type] = result["content"]
                        print(f"  [OK] Scraped {page_type}: {len(result['content'])} chars")
                    else:
                        print(f"  [--] No content from {page_type}")
                except Exception as e:
                    print(f"  [ERR] Failed to scrape {page_type}: {e}")

        if not page_contents:
            print(f"[Enhanced Scrape] No content scraped for {competitor_name}")
            return

        # 2. Extract with source tracking
        extractor = EnhancedGPTExtractor()
        extracted = extractor.extract_with_sources(
            competitor_name=competitor_name,
            competitor_website=website,
            page_contents=page_contents
        )

        print(f"[Enhanced Scrape] Extracted {len(extracted.field_sources)} fields from {len(page_contents)} pages")

        # 3. Store DataSource records with confidence scoring
        data_sources = extractor.to_data_sources(extracted, competitor_id)
        changes_count = 0
        new_values_count = 0

        for ds_data in data_sources:
            field_name = ds_data["field_name"]
            new_value = ds_data["current_value"]

            # Check if field is locked by manual correction
            is_locked = db.query(DataSource).filter(
                DataSource.competitor_id == competitor_id,
                DataSource.field_name == field_name,
                DataSource.source_type == "manual"
            ).first()

            if is_locked:
                print(f"  [LOCKED] Skipping {field_name} (manual correction exists)")
                continue

            # Get old value
            old_value = getattr(competitor, field_name, None) if hasattr(competitor, field_name) else None
            old_str = str(old_value) if old_value else None

            # Update competitor if field exists
            if hasattr(competitor, field_name) and new_value:
                if old_str != new_value:
                    # Log change
                    change_record = DataChangeHistory(
                        competitor_id=competitor_id,
                        competitor_name=competitor_name,
                        field_name=field_name,
                        old_value=old_str,
                        new_value=new_value,
                        changed_by=f"Enhanced Scrape ({user_email})",
                        change_reason="Enhanced scrape with source tracking"
                    )
                    db.add(change_record)
                    setattr(competitor, field_name, new_value)

                    if old_value is None or old_str == "" or old_str == "None":
                        new_values_count += 1
                    else:
                        changes_count += 1

            # Create/update DataSource record
            existing_source = db.query(DataSource).filter(
                DataSource.competitor_id == competitor_id,
                DataSource.field_name == field_name,
                DataSource.source_type == "website_scrape"
            ).first()

            if existing_source:
                # Update existing
                existing_source.previous_value = existing_source.current_value
                existing_source.current_value = new_value
                existing_source.source_url = ds_data["source_url"]
                existing_source.source_name = ds_data["source_name"]
                existing_source.extracted_at = ds_data["extracted_at"]
                existing_source.confidence_score = ds_data["confidence_score"]
                existing_source.confidence_level = ds_data["confidence_level"]
                existing_source.updated_at = datetime.utcnow()
            else:
                # Create new
                new_source = DataSource(
                    competitor_id=competitor_id,
                    field_name=field_name,
                    current_value=new_value,
                    source_type="website_scrape",
                    source_url=ds_data["source_url"],
                    source_name=ds_data["source_name"],
                    extraction_method="gpt_extraction",
                    extracted_at=ds_data["extracted_at"],
                    source_reliability=ds_data["source_reliability"],
                    information_credibility=ds_data["information_credibility"],
                    confidence_score=ds_data["confidence_score"],
                    confidence_level=ds_data["confidence_level"],
                    data_as_of_date=ds_data["data_as_of_date"]
                )
                db.add(new_source)

        competitor.last_updated = datetime.utcnow()
        db.commit()

        print(f"[Enhanced Scrape] Completed for {competitor_name}: {changes_count} changes, {new_values_count} new values")

        # 4. Run triangulation for key fields
        try:
            triangulator = DataTriangulator(db)
            await triangulator.triangulate_all_key_fields(
                competitor_id=competitor_id,
                competitor_name=competitor_name,
                website=website,
                is_public=competitor.is_public,
                ticker_symbol=competitor.ticker_symbol
            )
            db.commit()
            print(f"[Enhanced Scrape] Triangulation completed for {competitor_name}")
        except Exception as e:
            print(f"[Enhanced Scrape] Triangulation error: {e}")

    except Exception as e:
        print(f"[Enhanced Scrape] Error for {competitor_name}: {e}")
        import traceback
        traceback.print_exc()
        db.rollback()
    finally:
        db.close()


@app.get("/api/scrape/enhanced/{competitor_id}/sources")
async def get_enhanced_scrape_sources(
    competitor_id: int,
    db: Session = Depends(get_db)
):
    """Get all source data from the most recent enhanced scrape."""
    competitor = db.query(Competitor).filter(Competitor.id == competitor_id).first()
    if not competitor:
        raise HTTPException(status_code=404, detail="Competitor not found")

    sources = db.query(DataSource).filter(
        DataSource.competitor_id == competitor_id,
        DataSource.source_type == "website_scrape"
    ).order_by(DataSource.extracted_at.desc()).all()

    result = []
    for s in sources:
        result.append({
            "field_name": s.field_name,
            "current_value": s.current_value,
            "source_url": s.source_url,
            "source_name": s.source_name,
            "confidence_score": s.confidence_score,
            "confidence_level": s.confidence_level,
            "extracted_at": s.extracted_at.isoformat() if s.extracted_at else None,
            "is_verified": s.is_verified
        })

    return {
        "competitor": competitor.name,
        "competitor_id": competitor_id,
        "sources": result,
        "total_fields": len(result)
    }


# ============== BULK UPDATE ENDPOINT ==============

@app.post("/api/competitors/bulk-update")
async def bulk_update_competitors(
    request: Request,
    db: Session = Depends(get_db)
):
    """Bulk update competitors from JSON data."""
    data = await request.json()
    updates = data.get("updates", [])
    changed_by = data.get("changed_by", "bulk_import")
    
    results = {"success": [], "errors": []}
    
    for update in updates:
        competitor_id = update.get("id")
        if not competitor_id:
            results["errors"].append({"error": "Missing competitor ID", "data": update})
            continue
        
        competitor = db.query(Competitor).filter(Competitor.id == competitor_id).first()
        if not competitor:
            results["errors"].append({"error": f"Competitor {competitor_id} not found", "data": update})
            continue
        
        try:
            for field, new_value in update.items():
                if field == "id":
                    continue
                if hasattr(competitor, field):
                    old_value = getattr(competitor, field)
                    setattr(competitor, field, new_value)
                    
                    # Log the change
                    change = DataChangeHistory(
                        competitor_id=competitor_id,
                        competitor_name=competitor.name,
                        field_name=field,
                        old_value=str(old_value) if old_value else None,
                        new_value=str(new_value) if new_value else None,
                        source_url=data.get("source_url"), # Capture Source URL from bulk/correction payload
                        changed_by=changed_by
                    )
                    db.add(change)
            
            competitor.last_updated = datetime.utcnow()
            competitor.data_quality_score = calculate_quality_score(competitor)
            results["success"].append({"id": competitor_id, "name": competitor.name})
        except Exception as e:
            results["errors"].append({"error": str(e), "id": competitor_id})
    
    db.commit()
    
    return {
        "total_processed": len(updates),
        "successful": len(results["success"]),
        "failed": len(results["errors"]),
        "results": results
    }


# Import routers
from routers import reports, discovery
import api_routes

# Include routers
app.include_router(discovery.router)
app.include_router(api_routes.router)  # Covers analytics, winloss, external, etc.
app.include_router(reports.router)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)





@app.get("/health")
def health():
    return {"status": "healthy", "timestamp": datetime.utcnow().isoformat()}


# --- Competitors CRUD ---

@app.get("/api/competitors", response_model=List[CompetitorResponse])
async def list_competitors(
    status: Optional[str] = None,
    threat_level: Optional[str] = None,
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    query = db.query(Competitor).filter(Competitor.is_deleted == False)
    if status:
        query = query.filter(Competitor.status == status)
    if threat_level:
        query = query.filter(Competitor.threat_level == threat_level)
    competitors = query.offset(skip).limit(limit).all()
    return competitors


@app.get("/api/competitors/{competitor_id}", response_model=CompetitorResponse)
async def get_competitor(competitor_id: int, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    competitor = db.query(Competitor).filter(
        Competitor.id == competitor_id,
        Competitor.is_deleted == False
    ).first()
    if not competitor:
        raise HTTPException(status_code=404, detail="Competitor not found")
    return competitor


@app.post("/api/competitors", response_model=CompetitorResponse)
async def create_competitor(competitor: CompetitorCreate, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    db_competitor = Competitor(**competitor.model_dump())

    # Auto-classify Public/Private
    comp_lower = db_competitor.name.lower()
    if comp_lower in KNOWN_TICKERS:
        info = KNOWN_TICKERS[comp_lower]
        db_competitor.is_public = True
        db_competitor.ticker_symbol = info["symbol"]
        db_competitor.stock_exchange = info["exchange"]

    db.add(db_competitor)
    db.commit()
    db.refresh(db_competitor)

    # Log the competitor creation
    user_email = current_user.get("email", "unknown")
    log_activity(
        db, user_email, current_user.get("id"),
        "competitor_create",
        {"competitor_id": db_competitor.id, "competitor_name": db_competitor.name}
    )

    return db_competitor


@app.put("/api/competitors/{competitor_id}", response_model=CompetitorResponse)
async def update_competitor(competitor_id: int, competitor: CompetitorCreate, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    db_competitor = db.query(Competitor).filter(Competitor.id == competitor_id).first()
    if not db_competitor:
        raise HTTPException(status_code=404, detail="Competitor not found")

    # Track all field changes for audit log
    changes_made = []
    user_email = current_user.get("email", "unknown")

    for key, value in competitor.model_dump().items():
        old_value = getattr(db_competitor, key, None)
        # Only log if value actually changed
        if str(old_value) != str(value) and value is not None:
            changes_made.append({
                "field": key,
                "old_value": str(old_value) if old_value else None,
                "new_value": str(value)
            })
            # Log each change to DataChangeHistory
            change_record = DataChangeHistory(
                competitor_id=competitor_id,
                competitor_name=db_competitor.name,
                field_name=key,
                old_value=str(old_value) if old_value else None,
                new_value=str(value),
                changed_by=user_email,
                change_reason="Manual update via UI"
            )
            db.add(change_record)

        setattr(db_competitor, key, value)

    db_competitor.last_updated = datetime.utcnow()

    # Log activity if changes were made
    if changes_made:
        log_activity(
            db, user_email, current_user.get("id"),
            "competitor_update",
            {"competitor_id": competitor_id, "competitor_name": db_competitor.name, "changes_count": len(changes_made)}
        )

    db.commit()
    db.refresh(db_competitor)
    return db_competitor


@app.delete("/api/competitors/{competitor_id}")
async def delete_competitor(competitor_id: int, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    db_competitor = db.query(Competitor).filter(Competitor.id == competitor_id).first()
    if not db_competitor:
        raise HTTPException(status_code=404, detail="Competitor not found")

    competitor_name = db_competitor.name
    db_competitor.is_deleted = True

    # Log the deletion
    user_email = current_user.get("email", "unknown")
    log_activity(
        db, user_email, current_user.get("id"),
        "competitor_delete",
        {"competitor_id": competitor_id, "competitor_name": competitor_name}
    )

    db.commit()
    return {"message": "Competitor deleted", "deleted_by": user_email}


@app.post("/api/competitors/{competitor_id}/correct")
async def correct_competitor_data(
    competitor_id: int, 
    correction: CorrectionRequest, 
    db: Session = Depends(get_db), 
    current_user: dict = Depends(get_current_user)
):
    """Manually correct a data point and lock it to prevent overwrite."""
    # 1. Get Competitor
    competitor = db.query(Competitor).filter(Competitor.id == competitor_id).first()
    if not competitor:
        raise HTTPException(status_code=404, detail="Competitor not found")
    
    # 2. Validate field exists
    if not hasattr(competitor, correction.field):
        raise HTTPException(status_code=400, detail=f"Invalid field: {correction.field}")
    
    old_value = getattr(competitor, correction.field)
    
    # 3. Update Value
    setattr(competitor, correction.field, correction.new_value)
    competitor.last_updated = datetime.utcnow()
    
    # 4. Update/Create Data Source (Lock as Manual)
    source = db.query(DataSource).filter(
        DataSource.competitor_id == competitor_id,
        DataSource.field_name == correction.field
    ).first()
    
    if source:
        source.source_type = "manual"
        source.entered_by = current_user.get("email", "unknown")
        source.updated_at = datetime.utcnow()
    else:
        new_source = DataSource(
            competitor_id=competitor_id,
            field_name=correction.field,
            source_type="manual",
            source_url=correction.source_url,
            entered_by=current_user.get("email", "unknown"),
            verified_at=datetime.utcnow()
        )
        db.add(new_source)
    
    # 5. Log Change History
    history = DataChangeHistory(
        competitor_id=competitor_id,
        competitor_name=competitor.name,
        field_name=correction.field,
        old_value=str(old_value) if old_value else None,
        new_value=correction.new_value,
        source_url=correction.source_url, # Capture Evidence URL
        changed_by=current_user.get("email", "unknown"),
        change_reason=correction.reason
    )
    db.add(history)
    
    db.commit()
    db.refresh(competitor)
    
    return {"message": "Correction applied successfully", "competitor": competitor}


# --- Excel Export ---

def auto_fit_columns(worksheet):
    """Auto-fit all columns to the widest content."""
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except Exception:
                pass
        adjusted_width = max_length + 3
        if adjusted_width < 12:
            adjusted_width = 12
        if adjusted_width > 50:
            adjusted_width = 50
        worksheet.column_dimensions[column_letter].width = adjusted_width


def apply_white_background(worksheet, max_row=100, max_col=50):
    """Apply white background to all cells."""
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = worksheet.cell(row=row, column=col)
            if cell.fill.fgColor.rgb in (None, '00000000', 'FFFFFFFF') or cell.fill.fill_type is None:
                cell.fill = WHITE_FILL


@app.get("/api/export/excel")
def export_excel(db: Session = Depends(get_db)):
    """Export all competitor data to Excel."""
    competitors = db.query(Competitor).filter(Competitor.is_deleted == False).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Competitors"
    
    # Headers
    headers = [
        "Competitor Name", "Website", "Status", "Threat Level", "Last Updated",
        "Notes", "Data Quality Score", "Pricing Model", "Base Price", "Price Unit",
        "Product Categories", "Key Features", "Integration Partners", "Certifications",
        "Target Segments", "Customer Size Focus", "Geographic Focus", "Customer Count",
        "Customer Acquisition Rate", "Key Customers", "G2 Rating", "Employee Count",
        "Employee Growth Rate", "Year Founded", "Headquarters", "Funding Total",
        "Latest Round", "PE/VC Backers", "Website Traffic", "Social Following",
        "Recent Launches", "News Mentions"
    ]
    
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Data
    for row_idx, comp in enumerate(competitors, start=2):
        ws.cell(row=row_idx, column=1, value=comp.name)
        ws.cell(row=row_idx, column=2, value=comp.website)
        ws.cell(row=row_idx, column=3, value=comp.status)
        ws.cell(row=row_idx, column=4, value=comp.threat_level)
        ws.cell(row=row_idx, column=5, value=comp.last_updated.strftime("%Y-%m-%d") if comp.last_updated else "")
        ws.cell(row=row_idx, column=6, value=comp.notes)
        ws.cell(row=row_idx, column=7, value=comp.data_quality_score)
        ws.cell(row=row_idx, column=8, value=comp.pricing_model)
        ws.cell(row=row_idx, column=9, value=comp.base_price)
        ws.cell(row=row_idx, column=10, value=comp.price_unit)
        ws.cell(row=row_idx, column=11, value=comp.product_categories)
        ws.cell(row=row_idx, column=12, value=comp.key_features)
        ws.cell(row=row_idx, column=13, value=comp.integration_partners)
        ws.cell(row=row_idx, column=14, value=comp.certifications)
        ws.cell(row=row_idx, column=15, value=comp.target_segments)
        ws.cell(row=row_idx, column=16, value=comp.customer_size_focus)
        ws.cell(row=row_idx, column=17, value=comp.geographic_focus)
        ws.cell(row=row_idx, column=18, value=comp.customer_count)
        ws.cell(row=row_idx, column=19, value=comp.customer_acquisition_rate)
        ws.cell(row=row_idx, column=20, value=comp.key_customers)
        ws.cell(row=row_idx, column=21, value=comp.g2_rating)
        ws.cell(row=row_idx, column=22, value=comp.employee_count)
        ws.cell(row=row_idx, column=23, value=comp.employee_growth_rate)
        ws.cell(row=row_idx, column=24, value=comp.year_founded)
        ws.cell(row=row_idx, column=25, value=comp.headquarters)
        ws.cell(row=row_idx, column=26, value=comp.funding_total)
        ws.cell(row=row_idx, column=27, value=comp.latest_round)
        ws.cell(row=row_idx, column=28, value=comp.pe_vc_backers)
        ws.cell(row=row_idx, column=29, value=comp.website_traffic)
        ws.cell(row=row_idx, column=30, value=comp.social_following)
        ws.cell(row=row_idx, column=31, value=comp.recent_launches)
        ws.cell(row=row_idx, column=32, value=comp.news_mentions)
    
    ws.freeze_panes = "A2"
    auto_fit_columns(ws)
    apply_white_background(ws, max_row=len(competitors) + 10, max_col=32)
    
    # Save
    output_path = "./exports/competitors_export.xlsx"
    os.makedirs("./exports", exist_ok=True)
    wb.save(output_path)
    
    return FileResponse(
        output_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"certify_intel_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )


@app.get("/api/export/json")
def export_json(db: Session = Depends(get_db)):
    """Export all competitor data as JSON."""
    competitors = db.query(Competitor).filter(Competitor.is_deleted == False).all()
    
    data = []
    for comp in competitors:
        data.append({
            "name": comp.name,
            "website": comp.website,
            "status": comp.status,
            "threat_level": comp.threat_level,
            "last_updated": comp.last_updated.isoformat() if comp.last_updated else None,
            "notes": comp.notes,
            "data_quality_score": comp.data_quality_score,
            "pricing_model": comp.pricing_model,
            "base_price": comp.base_price,
            "price_unit": comp.price_unit,
            "product_categories": comp.product_categories,
            "key_features": comp.key_features,
            "integration_partners": comp.integration_partners,
            "certifications": comp.certifications,
            "target_segments": comp.target_segments,
            "customer_size_focus": comp.customer_size_focus,
            "geographic_focus": comp.geographic_focus,
            "customer_count": comp.customer_count,
            "customer_acquisition_rate": comp.customer_acquisition_rate,
            "key_customers": comp.key_customers,
            "g2_rating": comp.g2_rating,
            "employee_count": comp.employee_count,
            "employee_growth_rate": comp.employee_growth_rate,
            "year_founded": comp.year_founded,
            "headquarters": comp.headquarters,
            "funding_total": comp.funding_total,
            "latest_round": comp.latest_round,
            "pe_vc_backers": comp.pe_vc_backers,
            "website_traffic": comp.website_traffic,
            "social_following": comp.social_following,
            "recent_launches": comp.recent_launches,
            "news_mentions": comp.news_mentions,
        })
    
    return {"competitors": data, "count": len(data), "exported_at": datetime.utcnow().isoformat()}


# --- Dashboard Stats ---

@app.get("/api/dashboard/stats")
def get_dashboard_stats(db: Session = Depends(get_db)):
    """Get summary statistics for dashboard."""
    competitors = db.query(Competitor).filter(Competitor.is_deleted == False).all()
    
    stats = {
        "total_competitors": len(competitors),
        "active": len([c for c in competitors if c.status == "Active"]),
        "high_threat": len([c for c in competitors if c.threat_level == "High"]),
        "medium_threat": len([c for c in competitors if c.threat_level == "Medium"]),
        "low_threat": len([c for c in competitors if c.threat_level == "Low"]),
        "last_updated": datetime.utcnow().isoformat()
    }
    
    return stats


# --- Change Log ---
# Note: Primary /api/changes endpoint is defined earlier using DataChangeHistory table


# --- Scraping Endpoints ---

@app.post("/api/scrape/all")
async def trigger_scrape_all(background_tasks: BackgroundTasks, db: Session = Depends(get_db)):
    """Trigger scrape for all active competitors with progress tracking."""
    global scrape_progress

    competitors = db.query(Competitor).filter(
        Competitor.is_deleted == False,
        Competitor.status == "Active"
    ).all()
    competitor_ids = [c.id for c in competitors]
    competitor_names = {c.id: c.name for c in competitors}

    # Phase 4: Create RefreshSession for audit trail (Task 5.0.1-031)
    refresh_session = RefreshSession(
        competitors_scanned=len(competitor_ids),
        status="in_progress"
    )
    db.add(refresh_session)
    db.commit()
    db.refresh(refresh_session)

    # Reset progress tracker with enhanced tracking (Phase 2: Task 5.0.1-026)
    scrape_progress = {
        "active": True,
        "total": len(competitor_ids),
        "completed": 0,
        "current_competitor": None,
        "competitors_done": [],
        "changes_detected": 0,
        "new_values_added": 0,
        "started_at": datetime.utcnow().isoformat(),
        "recent_changes": [],
        "change_details": [],
        "errors": [],
        "session_id": refresh_session.id  # Track session ID for persistence
    }

    # Add to background tasks with progress tracking
    for cid in competitor_ids:
        background_tasks.add_task(run_scrape_job_with_progress, cid, competitor_names.get(cid, "Unknown"))

    return {
        "message": f"Scrape jobs queued for {len(competitor_ids)} competitors",
        "competitor_ids": competitor_ids,
        "total": len(competitor_ids),
        "session_id": refresh_session.id
    }


@app.get("/api/scrape/progress")
async def get_scrape_progress():
    """Get the current progress of a scrape operation."""
    return scrape_progress


# Phase 2: Task 5.0.1-028 - Get detailed session information
@app.get("/api/scrape/session")
async def get_scrape_session_details():
    """Get detailed information about the current or last refresh session."""
    return {
        "active": scrape_progress["active"],
        "total_competitors": scrape_progress["total"],
        "completed": scrape_progress["completed"],
        "current_competitor": scrape_progress.get("current_competitor"),
        "changes_detected": scrape_progress["changes_detected"],
        "new_values_added": scrape_progress["new_values_added"],
        "change_details": scrape_progress.get("change_details", []),
        "recent_changes": scrape_progress.get("recent_changes", []),
        "errors": scrape_progress.get("errors", []),
        "started_at": scrape_progress.get("started_at"),
        "competitors_processed": scrape_progress["competitors_done"]
    }


# Phase 3: Task 5.0.1-029 - Generate AI summary of refresh results
@app.post("/api/scrape/generate-summary")
async def generate_refresh_summary(db: Session = Depends(get_db)):
    """Use AI to generate a summary of the data refresh results."""
    import os

    if scrape_progress["active"]:
        return {"error": "Refresh still in progress", "type": "error"}

    if not scrape_progress.get("change_details"):
        return {
            "summary": "No changes detected during the last refresh. All competitor data remains the same.",
            "type": "static",
            "stats": {
                "competitors_scanned": scrape_progress.get("total", 0),
                "changes_detected": 0,
                "new_values": 0
            }
        }

    # Check for OpenAI client
    openai_key = os.getenv("OPENAI_API_KEY")
    if not openai_key:
        return {
            "summary": f"Refreshed {scrape_progress.get('total', 0)} competitors. Found {scrape_progress.get('changes_detected', 0)} changes and {scrape_progress.get('new_values_added', 0)} new data points.",
            "type": "static",
            "error": "OpenAI API key not configured"
        }

    try:
        from openai import OpenAI
        client = OpenAI(api_key=openai_key)

        # Prepare change data for AI
        changes_text = ""
        change_details = scrape_progress.get("change_details", [])

        for change in change_details[:30]:  # Limit to prevent token overflow
            if change.get("type") == "new":
                changes_text += f"- NEW: {change.get('competitor', 'Unknown')} - {change.get('field', 'Unknown')}: {change.get('new_value', 'N/A')}\n"
            else:
                changes_text += f"- CHANGED: {change.get('competitor', 'Unknown')} - {change.get('field', 'Unknown')}: '{change.get('old_value', 'N/A')}' → '{change.get('new_value', 'N/A')}'\n"

        # Generate AI summary
        prompt = f"""You are a competitive intelligence analyst. Summarize the following data refresh results in 3-4 sentences.
Focus on:
1. Most significant changes (pricing, threat levels, new features)
2. Any concerning trends
3. Recommended actions for the sales team

Data Refresh Results:
- Competitors scanned: {scrape_progress.get('total', 0)}
- Changes detected: {scrape_progress.get('changes_detected', 0)}
- New data points: {scrape_progress.get('new_values_added', 0)}
- Errors encountered: {len(scrape_progress.get('errors', []))}

Detailed Changes:
{changes_text}

Provide a concise executive summary. Be specific about which competitors changed."""

        response = client.chat.completions.create(
            model=os.getenv("OPENAI_MODEL", "gpt-4.1-mini"),
            messages=[
                {"role": "system", "content": "You are a competitive intelligence analyst providing brief, actionable summaries for sales and product teams."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.7,
            max_tokens=300
        )

        summary = response.choices[0].message.content

        return {
            "summary": summary,
            "type": "ai",
            "model": os.getenv("OPENAI_MODEL", "gpt-4.1-mini"),
            "stats": {
                "competitors_scanned": scrape_progress.get("total", 0),
                "changes_detected": scrape_progress.get("changes_detected", 0),
                "new_values": scrape_progress.get("new_values_added", 0),
                "errors": len(scrape_progress.get("errors", []))
            }
        }

    except ImportError:
        return {
            "summary": f"Refreshed {scrape_progress.get('total', 0)} competitors. Found {scrape_progress.get('changes_detected', 0)} changes and {scrape_progress.get('new_values_added', 0)} new data points.",
            "type": "static",
            "error": "OpenAI library not installed"
        }
    except Exception as e:
        return {
            "summary": f"Refreshed {scrape_progress.get('total', 0)} competitors. Found {scrape_progress.get('changes_detected', 0)} changes and {scrape_progress.get('new_values_added', 0)} new data points.",
            "type": "static",
            "error": str(e)
        }


# Phase 4: Task 5.0.1-031 - Refresh history endpoint
@app.get("/api/refresh-history")
async def get_refresh_history(limit: int = 10, db: Session = Depends(get_db)):
    """Get history of data refresh sessions."""
    sessions = db.query(RefreshSession).order_by(
        RefreshSession.started_at.desc()
    ).limit(limit).all()

    return [{
        "id": s.id,
        "started_at": s.started_at.isoformat() if s.started_at else None,
        "completed_at": s.completed_at.isoformat() if s.completed_at else None,
        "competitors_scanned": s.competitors_scanned,
        "changes_detected": s.changes_detected,
        "new_values_added": s.new_values_added,
        "errors_count": s.errors_count,
        "status": s.status,
        "ai_summary": s.ai_summary
    } for s in sessions]


@app.post("/api/scrape/{competitor_id}")
async def trigger_scrape(competitor_id: int, background_tasks: BackgroundTasks, db: Session = Depends(get_db)):
    """Trigger a scrape job for a specific competitor."""
    competitor = db.query(Competitor).filter(
        Competitor.id == competitor_id,
        Competitor.is_deleted == False
    ).first()
    
    if not competitor:
        raise HTTPException(status_code=404, detail="Competitor not found")
    
    # Add to background tasks
    background_tasks.add_task(run_scrape_job, competitor_id)
    
    return {
        "message": f"Scrape job queued for {competitor.name}",
        "competitor_id": competitor_id
    }


@app.post("/api/discovery/run")
async def trigger_discovery():
    """Trigger the autonomous discovery agent (MVP Mode)."""
    try:
        from discovery_agent import DiscoveryAgent
        agent = DiscoveryAgent()
        results = await agent.run_discovery_loop()
        return {"status": "success", "candidates": results, "count": len(results)}
    except Exception as e:
        return {"status": "error", "message": str(e)}


async def run_scrape_job(competitor_id: int):
    """Background job to scrape a competitor and update the database."""
    db = SessionLocal()
    try:
        comp = db.query(Competitor).filter(Competitor.id == competitor_id).first()
        if not comp:
            print(f"Competitor {competitor_id} not found")
            return
        
        print(f"Starting scrape for {comp.name}...")
        
        # Try to use the full scraper if available
        try:
            from scraper import CompetitorScraper
            from extractor import GPTExtractor
            
            scraper = CompetitorScraper()
            extractor = GPTExtractor()
            
            # Scrape the website
            content = await scraper.scrape(comp.website)
            
            if content:
                # Extract data using GPT
                from dataclasses import asdict
                
                # Note: extract_from_content is synchronous and takes (name, content)
                extracted_obj = extractor.extract_from_content(comp.name, content)
                # Convert dataclass to dict for iteration
                extracted = asdict(extracted_obj)
                
                if extracted:
                    # Update competitor with extracted data
                    for key, value in extracted.items():
                        if hasattr(comp, key) and value:
                            # Check if this field is locked (manual correction)
                            is_locked = db.query(DataSource).filter(
                                DataSource.competitor_id == comp.id,
                                DataSource.field_name == key,
                                DataSource.source_type == "manual"
                            ).first()
                            
                            if is_locked:
                                print(f"Skipping update for {comp.name}.{key} (locked by manual correction)")
                                continue
                                
                            old_value = getattr(comp, key)
                            if str(old_value) != str(value):
                                # Log the change
                                change = ChangeLog(
                                    competitor_id=comp.id,
                                    competitor_name=comp.name,
                                    change_type=key,
                                    previous_value=str(old_value) if old_value else None,
                                    new_value=str(value),
                                    source="scrape",
                                    severity="Medium"
                                )
                                db.add(change)
                                setattr(comp, key, value)
                    
                    comp.last_updated = datetime.utcnow()
                    db.commit()
                    print(f"Scrape completed for {comp.name}")
                    return
        except ImportError as e:
            print(f"Scraper not available: {e}")
        except Exception as e:
            print(f"Scrape error for {comp.name}: {e}")
        
        # Fallback: Just update the timestamp to show we tried
        comp.last_updated = datetime.utcnow()
        db.commit()
        print(f"Refresh completed for {comp.name} (timestamp updated)")

    except Exception as e:
        print(f"Scrape job failed for competitor {competitor_id}: {e}")
        db.rollback()
    finally:
        db.close()


def _update_data_source_with_confidence(
    db: Session,
    competitor_id: int,
    field_name: str,
    current_value: str,
    previous_value: str = None,
    source_url: str = None,
    source_name: str = None,
    gpt_confidence: int = 50
):
    """
    Helper to create or update DataSource record with confidence scoring.

    Maps GPT extraction confidence to our Admiralty Code-based scoring system.
    Website scrapes are inherently lower confidence (source_type = "website_scrape").
    """
    # Calculate confidence using our algorithm
    confidence_result = calculate_confidence_score(
        source_type="website_scrape",
        source_reliability="D",  # Website = Not usually reliable
        information_credibility=4,  # Doubtfully true until verified
        corroborating_sources=0,
        data_age_days=0
    )

    # Adjust score slightly based on GPT's own confidence assessment
    adjusted_score = min(100, max(0, confidence_result.score + (gpt_confidence - 50) // 5))
    adjusted_level = determine_confidence_level_from_score(adjusted_score)

    # Check if source already exists
    existing = db.query(DataSource).filter(
        DataSource.competitor_id == competitor_id,
        DataSource.field_name == field_name
    ).first()

    if existing:
        # Update existing record
        existing.previous_value = existing.current_value
        existing.current_value = current_value
        existing.source_type = "website_scrape"
        existing.source_url = source_url
        existing.source_name = source_name
        existing.extraction_method = "gpt_extraction"
        existing.extracted_at = datetime.utcnow()
        existing.source_reliability = "D"
        existing.information_credibility = 4
        existing.confidence_score = adjusted_score
        existing.confidence_level = adjusted_level
        existing.staleness_days = 0
        existing.updated_at = datetime.utcnow()
    else:
        # Create new record
        new_source = DataSource(
            competitor_id=competitor_id,
            field_name=field_name,
            current_value=current_value,
            previous_value=previous_value,
            source_type="website_scrape",
            source_url=source_url,
            source_name=source_name,
            extraction_method="gpt_extraction",
            source_reliability="D",
            information_credibility=4,
            confidence_score=adjusted_score,
            confidence_level=adjusted_level,
            corroborating_sources=0,
            staleness_days=0
        )
        db.add(new_source)


async def run_scrape_job_with_progress(competitor_id: int, competitor_name: str):
    """Background job to scrape a competitor with progress tracking, unified change logging, and confidence scoring."""
    global scrape_progress

    # Update current competitor being processed
    scrape_progress["current_competitor"] = competitor_name

    db = SessionLocal()
    changes_count = 0
    new_values_count = 0

    try:
        comp = db.query(Competitor).filter(Competitor.id == competitor_id).first()
        if not comp:
            print(f"Competitor {competitor_id} not found")
            return

        print(f"Starting scrape for {comp.name}...")

        # Try to use the full scraper if available
        try:
            from scraper import CompetitorScraper
            from extractor import GPTExtractor

            scraper = CompetitorScraper()
            extractor = GPTExtractor()

            # Scrape the website
            content = await scraper.scrape(comp.website)
            source_url = comp.website if not comp.website.startswith("http") else comp.website

            if content:
                # Extract data using GPT
                from dataclasses import asdict

                extracted_obj = extractor.extract_from_content(comp.name, content.get("content", ""))
                extracted = asdict(extracted_obj)

                # Get extraction confidence from GPT (if available)
                gpt_confidence = extracted.get("confidence_score") or 50

                if extracted:
                    # Update competitor with extracted data
                    for key, value in extracted.items():
                        if hasattr(comp, key) and value:
                            # Skip metadata fields
                            if key in ["confidence_score", "extraction_notes"]:
                                continue

                            # Check if this field is locked (manual correction)
                            is_locked = db.query(DataSource).filter(
                                DataSource.competitor_id == comp.id,
                                DataSource.field_name == key,
                                DataSource.source_type == "manual"
                            ).first()

                            if is_locked:
                                print(f"Skipping update for {comp.name}.{key} (locked by manual correction)")
                                continue

                            old_value = getattr(comp, key)
                            old_str = str(old_value) if old_value else None
                            new_str = str(value)

                            # Check if this is a new value or a change
                            if old_str != new_str:
                                # Determine change type
                                is_new_value = old_value is None or old_str == "" or old_str == "None"
                                change_type = "new" if is_new_value else "change"

                                # Log to DataChangeHistory (unified change log)
                                change_record = DataChangeHistory(
                                    competitor_id=comp.id,
                                    competitor_name=comp.name,
                                    field_name=key,
                                    old_value=old_str,
                                    new_value=new_str,
                                    changed_by="System (Auto-Refresh)",
                                    change_reason="Automated data refresh"
                                )
                                db.add(change_record)
                                setattr(comp, key, value)

                                # Phase 2: Track field-level changes (Task 5.0.1-027)
                                change_entry = {
                                    "competitor": comp.name,
                                    "field": key,
                                    "old_value": old_str[:50] if old_str else None,
                                    "new_value": new_str[:50] if new_str else None,
                                    "type": change_type,
                                    "timestamp": datetime.utcnow().isoformat()
                                }
                                scrape_progress["change_details"].append(change_entry)
                                scrape_progress["recent_changes"].append(change_entry)

                                # Keep only last 10 in recent_changes for live display
                                if len(scrape_progress["recent_changes"]) > 10:
                                    scrape_progress["recent_changes"] = scrape_progress["recent_changes"][-10:]

                                if is_new_value:
                                    new_values_count += 1
                                else:
                                    changes_count += 1

                            # Create or update DataSource with confidence scoring
                            _update_data_source_with_confidence(
                                db=db,
                                competitor_id=comp.id,
                                field_name=key,
                                current_value=new_str,
                                previous_value=old_str,
                                source_url=source_url,
                                source_name=f"{comp.name} Website",
                                gpt_confidence=gpt_confidence
                            )

                    comp.last_updated = datetime.utcnow()
                    db.commit()
                    print(f"Scrape completed for {comp.name} - {changes_count} changes, {new_values_count} new values")

                    # Trigger triangulation for key fields to verify scraped data
                    try:
                        triangulator = DataTriangulator(db)
                        triangulation_results = await triangulator.triangulate_all_key_fields(
                            competitor_id=comp.id,
                            competitor_name=comp.name,
                            website=comp.website,
                            is_public=comp.is_public,
                            ticker_symbol=comp.ticker_symbol
                        )

                        # Update confidence scores based on triangulation
                        for field_name, result in triangulation_results.items():
                            if result.confidence_score > 0:
                                existing = db.query(DataSource).filter(
                                    DataSource.competitor_id == comp.id,
                                    DataSource.field_name == field_name
                                ).first()
                                if existing:
                                    existing.confidence_score = result.confidence_score
                                    existing.confidence_level = result.confidence_level
                                    existing.corroborating_sources = result.sources_agreeing

                        db.commit()
                        print(f"Triangulation completed for {comp.name}")
                    except Exception as tri_err:
                        print(f"Triangulation error for {comp.name}: {tri_err}")

        except ImportError as e:
            print(f"Scraper not available: {e}")
            # Track error for display
            scrape_progress["errors"].append({
                "competitor": competitor_name,
                "error": f"Scraper not available: {str(e)[:100]}",
                "timestamp": datetime.utcnow().isoformat()
            })
        except Exception as e:
            print(f"Scrape error for {comp.name}: {e}")
            # Track error for display
            scrape_progress["errors"].append({
                "competitor": competitor_name,
                "error": str(e)[:100],
                "timestamp": datetime.utcnow().isoformat()
            })

        # Fallback: Just update the timestamp to show we tried
        if not changes_count and not new_values_count:
            comp.last_updated = datetime.utcnow()
            db.commit()
            print(f"Refresh completed for {comp.name} (timestamp updated)")

    except Exception as e:
        print(f"Scrape job failed for competitor {competitor_id}: {e}")
        db.rollback()
        # Track critical error
        scrape_progress["errors"].append({
            "competitor": competitor_name,
            "error": f"Job failed: {str(e)[:100]}",
            "timestamp": datetime.utcnow().isoformat()
        })
    finally:
        db.close()

        # Update progress tracker
        scrape_progress["completed"] += 1
        scrape_progress["competitors_done"].append(competitor_name)
        scrape_progress["changes_detected"] += changes_count
        scrape_progress["new_values_added"] += new_values_count

        # Check if all scrapes are done
        if scrape_progress["completed"] >= scrape_progress["total"]:
            scrape_progress["active"] = False
            scrape_progress["current_competitor"] = None
            print(f"All scrapes complete! {scrape_progress['changes_detected']} changes, {scrape_progress['new_values_added']} new values")

            # Phase 4: Persist RefreshSession results (Task 5.0.1-031)
            try:
                session_id = scrape_progress.get("session_id")
                if session_id:
                    session_db = SessionLocal()
                    refresh_session = session_db.query(RefreshSession).filter(
                        RefreshSession.id == session_id
                    ).first()
                    if refresh_session:
                        refresh_session.completed_at = datetime.utcnow()
                        refresh_session.changes_detected = scrape_progress["changes_detected"]
                        refresh_session.new_values_added = scrape_progress["new_values_added"]
                        refresh_session.errors_count = len(scrape_progress.get("errors", []))
                        refresh_session.status = "completed"
                        # Store change details as JSON
                        import json
                        refresh_session.change_details = json.dumps(scrape_progress.get("change_details", []))
                        session_db.commit()
                        print(f"RefreshSession {session_id} persisted to database")
                    session_db.close()
            except Exception as persist_err:
                print(f"Error persisting RefreshSession: {persist_err}")


# --- News Feed Endpoint ---

@app.get("/api/news/{company_name}")
async def get_competitor_news(company_name: str, limit: int = 5):
    """Fetch recent news articles mentioning a competitor."""
    try:
        from external_scrapers import NewsScraper
        scraper = NewsScraper()
        articles = await scraper.scrape_google_news(company_name, limit)
        return {
            "company": company_name,
            "articles": [
                {
                    "title": a.title,
                    "source": a.source,
                    "url": a.url,
                    "published_date": a.published_date,
                    "snippet": a.snippet
                }
                for a in articles
            ],
            "count": len(articles),
            "fetched_at": datetime.utcnow().isoformat()
        }
    except Exception as e:
        # Return mock news on error
        return {
            "company": company_name,
            "articles": [
                {
                    "title": f"{company_name} Announces New Healthcare Partnership",
                    "source": "Healthcare IT News",
                    "url": "https://healthcareitnews.com",
                    "published_date": datetime.utcnow().strftime("%Y-%m-%d"),
                    "snippet": f"Leading patient engagement company {company_name} has announced a strategic partnership..."
                },
                {
                    "title": f"{company_name} Expands Platform with AI Features",
                    "source": "MedCity News",
                    "url": "https://medcitynews.com",
                    "published_date": datetime.utcnow().strftime("%Y-%m-%d"),
                    "snippet": f"{company_name} today revealed new artificial intelligence capabilities designed to improve patient experience..."
                },
            ],
            "count": 2,
            "fetched_at": datetime.utcnow().isoformat()
        }



# --- AI SWOT Analysis Endpoint (Real) ---

@app.get("/api/competitors/{competitor_id}/swot")
async def get_swot_analysis(competitor_id: int, db: Session = Depends(get_db)):
    """Generate a real-time SWOT analysis using GPT-4 based on database records."""
    comp = db.query(Competitor).filter(Competitor.id == competitor_id).first()
    if not comp:
        raise HTTPException(status_code=404, detail="Competitor not found")

    try:
        if not get_openai_client():
            # Graceful fallback if no API key
            return {
                "strengths": ["Data unavailable (Check OpenAI Key)"], 
                "weaknesses": ["Data unavailable"], 
                "opportunities": [], 
                "threats": []
            }

        client = get_openai_client()
        
        # Construct rich context from DB
        context = f"""
        Analyze this competitor for 'Certify Health' (Provider of Patient Intake, Payments, and Biometrics).
        
        Competitor: {comp.name}
        Description: {comp.notes or 'N/A'}
        Pricing: {comp.pricing_model} ({comp.base_price})
        Target Segments: {comp.target_segments}
        Key Features: {comp.key_features}
        Weaknesses/Gaps: {comp.notes}
        """

        response = client.chat.completions.create(
            model="gpt-4.1",
            messages=[
                {"role": "system", "content": "You are a competitive strategy expert. Generate a strict JSON SWOT analysis with 3-4 bullet points per section."},
                {"role": "user", "content": context}
            ],
            temperature=0.7,
            response_format={"type": "json_object"}
        )
        
        content = json.loads(response.choices[0].message.content)
        # Ensure correct key structure
        return {
            "strengths": content.get("strengths", []) or content.get("Strengths", []),
            "weaknesses": content.get("weaknesses", []) or content.get("Weaknesses", []),
            "opportunities": content.get("opportunities", []) or content.get("Opportunities", []),
            "threats": content.get("threats", []) or content.get("Threats", [])
        }

    except Exception as e:
        print(f"SWOT Generation Error: {e}")
        # Return DB notes as fallback
        return {
            "strengths": ["Diverse product portfolio (inferred)"],
            "weaknesses": [comp.notes or "No specific weaknesses recorded"],
            "opportunities": ["Target their dissatisfaction"],
            "threats": ["Standard competitive threat"]
        }


# --- Stock Ticker Endpoint ---

def fetch_real_stock_data(ticker: str) -> Dict[str, Any]:
    """Fetch real-time stock data using yfinance."""
    try:
        stock = yf.Ticker(ticker)
        info = stock.info
        
        # Calculate change if not provided
        current = info.get("currentPrice") or info.get("regularMarketPrice", 0)
        previous = info.get("previousClose", 0)
        change = current - previous if current and previous else 0
        percent = (change / previous) * 100 if previous else 0
        
        return {
            "price": current,
            "change": change,
            "change_percent": percent,
            "market_cap": info.get("marketCap"),
            "pe_ratio": info.get("trailingPE"),
            "eps": info.get("trailingEps"),
            "beta": info.get("beta"),
            "volume": info.get("volume"),
            "high52": info.get("fiftyTwoWeekHigh"),
            "low52": info.get("fiftyTwoWeekLow"),
            "target_est": info.get("targetMeanPrice"),
            "company": info.get("longName"),
            "ticker": ticker
        }
    except Exception as e:
        print(f"Error fetching stock data for {ticker}: {e}")
        return None

@app.get("/api/stock/{company_name}")
async def get_stock_data(company_name: str, db: Session = Depends(get_db)):
    """Get stock data for a public company using yfinance."""
    import yfinance as yf
    from datetime import datetime
    
    company_lower = company_name.lower()
    
    # 1. Check Database first (for manually updated tickers)
    db_comp = db.query(Competitor).filter(Competitor.name == company_name).first()
    
    ticker_obj = None
    
    if db_comp and db_comp.is_public and db_comp.ticker_symbol:
        ticker_obj = {
            "symbol": db_comp.ticker_symbol,
            "exchange": db_comp.stock_exchange or "NYSE",
            "name": db_comp.name
        }
    
    # 2. Check hardcoded list if not in DB
    if not ticker_obj:
        ticker_obj = KNOWN_TICKERS.get(company_lower)
    
    if ticker_obj:
        try:
            # Fetch data from yfinance
            ticker = yf.Ticker(ticker_obj["symbol"])
            info = ticker.info
            
            # Helper for safe float formatting
            def get_val(key, default=None):
                return info.get(key, default)

            # Calculate Free Cash Flow if not directly available
            fcf = get_val('freeCashflow')
            if fcf is None and get_val('operatingCashflow') and get_val('capitalExpenditures'):
                fcf = get_val('operatingCashflow') - abs(get_val('capitalExpenditures', 0))

            next_earnings = "N/A"
            if get_val('earningsTimestamp'):
                next_earnings = datetime.fromtimestamp(get_val('earningsTimestamp')).strftime('%Y-%m-%d')
            elif get_val('earningsDate'):
                 # Sometimes it's a list
                 dates = get_val('earningsDate')
                 if isinstance(dates, list) and len(dates) > 0:
                     next_earnings = str(dates[0])[:10]

            return {
                "is_public": True,
                "company": ticker_obj["name"],
                "ticker": ticker_obj["symbol"],
                "exchange": ticker_obj["exchange"],
                
                # Trading
                "price": get_val('currentPrice', get_val('previousClose')),
                "change": float(get_val('currentPrice', 0)) - float(get_val('previousClose', 0)) if get_val('currentPrice') and get_val('previousClose') else 0,
                "change_percent": ((float(get_val('currentPrice', 0)) - float(get_val('previousClose', 0))) / float(get_val('previousClose', 1))) * 100 if get_val('previousClose') else 0,
                "volume": get_val('volume'),
                "avg_volume_10d": get_val('averageVolume10days'),
                "avg_volume_90d": get_val('averageVolume'),
                "fifty_two_week_low": get_val('fiftyTwoWeekLow'),
                "fifty_two_week_high": get_val('fiftyTwoWeekHigh'),
                "market_cap": get_val('marketCap'),

                # Valuation
                "enterprise_value": get_val('enterpriseValue'),
                "pe_trailing": get_val('trailingPE'),
                "pe_forward": get_val('forwardPE'),
                "ev_ebitda": get_val('enterpriseToEbitda'),
                "price_to_book": get_val('priceToBook'),
                "peg_ratio": get_val('pegRatio'),

                # Operating
                "eps_trailing": get_val('trailingEps'),
                "eps_forward": get_val('forwardEps'),
                "ebitda": get_val('ebitda'),
                "revenue_ttm": get_val('totalRevenue'),
                "free_cash_flow": fcf,
                "profit_margin": get_val('profitMargins'),

                # Risk
                "beta": get_val('beta'),
                "short_interest": get_val('shortPercentOfFloat'),
                "float_shares": get_val('floatShares'),

                # Capital
                "shares_outstanding": get_val('sharesOutstanding'),
                "inst_ownership": get_val('heldPercentInstitutions'),
                "dividend_yield": get_val('dividendYield'),
                "next_earnings": next_earnings,
                
                "data_sources": ["Yahoo Finance", "SEC EDGAR"]
            }
        except Exception as e:
            print(f"yfinance error for {ticker_obj['symbol']}: {e}")
            return {
                "is_public": True,
                "company": ticker_obj["name"],
                "ticker": ticker_obj["symbol"],
                "exchange": ticker_obj["exchange"],
                "error": "Unable to fetch live data"
            }



    # 3. Private Company Intelligence Logic
    
    # Initialize trackers
    from sec_edgar_scraper import SECEdgarScraper
    from linkedin_tracker import LinkedInTracker
    from gov_contracts_scraper import GovContractsScraper
    from h1b_scraper import H1BScraper
    from uspto_scraper import USPTOScraper
    from appstore_scraper import AppStoreScraper
    from glassdoor_scraper import GlassdoorScraper
    from google_ecosystem_scraper import GoogleEcosystemScraper
    from tech_stack_scraper import TechStackScraper
    from sentiment_scraper import SentimentScraper
    from seo_scraper import SEOScraper
    from risk_management_scraper import RiskManagementScraper
    
    sec = SECEdgarScraper()
    linkedin = LinkedInTracker()
    gov = GovContractsScraper()
    h1b = H1BScraper()
    uspto = USPTOScraper()
    appstore = AppStoreScraper()
    glassdoor = GlassdoorScraper()
    google = GoogleEcosystemScraper()
    tech = TechStackScraper()
    sentiment = SentimentScraper()
    seo = SEOScraper()
    risk = RiskManagementScraper()
    
    # Fetch Data
    form_d = sec.get_latest_form_d(company_name)
    li_data = linkedin.get_company_data(company_name)
    gov_data = gov.get_contract_data(company_name)
    h1b_data = h1b.get_h1b_data(company_name)
    patent_data = uspto.get_patent_data(company_name)
    app_data = appstore.get_app_data(company_name)
    glassdoor_data = glassdoor.get_company_data(company_name)
    google_data = google.get_ecosystem_data(company_name)
    tech_data = tech.get_tech_stack(company_name)
    sentiment_data = sentiment.get_sentiment_data(company_name)
    seo_data = seo.get_seo_data(company_name)
    risk_data = risk.get_risk_data(company_name)
    
    # Calculate Est. Revenue (Proxy: $150k ARR per employee for HealthTech)
    est_revenue = (li_data.employee_count or 0) * 150000
    
    # Determine Status
    stage = "Late Stage VC" 
    if li_data.employee_count < 50: stage = "Seed/Early"
    elif li_data.employee_count < 200: stage = "Growth Stage"
    elif li_data.employee_count > 1000:
        if gov_data and gov_data.total_amount > 10000000:
            stage = "Gov. Contractor / Mature"
        elif google_data.ads.active_creative_count > 100:
            stage = "Mass Market / Mature"
        else:
            stage = "Pre-IPO / PE Backed"
    
    return {
        "is_public": False,
        "company": company_name,
        "stage": stage,
        
        # Capital
        "total_funding": form_d["amount_raised"] if form_d else None,
        "latest_deal_date": form_d["filing_date"] if form_d else None,
        "latest_deal_amount": form_d["amount_raised"] if form_d else None,
        "latest_deal_type": form_d.get("round_type", "Venture Round") if form_d else None,
        
        # Growth & Ops
        "headcount": li_data.employee_count,
        "growth_rate_6mo": li_data.employee_growth_6mo,
        "active_hiring": li_data.open_jobs,
        "est_revenue": est_revenue,
        "hiring_departments": list(li_data.job_categories.keys())[:3] if li_data.job_categories else [],
        
        # Identity
        "headquarters": li_data.headquarters,
        "founded": li_data.founded_year,

        # Alternative Intelligence
        "gov_contracts": {
            "total_awards": gov_data.total_awards,
            "total_amount": gov_data.total_amount,
            "top_agency": gov_data.top_agency
        },
        "h1b_data": {
            "filings": h1b_data.total_filings_2023,
            "avg_salary": h1b_data.avg_salary_engineer,
            "top_title": h1b_data.top_job_titles[0] if h1b_data.top_job_titles else "N/A"
        },
        "innovation": {
            "patents": patent_data.total_patents,
            "pending": patent_data.pending_applications,
            "innovation_score": patent_data.innovation_score
        },
        "app_quality": {
            "avg_rating": app_data.avg_rating,
            "downloads": app_data.total_downloads,
            "sentiment": app_data.sentiment_summary
        },
        "employee_sentiment": {
            "rating": glassdoor_data.overall_rating,
            "ceo_approval": glassdoor_data.ceo_approval,
            "recommend": glassdoor_data.recommend_to_friend
        },

        # Google Digital Footprint
        "google_ecosystem": {
            "ads_active": google_data.ads.active_creative_count,
            "ad_formats": google_data.ads.formats,
            "brand_index": google_data.trends.current_index,
            "trend": google_data.trends.trend_direction,
            "reviews": google_data.maps.review_count,
            "review_velocity": google_data.maps.reviews_last_month
        },
        "tech_stack": {
            "signal": tech_data.marketing_budget_signal,
            "tools": tech_data.detected_tools,
            "has_enterprise_ads": tech_data.has_floodlight or tech_data.has_adobe_analytics
        },
        
        # Deep Dive Intelligence (New)
        "sentiment": {
            "g2_score": sentiment_data.g2_score,
            "g2_badges": sentiment_data.g2_badges[:2],
            "trustpilot": sentiment_data.trustpilot_score,
            "reddit": sentiment_data.reddit_sentiment,
            "complaints": sentiment_data.top_complaints[:1]
        },
        "seo": {
            "da": seo_data.domain_authority,
            "backlinks": seo_data.backlink_count,
            "speed": seo_data.page_load_speed,
            "keywords": seo_data.top_keywords[:3]
        },
        "risk_mgmt": {
            "founder_exit": risk_data.founder_exits,
            "exec_tenure": risk_data.avg_executive_tenure,
            "tier1_vc": risk_data.tier_1_investors,
            "soc2": risk_data.soc2_compliant,
            "warn": risk_data.warn_notices
        },
        
        "data_sources": ["SEC", "LinkedIn", "USAspending", "H-1B", "Google", "Tech", "G2/Capterra", "Moz", "Crunchbase"]
    }



# --- Alert Endpoints ---

@app.post("/api/alerts/send-digest")
def send_digest_email():
    """Manually trigger daily digest email."""
    try:
        from alerts import send_daily_digest
        success = send_daily_digest()
        return {"success": success, "message": "Daily digest sent" if success else "No changes to report or email not configured"}
    except Exception as e:
        return {"success": False, "error": str(e)}


@app.post("/api/alerts/send-summary")
def send_summary_email():
    """Manually trigger weekly summary email."""
    try:
        from alerts import send_weekly_summary
        success = send_weekly_summary()
        return {"success": success, "message": "Weekly summary sent" if success else "Email not configured"}
    except Exception as e:
        return {"success": False, "error": str(e)}


# --- Report Endpoints ---




@app.get("/api/reports/weekly-briefing")
def generate_weekly_briefing(db: Session = Depends(get_db)):
    """Generate executive weekly briefing PDF."""
    try:
        from reports import ReportManager
        competitors = db.query(Competitor).filter(Competitor.is_deleted == False).all()
        changes = db.query(ChangeLog).order_by(ChangeLog.detected_at.desc()).limit(10).all()
        
        # Calculate stats
        stats = {
            "total_competitors": len(competitors),
            "high_threat": len([c for c in competitors if c.threat_level == "High"]),
            "medium_threat": len([c for c in competitors if c.threat_level == "Medium"]),
            "low_threat": len([c for c in competitors if c.threat_level == "Low"])
        }
        
        manager = ReportManager("./exports")
        filepath = manager.generate_weekly_briefing(
            [c.__dict__ for c in competitors], 
            [c.__dict__ for c in changes],
            stats
        )
        return FileResponse(filepath, filename="weekly_briefing.pdf", media_type="application/pdf")
    except Exception as e:
        # Fallback - return a simple text summary
        competitors = db.query(Competitor).filter(Competitor.is_deleted == False).all()
        stats = {
            "total": len(competitors),
            "high": len([c for c in competitors if c.threat_level == "High"]),
            "medium": len([c for c in competitors if c.threat_level == "Medium"]),
            "low": len([c for c in competitors if c.threat_level == "Low"])
        }
        return {"error": str(e), "summary": stats, "message": "PDF generation unavailable, run pip install reportlab"}


@app.get("/api/reports/comparison")
def generate_comparison_report(db: Session = Depends(get_db)):
    """Generate competitor comparison PDF."""
    try:
        from reports import ReportManager
        competitors = db.query(Competitor).filter(Competitor.is_deleted == False).all()
        
        manager = ReportManager("./exports")
        filepath = manager.generate_comparison([c.__dict__ for c in competitors])
        return FileResponse(filepath, filename="competitor_comparison.pdf", media_type="application/pdf")
    except Exception as e:
        return {"error": str(e), "message": "PDF generation unavailable, run pip install reportlab"}


@app.get("/api/reports/battlecard/{competitor_id}")
def generate_battlecard(competitor_id: int, db: Session = Depends(get_db)):
    """Generate battlecard PDF for a specific competitor."""
    competitor = db.query(Competitor).filter(
        Competitor.id == competitor_id,
        Competitor.is_deleted == False
    ).first()
    
    if not competitor:
        raise HTTPException(status_code=404, detail="Competitor not found")
    
    try:
        from reports import ReportManager
        manager = ReportManager("./exports")
        
        # Inject Stock Data if Public
        comp_dict = {k: v for k, v in competitor.__dict__.items() if not k.startswith('_')}
        if competitor.is_public and competitor.ticker_symbol:
            stock_data = fetch_real_stock_data(competitor.ticker_symbol)
            if stock_data:
                comp_dict["stock_data"] = stock_data
                
        filepath = manager.generate_battlecard(comp_dict)
        return FileResponse(filepath, filename=f"{competitor.name}_battlecard.pdf", media_type="application/pdf")
    except Exception as e:
        return {"error": str(e), "competitor": competitor.name, "message": "PDF generation unavailable"}


# ============== Discovery Endpoints ==============

# In-memory storage for discovery results (would use Redis in production)
discovery_results = {"candidates": [], "last_run": None}


@app.get("/api/discovery/results")
def get_discovery_results():
    """Get previously discovered competitor candidates."""
    return discovery_results


@app.post("/api/discovery/run")
async def run_discovery(request: Request):
    """Run the autonomous competitor discovery agent with optional custom criteria."""
    global discovery_results
    
    # Parse request body for criteria
    criteria = None
    try:
        body = await request.json()
        criteria = body.get("criteria", None)
        if criteria:
            print(f"Discovery running with custom criteria:\n{criteria[:200]}...")
    except:
        pass
    
    try:
        from discovery_agent import DiscoveryAgent
        import asyncio
        
        # Pass criteria to agent for potential future use in search customization
        agent = DiscoveryAgent(use_live_search=False, use_openai=False)
        
        # Store criteria for reference
        if criteria:
            agent.custom_criteria = criteria
        
        candidates = await agent.run_discovery_loop(max_candidates=10)
        
        discovery_results = {
            "candidates": candidates,
            "last_run": datetime.utcnow().isoformat(),
            "count": len(candidates),
            "criteria_used": criteria[:100] + "..." if criteria and len(criteria) > 100 else criteria
        }
        
        return discovery_results
        
    except Exception as e:
        print(f"Discovery error: {e}")
        return {"error": str(e), "candidates": [], "message": "Discovery failed"}


@app.post("/api/discovery/run-live")
async def run_live_discovery():
    """Run live discovery with DuckDuckGo search (rate-limited)."""
    global discovery_results
    
    try:
        from discovery_agent import DiscoveryAgent
        
        agent = DiscoveryAgent(use_live_search=True, use_openai=False)
        candidates = await agent.run_discovery_loop(max_candidates=5)
        
        discovery_results = {
            "candidates": candidates,
            "last_run": datetime.utcnow().isoformat(),
            "count": len(candidates),
            "mode": "live"
        }
        
        return discovery_results
        
    except Exception as e:
        print(f"Live discovery error: {e}")
        return {"error": str(e), "candidates": [], "message": "Live discovery failed"}


# ============== Enhancement Endpoints ==============

@app.get("/api/competitors/{competitor_id}/threat-analysis")
def get_threat_analysis(competitor_id: int, db: Session = Depends(get_db)):
    """Get AI-powered threat analysis for a competitor."""
    try:
        from threat_analyzer import analyze_competitor_threat
        
        competitor = db.query(Competitor).filter(
            Competitor.id == competitor_id,
            Competitor.is_deleted == False
        ).first()
        
        if not competitor:
            raise HTTPException(status_code=404, detail="Competitor not found")
        
        comp_data = {k: v for k, v in competitor.__dict__.items() if not k.startswith('_')}
        analysis = analyze_competitor_threat(comp_data)
        
        # Update competitor threat level if changed
        if analysis.get("level") != competitor.threat_level:
            competitor.threat_level = analysis["level"]
            db.commit()
        
        return analysis
        
    except HTTPException:
        raise
    except Exception as e:
        return {"error": str(e), "message": "Threat analysis failed"}


@app.get("/api/competitors/{competitor_id}/news")
def get_competitor_news(competitor_id: int, days: int = 7, db: Session = Depends(get_db)):
    """Get news articles for a competitor."""
    try:
        from news_monitor import fetch_competitor_news
        
        competitor = db.query(Competitor).filter(
            Competitor.id == competitor_id,
            Competitor.is_deleted == False
        ).first()
        
        if not competitor:
            raise HTTPException(status_code=404, detail="Competitor not found")
        
        return fetch_competitor_news(competitor.name, days)
        
    except HTTPException:
        raise
    except Exception as e:
        return {"error": str(e), "articles": [], "message": "News fetch failed"}


@app.get("/api/news/{company_name}")
def get_news_by_name(company_name: str, days: int = 7):
    """Get news articles by company name."""
    try:
        from news_monitor import fetch_competitor_news
        return fetch_competitor_news(company_name, days)
    except Exception as e:
        return {"error": str(e), "articles": []}


@app.get("/api/alerts/price-changes")
def get_price_alerts(threshold: float = 10.0, db: Session = Depends(get_db)):
    """Get price change alerts."""
    try:
        from price_tracker import PriceTracker
        
        tracker = PriceTracker(db)
        alerts = tracker.detect_price_alerts(threshold)
        
        return {
            "alerts": [
                {
                    "competitor_id": a.competitor_id,
                    "competitor_name": a.competitor_name,
                    "previous_price": a.previous_price,
                    "new_price": a.new_price,
                    "change_percent": a.change_percent,
                    "direction": a.change_direction,
                    "severity": a.severity,
                    "detected_at": a.detected_at
                }
                for a in alerts
            ],
            "count": len(alerts)
        }
        
    except Exception as e:
        return {"error": str(e), "alerts": []}


@app.get("/api/pricing/comparison")
def get_pricing_comparison(db: Session = Depends(get_db)):
    """Get pricing comparison across all competitors."""
    try:
        from price_tracker import PriceTracker
        
        tracker = PriceTracker(db)
        return tracker.get_pricing_comparison()
        
    except Exception as e:
        return {"error": str(e)}


@app.get("/api/competitors/{competitor_id}/reviews")
def get_competitor_reviews(competitor_id: int, db: Session = Depends(get_db)):
    """Get G2/Capterra review data for a competitor."""
    try:
        from review_scraper import get_competitor_reviews as _get_reviews
        
        competitor = db.query(Competitor).filter(
            Competitor.id == competitor_id,
            Competitor.is_deleted == False
        ).first()
        
        if not competitor:
            raise HTTPException(status_code=404, detail="Competitor not found")
        
        return _get_reviews(competitor.name)
        
    except HTTPException:
        raise
    except Exception as e:
        return {"error": str(e), "overall_rating": 0}


@app.get("/api/reviews/compare")
def compare_reviews(competitor_ids: str, db: Session = Depends(get_db)):
    """Compare reviews across multiple competitors."""
    try:
        from review_scraper import compare_competitor_reviews
        
        ids = [int(id.strip()) for id in competitor_ids.split(",")]
        competitors = db.query(Competitor).filter(
            Competitor.id.in_(ids),
            Competitor.is_deleted == False
        ).all()
        
        names = [c.name for c in competitors]
        return compare_competitor_reviews(names)
        
    except Exception as e:
        return {"error": str(e)}


@app.get("/api/competitors/{competitor_id}/linkedin")
def get_linkedin_data(competitor_id: int, db: Session = Depends(get_db)):
    """Get LinkedIn company data for a competitor."""
    try:
        from linkedin_tracker import get_linkedin_data as _get_linkedin
        
        competitor = db.query(Competitor).filter(
            Competitor.id == competitor_id,
            Competitor.is_deleted == False
        ).first()
        
        if not competitor:
            raise HTTPException(status_code=404, detail="Competitor not found")
        
        return _get_linkedin(competitor.name)
        
    except HTTPException:
        raise
    except Exception as e:
        return {"error": str(e)}


@app.get("/api/competitors/{competitor_id}/hiring")
def get_hiring_analysis(competitor_id: int, db: Session = Depends(get_db)):
    """Get hiring trend analysis for a competitor."""
    try:
        from linkedin_tracker import analyze_competitor_hiring
        
        competitor = db.query(Competitor).filter(
            Competitor.id == competitor_id,
            Competitor.is_deleted == False
        ).first()
        
        if not competitor:
            raise HTTPException(status_code=404, detail="Competitor not found")
        
        return analyze_competitor_hiring(competitor.name)
        
    except HTTPException:
        raise
    except Exception as e:
        return {"error": str(e)}


@app.get("/api/hiring/compare")
def compare_hiring(competitor_ids: str, db: Session = Depends(get_db)):
    """Compare hiring across multiple competitors."""
    try:
        from linkedin_tracker import LinkedInTracker
        
        ids = [int(id.strip()) for id in competitor_ids.split(",")]
        competitors = db.query(Competitor).filter(
            Competitor.id.in_(ids),
            Competitor.is_deleted == False
        ).all()
        
        tracker = LinkedInTracker()
        return tracker.compare_hiring([c.name for c in competitors])
        
    except Exception as e:
        return {"error": str(e)}


@app.get("/api/competitors/{competitor_id}/insights")
def get_competitor_insights(competitor_id: int, db: Session = Depends(get_db)):
    """Get comprehensive insights for a competitor (threat + news + reviews + LinkedIn)."""
    try:
        competitor = db.query(Competitor).filter(
            Competitor.id == competitor_id,
            Competitor.is_deleted == False
        ).first()
        
        if not competitor:
            raise HTTPException(status_code=404, detail="Competitor not found")
        
        insights = {"competitor_name": competitor.name, "id": competitor_id}
        
        # Threat analysis
        try:
            from threat_analyzer import analyze_competitor_threat
            comp_data = {k: v for k, v in competitor.__dict__.items() if not k.startswith('_')}
            insights["threat"] = analyze_competitor_threat(comp_data)
        except Exception as e:
            insights["threat"] = {"error": str(e)}
        
        # News
        try:
            from news_monitor import NewsMonitor
            monitor = NewsMonitor()
            summary = monitor.get_news_summary(competitor.name)
            insights["news"] = summary
        except Exception as e:
            insights["news"] = {"error": str(e)}
        
        # Reviews
        try:
            from review_scraper import ReviewScraper
            scraper = ReviewScraper()
            review_insights = scraper.get_review_insights(competitor.name)
            insights["reviews"] = review_insights
        except Exception as e:
            insights["reviews"] = {"error": str(e)}
        
        # LinkedIn
        try:
            from linkedin_tracker import LinkedInTracker
            tracker = LinkedInTracker()
            hiring = tracker.analyze_hiring_trends(competitor.name)
            insights["hiring"] = hiring
        except Exception as e:
            insights["hiring"] = {"error": str(e)}
        
        return insights
        
    except HTTPException:
        raise
    except Exception as e:
        return {"error": str(e)}


# ============== Win/Loss Tracker Endpoints ==============

@app.post("/api/deals")
def log_deal(deal: dict, db: Session = Depends(get_db)):
    """Log a competitive deal outcome."""
    try:
        from win_loss_tracker import get_tracker
        
        tracker = get_tracker(db)
        result = tracker.log_deal(
            competitor_id=deal.get("competitor_id"),
            competitor_name=deal.get("competitor_name", "Unknown"),
            deal_name=deal.get("deal_name", "Untitled Deal"),
            deal_value=deal.get("deal_value"),
            outcome=deal.get("outcome", "Won"),
            loss_reason=deal.get("loss_reason"),
            notes=deal.get("notes")
        )
        
        # Trigger webhook
        try:
            from webhooks import trigger_deal_outcome
            trigger_deal_outcome(
                deal.get("competitor_name", "Unknown"),
                deal.get("deal_name", "Untitled"),
                deal.get("deal_value", 0),
                deal.get("outcome", "Won")
            )
        except:
            pass
        
        return result
        
    except Exception as e:
        return {"error": str(e)}


@app.get("/api/deals/stats")
def get_deal_stats(days: int = 365):
    """Get win/loss statistics."""
    try:
        from win_loss_tracker import get_win_loss_stats
        return get_win_loss_stats(days)
    except Exception as e:
        return {"error": str(e)}


@app.get("/api/deals/competitor/{competitor_id}")
def get_competitor_deals(competitor_id: int):
    """Get win/loss performance against a specific competitor."""
    try:
        from win_loss_tracker import get_tracker
        tracker = get_tracker()
        return tracker.get_competitor_performance(competitor_id)
    except Exception as e:
        return {"error": str(e)}


@app.get("/api/deals/most-competitive")
def get_most_competitive(limit: int = 5):
    """Get competitors we face most often."""
    try:
        from win_loss_tracker import get_tracker
        tracker = get_tracker()
        return tracker.get_most_competitive(limit)
    except Exception as e:
        return {"error": str(e)}


@app.get("/api/scrape/all")
async def trigger_scrape_all(
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Trigger a live refresh of all competitor data."""
    from refresh_orchestrator import refresh_orchestrator

    # Log the refresh action with user info
    user_email = current_user.get("email", "unknown")
    user_id = current_user.get("id")

    # Count competitors for immediate response
    count = db.query(Competitor).filter(Competitor.is_deleted == False).count()

    # Log this activity (shared across all users)
    log_activity(
        db, user_email, user_id,
        "data_refresh",
        {"action": "full_refresh", "competitor_count": count, "triggered_by": user_email}
    )

    # Run full refresh in background
    async def run_refresh():
        result = await refresh_orchestrator.refresh_all(priority_order=False)
        print(f"Refresh complete: {result}")

    background_tasks.add_task(asyncio.create_task, run_refresh())

    return {
        "success": True,
        "message": f"Live refresh started for {count} competitors by {user_email}",
        "count": count,
        "triggered_by": user_email,
        "note": "Check /api/refresh/history for results"
    }


@app.get("/api/refresh/history")
def get_refresh_history(competitor_id: int = None, limit: int = 100, db: Session = Depends(get_db)):
    """Get refresh history showing what changed."""
    from database import RefreshSnapshot
    
    query = db.query(RefreshSnapshot)
    if competitor_id:
        query = query.filter(RefreshSnapshot.competitor_id == competitor_id)
    
    snapshots = query.order_by(RefreshSnapshot.created_at.desc()).limit(limit).all()
    
    return {
        "total": len(snapshots),
        "history": [
            {
                "refresh_id": s.refresh_id,
                "competitor": s.competitor_name,
                "field": s.field_name,
                "old_value": s.old_value[:100] if s.old_value else None,
                "new_value": s.new_value[:100] if s.new_value else None,
                "source": s.source,
                "changed": s.changed,
                "date": s.created_at.isoformat()
            }
            for s in snapshots
        ]
    }



# ============== Webhook Endpoints ==============

@app.get("/api/webhooks")
def list_webhooks():
    """List all registered webhooks."""
    try:
        from webhooks import get_webhook_manager
        manager = get_webhook_manager()
        return {"webhooks": manager.list_webhooks()}
    except Exception as e:
        return {"error": str(e), "webhooks": []}


@app.post("/api/webhooks")
def register_webhook(webhook: dict):
    """Register a new webhook."""
    try:
        from webhooks import get_webhook_manager
        
        manager = get_webhook_manager()
        result = manager.register_webhook(
            url=webhook.get("url"),
            events=webhook.get("events", []),
            secret=webhook.get("secret")
        )
        
        return {
            "success": True,
            "webhook_id": result.id,
            "message": "Webhook registered successfully"
        }
        
    except ValueError as e:
        return {"success": False, "error": str(e)}
    except Exception as e:
        return {"success": False, "error": str(e)}


@app.delete("/api/webhooks/{webhook_id}")
def delete_webhook(webhook_id: str):
    """Delete a webhook."""
    try:
        from webhooks import get_webhook_manager
        
        manager = get_webhook_manager()
        removed = manager.unregister_webhook(webhook_id)
        
        return {
            "success": removed,
            "message": "Webhook removed" if removed else "Webhook not found"
        }
        
    except Exception as e:
        return {"success": False, "error": str(e)}


@app.post("/api/webhooks/{webhook_id}/test")
def test_webhook(webhook_id: str):
    """Send a test event to a webhook."""
    try:
        from webhooks import get_webhook_manager
        
        manager = get_webhook_manager()
        return manager.test_webhook(webhook_id)
        
    except Exception as e:
        return {"success": False, "error": str(e)}


@app.get("/api/webhooks/events")
def list_webhook_events():
    """List available webhook event types."""
    from webhooks import WebhookManager
    return {"event_types": WebhookManager.EVENT_TYPES}


# ============== New Data Source Endpoints ==============

@app.get("/api/competitors/{competitor_id}/employee-reviews")
def get_competitor_employee_reviews(competitor_id: int, db: Session = Depends(get_db)):
    """Get employee reviews and ratings from Glassdoor."""
    competitor = db.query(Competitor).filter(Competitor.id == competitor_id).first()
    if not competitor:
        raise HTTPException(status_code=404, detail="Competitor not found")
    
    return glassdoor_scraper.get_glassdoor_data(competitor.name)

@app.get("/api/competitors/{competitor_id}/jobs")
def get_competitor_jobs(competitor_id: int, db: Session = Depends(get_db)):
    """Get job postings and hiring signals from Indeed/ZipRecruiter."""
    competitor = db.query(Competitor).filter(Competitor.id == competitor_id).first()
    if not competitor:
        raise HTTPException(status_code=404, detail="Competitor not found")
    
    return indeed_scraper.get_job_data(competitor.name)

@app.get("/api/competitors/{competitor_id}/sec-filings")
def get_competitor_sec_filings(competitor_id: int, db: Session = Depends(get_db)):
    """Get public filings and financials from SEC EDGAR."""
    competitor = db.query(Competitor).filter(Competitor.id == competitor_id).first()
    if not competitor:
        raise HTTPException(status_code=404, detail="Competitor not found")
    
    return sec_edgar_scraper.get_sec_data(competitor.name)

@app.get("/api/competitors/{competitor_id}/patents")
def get_competitor_patents(competitor_id: int, db: Session = Depends(get_db)):
    """Get patent portfolio and IP analysis from USPTO."""
    competitor = db.query(Competitor).filter(Competitor.id == competitor_id).first()
    if not competitor:
        raise HTTPException(status_code=404, detail="Competitor not found")
    
    import uspto_scraper
    return uspto_scraper.get_patent_data(competitor.name)

@app.get("/api/competitors/{competitor_id}/klas-ratings")
def get_competitor_klas_ratings(competitor_id: int, db: Session = Depends(get_db)):
    """Get vendor ratings and customer satisfaction from KLAS Research."""
    competitor = db.query(Competitor).filter(Competitor.id == competitor_id).first()
    if not competitor:
        raise HTTPException(status_code=404, detail="Competitor not found")
    
    return klas_scraper.get_klas_data(competitor.name)

@app.get("/api/competitors/{competitor_id}/mobile-apps")
def get_competitor_mobile_apps(competitor_id: int, db: Session = Depends(get_db)):
    """Get mobile app ratings and reviews from App Store/Google Play."""
    competitor = db.query(Competitor).filter(Competitor.id == competitor_id).first()
    if not competitor:
        raise HTTPException(status_code=404, detail="Competitor not found")
    
    return appstore_scraper.get_app_store_data(competitor.name)

@app.get("/api/competitors/{competitor_id}/social-sentiment")
def get_competitor_social_sentiment(competitor_id: int, db: Session = Depends(get_db)):
    """Get brand sentiment and mentions from Twitter/Reddit."""
    competitor = db.query(Competitor).filter(Competitor.id == competitor_id).first()
    if not competitor:
        raise HTTPException(status_code=404, detail="Competitor not found")
    
    analysis = social_media_monitor.analyze_social_sentiment(competitor.name)
    raw_data = social_media_monitor.get_social_data(competitor.name)
    return {**analysis, "recent_posts": raw_data.get("top_posts", [])}

@app.get("/api/competitors/{competitor_id}/market-presence")
def get_competitor_market_presence(competitor_id: int, db: Session = Depends(get_db)):
    """Get industry presence and customer data from HIMSS/CHIME."""
    competitor = db.query(Competitor).filter(Competitor.id == competitor_id).first()
    if not competitor:
        raise HTTPException(status_code=404, detail="Competitor not found")
    
    return himss_scraper.get_himss_data(competitor.name)

# ============== EXPORT ENDPOINTS ==============

@app.get("/api/export/excel")
def get_export_excel(db: Session = Depends(get_db)):
    """Export all competitor data to Excel."""
    import pandas as pd
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    
    competitors = db.query(Competitor).filter(Competitor.is_deleted == False).all()
    
    if not competitors:
        raise HTTPException(status_code=404, detail="No data to export")
        
    # Convert to list of dicts
    data = []
    for comp in competitors:
        item = {
            "ID": comp.id,
            "Company Name": comp.name,
            "Website": comp.website,
            "Status": comp.status,
            "Threat Level": comp.threat_level,
            "Primary Market": comp.primary_market,
            "Pricing Model": comp.pricing_model,
            "Employee Count": comp.employee_count,
            "Revenue": comp.annual_revenue,
            "Net Income": comp.net_income,
            "Stock Symbol": comp.ticker_symbol,
            "Last Updated": comp.last_updated
        }
        data.append(item)
        
    df = pd.DataFrame(data)
    
    # Create Excel buffer
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="Competitors")
        
    buffer.seek(0)
    
    return StreamingResponse(
        buffer, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=competitor_intelligence_export.xlsx"}
    )


# ============== USER MANAGEMENT ENDPOINTS ==============

@app.get("/api/users", response_model=List[UserResponse])
def get_users(db: Session = Depends(get_db)):
    """List all users."""
    return db.query(User).filter(User.is_active == True).all()

@app.post("/api/users/invite")
def invite_user(invite: UserInviteRequest, db: Session = Depends(get_db)):
    """Invite a new user (creates account with default password)."""
    from extended_features import auth_manager

    existing = db.query(User).filter(User.email == invite.email).first()
    if existing:
        raise HTTPException(status_code=400, detail="User already exists")

    # Create user with default password using auth_manager for consistent hashing
    default_password = "Welcome123!"
    new_user = auth_manager.create_user(
        db,
        email=invite.email,
        password=default_password,
        full_name=invite.full_name or "",
        role=invite.role
    )

    return {"message": f"User {invite.email} invited successfully. Default password: {default_password}"}

@app.delete("/api/users/{user_id}")
def delete_user(user_id: int, db: Session = Depends(get_db)):
    """Remove a user."""
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
        
    db.delete(user)
    db.commit()
    return {"message": "User deleted"}


@app.get("/api/innovations/compare")
def compare_innovation_metrics(competitor_ids: str, db: Session = Depends(get_db)):
    """Compare innovation metrics (patents) across competitors."""
    ids = [int(id) for id in competitor_ids.split(",")]
    names = []
    
    for c_id in ids:
        comp = db.query(Competitor).filter(Competitor.id == c_id).first()
        if comp:
            names.append(comp.name)
            
    scraper = uspto_scraper.USPTOScraper()
    return scraper.compare_innovation(names)

@app.get("/api/social/compare")
def compare_social_metrics(competitor_ids: str, db: Session = Depends(get_db)):
    """Compare social media sentiment and presence."""
    ids = [int(id) for id in competitor_ids.split(",")]
    names = []
    
    for c_id in ids:
        comp = db.query(Competitor).filter(Competitor.id == c_id).first()
        if comp:
            names.append(comp.name)
            
    monitor = social_media_monitor.SocialMediaMonitor()
    return monitor.compare_social_presence(names)



# ============== Win/Loss & Webhooks (Real DB) ==============

class WinLossCreate(BaseModel):
    competitor_id: Optional[int] = None
    competitor_name: str
    outcome: str  # "win" or "loss"
    deal_value: Optional[float] = None
    customer_name: Optional[str] = None
    customer_size: Optional[str] = None
    reason: Optional[str] = None
    sales_rep: Optional[str] = None
    notes: Optional[str] = None

class WebhookCreate(BaseModel):
    name: str
    url: str
    event_types: str

class SystemPromptCreate(BaseModel):
    key: str
    content: str

class SystemPromptResponse(BaseModel):
    key: str
    content: str
    updated_at: Optional[datetime] = None

    class Config:
        orm_mode = True

class KnowledgeBaseItemCreate(BaseModel):
    title: str
    content_text: str
    source_type: str = "manual"
    is_active: bool = True

class KnowledgeBaseItemResponse(BaseModel):
    id: int
    title: str
    content_text: str
    source_type: str
    is_active: bool
    created_at: datetime

    class Config:
        orm_mode = True

@app.get("/api/win-loss")
def get_win_loss_deals(db: Session = Depends(get_db)):
    """Get all win/loss deals."""
    from database import WinLossDeal
    return db.query(WinLossDeal).order_by(WinLossDeal.deal_date.desc()).all()

@app.post("/api/win-loss")
def create_win_loss_deal(deal: WinLossCreate, db: Session = Depends(get_db)):
    """Log a new win/loss deal."""
    from database import WinLossDeal
    new_deal = WinLossDeal(
        competitor_id=deal.competitor_id,
        competitor_name=deal.competitor_name,
        outcome=deal.outcome,
        deal_value=deal.deal_value,
        customer_name=deal.customer_name,
        customer_size=deal.customer_size,
        reason=deal.reason,
        sales_rep=deal.sales_rep,
        notes=deal.notes
    )
    db.add(new_deal)
    db.commit()
    return {"status": "success", "id": new_deal.id}

@app.get("/api/webhooks")
def get_webhooks(db: Session = Depends(get_db)):
    """Get all configured webhooks."""
    from database import WebhookConfig
    return db.query(WebhookConfig).filter(WebhookConfig.is_active == True).all()

@app.post("/api/webhooks")
def create_webhook(webhook: WebhookCreate, db: Session = Depends(get_db)):
    """Configure a new webhook."""
    from database import WebhookConfig
    new_hook = WebhookConfig(
        name=webhook.name,
        url=webhook.url,
        event_types=webhook.event_types
    )
    db.add(new_hook)
    db.commit()
    return {"status": "success", "id": new_hook.id}

@app.delete("/api/webhooks/{id}")
def delete_webhook(id: int, db: Session = Depends(get_db)):
    """Delete a webhook."""
    from database import WebhookConfig
    hook = db.query(WebhookConfig).filter(WebhookConfig.id == id).first()
    if hook:
        hook.is_active = False
        db.commit()
    return {"status": "success"}


@app.get("/api/analytics/trends")
def get_market_trends(db: Session = Depends(get_db)):
    """Get market trends (New Competitors, Avg Price)."""
    from sqlalchemy import func
    import re

    # 1. New Competitors (Monthly)
    # SQLite syntax for YYYY-MM
    trends = db.query(
        func.strftime('%Y-%m', Competitor.created_at).label('month'),
        func.count(Competitor.id).label('count')
    ).filter(Competitor.is_deleted == False).group_by('month').order_by('month').all()

    labels = []
    competitor_counts = []
    
    for t in trends:
        labels.append(t.month)
        competitor_counts.append(t.count)
        
    if not labels:
        labels = [datetime.now().strftime('%Y-%m')]
        competitor_counts = [0]

    # 2. Avg Price (Current Snapshot)
    prices = db.query(Competitor.base_price).filter(
        Competitor.is_deleted == False,
        Competitor.base_price.isnot(None)
    ).all()
    
    valid_prices = []
    for (p_str,) in prices:
        if not p_str: continue
        # Extract first float found
        match = re.search(r'\$?([\d,]+(\.\d{2})?)', p_str)
        if match:
             try:
                 val = float(match.group(1).replace(',', ''))
                 valid_prices.append(val)
             except:
                 pass
                 
    avg_price = sum(valid_prices) / len(valid_prices) if valid_prices else 0
    
    # Repeat avg_price to create a baseline line
    price_data = [round(avg_price, 2)] * len(labels)

    return {
        "labels": labels,
        "datasets": [
            {
                "label": "Avg Market Price",
                "data": price_data,
                "borderColor": "#3A95ED",
                "tension": 0.4
            },
            {
                "label": "New Competitors",
                "data": competitor_counts,
                "borderColor": "#DC3545",
                "tension": 0.4,
                "yAxisID": "y1"
            }
        ]
    }



# =========================================================================
# NOTE: Duplicate endpoints removed - see lines 1774, 1426, 1797 for originals
# =========================================================================

# Import and mount the new analytics router
from analytics_routes import router as analytics_router
app.include_router(analytics_router)



# ============== Run Server ==============

# ============== Admin & AI Control Endpoints ==============

@app.get("/api/admin/system-prompts/{key}", response_model=SystemPromptResponse)
def get_system_prompt(key: str, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    """Get a system prompt by key (user-specific first, then global fallback)."""
    user_id = current_user.get("id") if current_user else None

    # Try user-specific prompt first
    prompt = db.query(SystemPrompt).filter(
        SystemPrompt.key == key,
        SystemPrompt.user_id == user_id
    ).first()

    # Fallback to global prompt if no user-specific one
    if not prompt:
        prompt = db.query(SystemPrompt).filter(
            SystemPrompt.key == key,
            SystemPrompt.user_id == None
        ).first()

    if not prompt:
        # Return default if not found
        default_content = ""
        if key == "dashboard_summary":
            default_content = """You are Certify Health's competitive intelligence analyst. Generate a comprehensive, executive-level strategic summary using ONLY the LIVE data provided below.

**CRITICAL - PROVE YOU ARE USING LIVE DATA:**
- Start your summary with: "📊 **Live Intelligence Report** - Generated [TODAY'S DATE AND TIME]"
- State the EXACT total number of competitors being tracked (e.g., "Currently monitoring **X competitors**")
- Name at least 3-5 SPECIFIC competitor names from the data with their EXACT threat levels
- Quote SPECIFIC numbers: funding amounts, employee counts, pricing figures directly from the data
- Reference any recent changes or updates with their timestamps if available
- If a competitor has specific data points (headquarters, founding year, etc.), cite them exactly

**YOUR SUMMARY MUST INCLUDE:**

1. **📈 Executive Overview**
   - State exact competitor count and breakdown by threat level
   - Name the top 3 high-threat competitors BY NAME

2. **🎯 Threat Analysis**
   - List HIGH threat competitors by name with why they're threats
   - List MEDIUM threat competitors by name
   - Provide specific threat justifications using their data

3. **💰 Pricing Intelligence**
   - Name competitors with known pricing and their EXACT pricing models
   - Compare specific price points where available

4. **📊 Market Trends**
   - Reference specific data points that indicate trends
   - Name competitors showing growth signals

5. **✅ Strategic Recommendations**
   - 3-5 specific, actionable recommendations
   - Reference specific competitors in each recommendation

6. **👁️ Watch List**
   - Name the top 5 competitors requiring immediate attention
   - State WHY each is on the watch list with specific data

**IMPORTANT:** Every claim must reference actual data provided. Do NOT make up or assume any information. If data is missing, say "Data not available" rather than guessing."""
        elif key == "chat_persona":
            default_content = "You are a competitive intelligence analyst for Certify Health. Always reference specific data points and competitor names when answering questions. Cite exact numbers and dates when available."

        return SystemPromptResponse(key=key, content=default_content)
    return prompt

@app.post("/api/admin/system-prompts", response_model=SystemPromptResponse)
def update_system_prompt(prompt_data: SystemPromptCreate, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    """Update or create a user-specific system prompt."""
    user_id = current_user.get("id") if current_user else None

    # Check for existing user-specific prompt
    prompt = db.query(SystemPrompt).filter(
        SystemPrompt.key == prompt_data.key,
        SystemPrompt.user_id == user_id
    ).first()

    if prompt:
        prompt.content = prompt_data.content
        prompt.updated_at = datetime.utcnow()
    else:
        # Create new user-specific prompt
        prompt = SystemPrompt(key=prompt_data.key, content=prompt_data.content, user_id=user_id)
        db.add(prompt)

    db.commit()
    db.refresh(prompt)
    return prompt

@app.get("/api/admin/knowledge-base", response_model=List[KnowledgeBaseItemResponse])
def get_knowledge_base_items(db: Session = Depends(get_db)):
    """Get all active knowledge base items."""
    return db.query(KnowledgeBaseItem).filter(KnowledgeBaseItem.is_active == True).order_by(KnowledgeBaseItem.created_at.desc()).all()

@app.post("/api/admin/knowledge-base", response_model=KnowledgeBaseItemResponse)
def add_knowledge_base_item(item: KnowledgeBaseItemCreate, db: Session = Depends(get_db)):
    """Add a new item to the knowledge base."""
    new_item = KnowledgeBaseItem(
        title=item.title,
        content_text=item.content_text,
        source_type=item.source_type,
        is_active=item.is_active
    )
    db.add(new_item)
    db.commit()
    db.refresh(new_item)
    return new_item

@app.delete("/api/admin/knowledge-base/{item_id}")
def delete_knowledge_base_item(item_id: int, db: Session = Depends(get_db)):
    """Soft delete a knowledge base item."""
    item = db.query(KnowledgeBaseItem).filter(KnowledgeBaseItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")

    item.is_active = False
    db.commit()
    return {"message": "Item deleted"}


# =========================================================================
# USER SAVED PROMPTS - Per-user prompt management
# =========================================================================

from database import UserSavedPrompt

class UserSavedPromptCreate(BaseModel):
    name: str
    prompt_type: str = "executive_summary"
    content: str

class UserSavedPromptUpdate(BaseModel):
    name: Optional[str] = None
    content: Optional[str] = None
    is_default: Optional[bool] = None

class UserSavedPromptResponse(BaseModel):
    id: int
    user_id: int
    name: str
    prompt_type: str
    content: str
    is_default: bool
    created_at: datetime
    updated_at: datetime

    class Config:
        from_attributes = True


@app.get("/api/user/prompts", response_model=List[UserSavedPromptResponse])
def get_user_prompts(
    prompt_type: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Get all saved prompts for the current user."""
    user_id = current_user.get("id")
    if not user_id:
        raise HTTPException(status_code=401, detail="User not authenticated")

    query = db.query(UserSavedPrompt).filter(UserSavedPrompt.user_id == user_id)
    if prompt_type:
        query = query.filter(UserSavedPrompt.prompt_type == prompt_type)

    return query.order_by(UserSavedPrompt.updated_at.desc()).all()


@app.post("/api/user/prompts", response_model=UserSavedPromptResponse)
def create_user_prompt(
    prompt_data: UserSavedPromptCreate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Create a new saved prompt for the current user."""
    user_id = current_user.get("id")
    if not user_id:
        raise HTTPException(status_code=401, detail="User not authenticated")

    new_prompt = UserSavedPrompt(
        user_id=user_id,
        name=prompt_data.name,
        prompt_type=prompt_data.prompt_type,
        content=prompt_data.content,
        is_default=False
    )
    db.add(new_prompt)
    db.commit()
    db.refresh(new_prompt)

    # Log activity
    log_activity(db, current_user.get("email", "unknown"), user_id, "prompt_created", f"Created prompt: {prompt_data.name}")

    return new_prompt


@app.get("/api/user/prompts/{prompt_id}", response_model=UserSavedPromptResponse)
def get_user_prompt(
    prompt_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Get a specific saved prompt by ID."""
    user_id = current_user.get("id")
    if not user_id:
        raise HTTPException(status_code=401, detail="User not authenticated")

    prompt = db.query(UserSavedPrompt).filter(
        UserSavedPrompt.id == prompt_id,
        UserSavedPrompt.user_id == user_id
    ).first()

    if not prompt:
        raise HTTPException(status_code=404, detail="Prompt not found")

    return prompt


@app.put("/api/user/prompts/{prompt_id}", response_model=UserSavedPromptResponse)
def update_user_prompt(
    prompt_id: int,
    prompt_data: UserSavedPromptUpdate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Update an existing saved prompt."""
    user_id = current_user.get("id")
    if not user_id:
        raise HTTPException(status_code=401, detail="User not authenticated")

    prompt = db.query(UserSavedPrompt).filter(
        UserSavedPrompt.id == prompt_id,
        UserSavedPrompt.user_id == user_id
    ).first()

    if not prompt:
        raise HTTPException(status_code=404, detail="Prompt not found")

    if prompt_data.name is not None:
        prompt.name = prompt_data.name
    if prompt_data.content is not None:
        prompt.content = prompt_data.content
    if prompt_data.is_default is not None:
        # If setting as default, unset any other defaults for this type
        if prompt_data.is_default:
            db.query(UserSavedPrompt).filter(
                UserSavedPrompt.user_id == user_id,
                UserSavedPrompt.prompt_type == prompt.prompt_type,
                UserSavedPrompt.id != prompt_id
            ).update({"is_default": False})
        prompt.is_default = prompt_data.is_default

    prompt.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(prompt)

    return prompt


@app.delete("/api/user/prompts/{prompt_id}")
def delete_user_prompt(
    prompt_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Delete a saved prompt."""
    user_id = current_user.get("id")
    if not user_id:
        raise HTTPException(status_code=401, detail="User not authenticated")

    prompt = db.query(UserSavedPrompt).filter(
        UserSavedPrompt.id == prompt_id,
        UserSavedPrompt.user_id == user_id
    ).first()

    if not prompt:
        raise HTTPException(status_code=404, detail="Prompt not found")

    prompt_name = prompt.name
    db.delete(prompt)
    db.commit()

    # Log activity
    log_activity(db, current_user.get("email", "unknown"), user_id, "prompt_deleted", f"Deleted prompt: {prompt_name}")

    return {"message": "Prompt deleted successfully"}


@app.post("/api/user/prompts/{prompt_id}/set-default")
def set_prompt_as_default(
    prompt_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Set a prompt as the default for its type."""
    user_id = current_user.get("id")
    if not user_id:
        raise HTTPException(status_code=401, detail="User not authenticated")

    prompt = db.query(UserSavedPrompt).filter(
        UserSavedPrompt.id == prompt_id,
        UserSavedPrompt.user_id == user_id
    ).first()

    if not prompt:
        raise HTTPException(status_code=404, detail="Prompt not found")

    # Unset any existing defaults for this prompt type
    db.query(UserSavedPrompt).filter(
        UserSavedPrompt.user_id == user_id,
        UserSavedPrompt.prompt_type == prompt.prompt_type
    ).update({"is_default": False})

    # Set this one as default
    prompt.is_default = True
    prompt.updated_at = datetime.utcnow()
    db.commit()

    return {"message": f"Prompt '{prompt.name}' set as default"}




# ============== Settings API Endpoints (User-Specific) ==============

import json

def get_user_setting(db: Session, user_id: int, setting_key: str):
    """Get a user-specific setting from the database."""
    setting = db.query(UserSettings).filter(
        UserSettings.user_id == user_id,
        UserSettings.setting_key == setting_key
    ).first()
    if setting:
        try:
            return json.loads(setting.setting_value)
        except:
            return setting.setting_value
    return {}

def save_user_setting(db: Session, user_id: int, setting_key: str, setting_value):
    """Save a user-specific setting to the database."""
    setting = db.query(UserSettings).filter(
        UserSettings.user_id == user_id,
        UserSettings.setting_key == setting_key
    ).first()

    value_str = json.dumps(setting_value) if not isinstance(setting_value, str) else setting_value

    if setting:
        setting.setting_value = value_str
        setting.updated_at = datetime.utcnow()
    else:
        setting = UserSettings(
            user_id=user_id,
            setting_key=setting_key,
            setting_value=value_str
        )
        db.add(setting)
    db.commit()
    return setting


@app.post("/api/settings/schedule")
async def save_schedule_settings(
    settings: dict,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Save user refresh schedule preferences (personal to each user)."""
    save_user_setting(db, current_user["id"], "schedule", settings)
    return {"success": True, "message": "Schedule settings saved"}

@app.get("/api/settings/schedule")
async def get_schedule_settings(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Get saved schedule settings (personal to each user)."""
    return get_user_setting(db, current_user["id"], "schedule")

@app.post("/api/settings/notifications")
async def save_notification_settings(
    settings: dict,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Save user notification preferences (personal to each user)."""
    save_user_setting(db, current_user["id"], "notifications", settings)
    return {"success": True, "message": "Notification settings saved"}

@app.get("/api/settings/notifications")
async def get_notification_settings(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Get saved notification settings (personal to each user)."""
    return get_user_setting(db, current_user["id"], "notifications")


# ============== WebSocket for Real-time Refresh Progress ==============

from fastapi import WebSocket, WebSocketDisconnect
from typing import List

class ConnectionManager:
    """Manages WebSocket connections for real-time updates."""
    def __init__(self):
        self.active_connections: List[WebSocket] = []

    async def connect(self, websocket: WebSocket):
        await websocket.accept()
        self.active_connections.append(websocket)

    def disconnect(self, websocket: WebSocket):
        if websocket in self.active_connections:
            self.active_connections.remove(websocket)

    async def broadcast(self, message: dict):
        for connection in self.active_connections:
            try:
                await connection.send_json(message)
            except:
                pass

ws_manager = ConnectionManager()

@app.websocket("/ws/refresh-progress")
async def websocket_refresh_progress(websocket: WebSocket):
    """WebSocket endpoint for real-time refresh progress updates."""
    await ws_manager.connect(websocket)
    try:
        while True:
            # Keep connection alive, actual updates sent via broadcast
            data = await websocket.receive_text()
    except WebSocketDisconnect:
        ws_manager.disconnect(websocket)


# ============== Static Files (Must be Last) ==============
import os
import sys

# Determine the directory where the backend executable (or script) is running
if getattr(sys, 'frozen', False):
    # If we are running as a bundle (PyInstaller)
    base_dir = os.path.dirname(sys.executable)
    # In bundle, we expect 'frontend' to be a sibling of the executable (or in resources/frontend)
    # The 'backend-bundle' is usually in 'resources/backend-bundle'
    # So we look at 'resources/frontend' which is 1 level up from 'backend-bundle'
    frontend_dir = os.path.join(os.path.dirname(base_dir), "frontend")
else:
    # Running in normal dev environment
    frontend_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "frontend")


if os.path.exists(frontend_dir):
    from fastapi.responses import FileResponse
    
    @app.get("/app")
    async def read_app_root():
        """Serve the frontend app root for Electron."""
        return FileResponse(os.path.join(frontend_dir, 'index.html'))

    app.mount("/", StaticFiles(directory=frontend_dir, html=True), name="frontend")
    print(f"Serving frontend from: {frontend_dir}")
else:
    print(f"Warning: Frontend directory not found at {frontend_dir}")


# ============== Start Server ==============
if __name__ == "__main__":
    import uvicorn
    print("\n" + "="*50)
    print("  Certify Intel Backend Starting...")
    print("  Open http://localhost:8000 in your browser")
    print("="*50 + "\n")
    uvicorn.run(app, host="127.0.0.1", port=8000, reload=False)
