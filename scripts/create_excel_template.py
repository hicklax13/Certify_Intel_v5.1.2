"""
Certify Intel - Excel Template Generator
Generates the base Excel template with all 32 data columns and proper formatting.
AUTO-FIT COLUMNS: All columns are sized to fit the widest content.
WHITE BACKGROUND: All cells have white fill (no empty/no-fill cells).
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# White fill for all data cells
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

def auto_fit_columns(worksheet):
    """Auto-fit all columns to the widest content."""
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        # Add padding and set minimum/maximum widths
        adjusted_width = max_length + 3
        if adjusted_width < 12:
            adjusted_width = 12
        if adjusted_width > 50:
            adjusted_width = 50
        worksheet.column_dimensions[column_letter].width = adjusted_width

def apply_white_background(worksheet, max_row=100, max_col=50):
    """Apply white background to all cells in the worksheet."""
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = worksheet.cell(row=row, column=col)
            # Only apply white fill if cell doesn't already have a colored fill (headers)
            if cell.fill.fgColor.rgb in (None, '00000000', 'FFFFFFFF') or cell.fill.fill_type is None:
                cell.fill = WHITE_FILL

def create_excel_template():
    """Create the Certify Intel Excel Dashboard template."""
    
    wb = Workbook()
    
    # ============== SHEET 1: Competitors ==============
    ws_competitors = wb.active
    ws_competitors.title = "Competitors"
    
    # Define all 32 columns with their categories
    columns = [
        # Core Info (1-7)
        ("Competitor Name", "Core"),
        ("Website", "Core"),
        ("Status", "Core"),
        ("Threat Level", "Core"),
        ("Last Updated", "Core"),
        ("Notes", "Core"),
        ("Data Quality Score", "Core"),
        
        # Pricing & Product (8-14)
        ("Pricing Model", "Pricing"),
        ("Base Price", "Pricing"),
        ("Price Unit", "Pricing"),
        ("Product Categories", "Product"),
        ("Key Features", "Product"),
        ("Integration Partners", "Product"),
        ("Certifications", "Product"),
        
        # Customer & Market (15-21)
        ("Target Segments", "Market"),
        ("Customer Size Focus", "Market"),
        ("Geographic Focus", "Market"),
        ("Customer Count", "Market"),
        ("Customer Acquisition Rate", "Market"),
        ("Key Customers", "Market"),
        ("G2 Rating", "Market"),
        
        # Company Health (22-28)
        ("Employee Count", "Company"),
        ("Employee Growth Rate", "Company"),
        ("Year Founded", "Company"),
        ("Headquarters", "Company"),
        ("Funding Total", "Company"),
        ("Latest Round", "Company"),
        ("PE/VC Backers", "Company"),
        
        # Digital Presence (29-32)
        ("Website Traffic (Monthly)", "Digital"),
        ("Social Following", "Digital"),
        ("Recent Product Launches", "Digital"),
        ("News Mentions (30d)", "Digital"),
    ]
    
    # Styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill_core = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_fill_pricing = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    header_fill_product = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
    header_fill_market = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
    header_fill_company = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    header_fill_digital = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")
    
    category_fills = {
        "Core": header_fill_core,
        "Pricing": header_fill_pricing,
        "Product": header_fill_product,
        "Market": header_fill_market,
        "Company": header_fill_company,
        "Digital": header_fill_digital,
    }
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col_idx, (col_name, category) in enumerate(columns, start=1):
        cell = ws_competitors.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = category_fills[category]
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cell.border = thin_border
    
    # Freeze header row
    ws_competitors.freeze_panes = "A2"
    
    # Data Validations
    # Status dropdown
    status_dv = DataValidation(
        type="list",
        formula1='"Active,Inactive,Acquired,Watch"',
        showDropDown=False
    )
    ws_competitors.add_data_validation(status_dv)
    status_dv.add(f"C2:C100")
    
    # Threat Level dropdown
    threat_dv = DataValidation(
        type="list",
        formula1='"High,Medium,Low,Watch"',
        showDropDown=False
    )
    ws_competitors.add_data_validation(threat_dv)
    threat_dv.add(f"D2:D100")
    
    # Pricing Model dropdown
    pricing_dv = DataValidation(
        type="list",
        formula1='"Per User,Per Provider,Per Location,Per Visit,Flat Rate,Tiered,Custom,Unknown"',
        showDropDown=False
    )
    ws_competitors.add_data_validation(pricing_dv)
    pricing_dv.add(f"H2:H100")
    
    # Customer Size Focus dropdown
    size_dv = DataValidation(
        type="list",
        formula1='"Solo,Small (1-15),Medium (15-50),Large (50+),Enterprise,All Sizes"',
        showDropDown=False
    )
    ws_competitors.add_data_validation(size_dv)
    size_dv.add(f"P2:P100")
    
    # Auto-fit columns
    auto_fit_columns(ws_competitors)
    
    # ============== SHEET 2: Dashboard ==============
    ws_dashboard = wb.create_sheet("Dashboard")
    
    # Dashboard header
    ws_dashboard["A1"] = "Certify Intel - Competitive Intelligence Dashboard"
    ws_dashboard["A1"].font = Font(bold=True, size=18, color="2F5496")
    ws_dashboard.merge_cells("A1:E1")
    
    ws_dashboard["A3"] = "Summary Metrics"
    ws_dashboard["A3"].font = Font(bold=True, size=14)
    
    # Metric labels
    metrics = [
        ("Total Competitors", "=COUNTA(Competitors!A:A)-1"),
        ("High Threat", '=COUNTIF(Competitors!D:D,"High")'),
        ("Medium Threat", '=COUNTIF(Competitors!D:D,"Medium")'),
        ("Low Threat", '=COUNTIF(Competitors!D:D,"Low")'),
        ("Active Competitors", '=COUNTIF(Competitors!C:C,"Active")'),
    ]
    
    for idx, (label, formula) in enumerate(metrics, start=4):
        ws_dashboard[f"A{idx}"] = label
        ws_dashboard[f"B{idx}"] = formula
        ws_dashboard[f"A{idx}"].font = Font(bold=True)
    
    auto_fit_columns(ws_dashboard)
    
    # ============== SHEET 3: Pricing Matrix ==============
    ws_pricing = wb.create_sheet("Pricing Matrix")
    ws_pricing["A1"] = "Pricing Comparison Matrix"
    ws_pricing["A1"].font = Font(bold=True, size=14)
    
    pricing_headers = ["Competitor", "Pricing Model", "Base Price", "Price Unit", "Notes"]
    for col_idx, header in enumerate(pricing_headers, start=1):
        cell = ws_pricing.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill_pricing
    
    auto_fit_columns(ws_pricing)
    
    # ============== SHEET 4: Feature Matrix ==============
    ws_features = wb.create_sheet("Feature Matrix")
    ws_features["A1"] = "Feature Comparison Matrix"
    ws_features["A1"].font = Font(bold=True, size=14)
    
    feature_headers = ["Competitor", "Patient Intake", "Insurance Verification", "Payments", "Scheduling", "Patient Portal", "Telehealth", "RCM"]
    for col_idx, header in enumerate(feature_headers, start=1):
        cell = ws_features.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill_product
    
    auto_fit_columns(ws_features)
    
    # ============== SHEET 5: Market Segments ==============
    ws_segments = wb.create_sheet("Market Segments")
    ws_segments["A1"] = "Market Segment Analysis"
    ws_segments["A1"].font = Font(bold=True, size=14)
    
    segment_headers = ["Competitor", "Target Segments", "Customer Size", "Geographic Focus", "Threat Level"]
    for col_idx, header in enumerate(segment_headers, start=1):
        cell = ws_segments.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill_market
    
    auto_fit_columns(ws_segments)
    
    # ============== SHEET 6: Change Log ==============
    ws_changelog = wb.create_sheet("Change Log")
    ws_changelog["A1"] = "Competitor Change Log"
    ws_changelog["A1"].font = Font(bold=True, size=14)
    
    log_headers = ["Date", "Competitor", "Change Type", "Previous Value", "New Value", "Source", "Severity"]
    for col_idx, header in enumerate(log_headers, start=1):
        cell = ws_changelog.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill_core
    
    auto_fit_columns(ws_changelog)
    
    # ============== SHEET 7: Data Dictionary ==============
    ws_dict = wb.create_sheet("Data Dictionary")
    ws_dict["A1"] = "Data Dictionary - Column Definitions"
    ws_dict["A1"].font = Font(bold=True, size=14)
    
    dict_headers = ["Column Name", "Category", "Data Type", "Description", "Example", "Source"]
    for col_idx, header in enumerate(dict_headers, start=1):
        cell = ws_dict.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill_core
    
    # Add column definitions
    definitions = [
        ("Competitor Name", "Core", "Text", "Official company name", "Phreesia", "Website"),
        ("Website", "Core", "URL", "Primary company website", "https://www.phreesia.com", "Direct"),
        ("Status", "Core", "Dropdown", "Active, Inactive, Acquired, Watch", "Active", "Manual"),
        ("Threat Level", "Core", "Dropdown", "High, Medium, Low, Watch", "High", "Analysis"),
        ("Last Updated", "Core", "Date", "Date of last data update", "2026-01-15", "System"),
        ("Pricing Model", "Pricing", "Dropdown", "How they charge customers", "Per Visit", "Pricing Page"),
        ("Base Price", "Pricing", "Currency", "Starting/base price", "$3.00", "Pricing Page"),
        ("Customer Count", "Market", "Number", "Estimated customer count", "3000+", "Website/PR"),
        ("Employee Count", "Company", "Number", "Current headcount", "1500", "LinkedIn"),
        ("Funding Total", "Company", "Currency", "Total funding raised", "$300M", "Crunchbase"),
    ]
    
    for row_idx, (col_name, category, dtype, desc, example, source) in enumerate(definitions, start=4):
        ws_dict.cell(row=row_idx, column=1, value=col_name)
        ws_dict.cell(row=row_idx, column=2, value=category)
        ws_dict.cell(row=row_idx, column=3, value=dtype)
        ws_dict.cell(row=row_idx, column=4, value=desc)
        ws_dict.cell(row=row_idx, column=5, value=example)
        ws_dict.cell(row=row_idx, column=6, value=source)
    
    auto_fit_columns(ws_dict)
    
    # Apply white background to all cells in all sheets
    for sheet in wb.worksheets:
        apply_white_background(sheet)
    
    # Save workbook
    output_path = r"C:\Users\conno\Downloads\Certify_Health_Intelv1\Certify_Intel_Dashboard_Template.xlsx"
    wb.save(output_path)
    print(f"Excel template created: {output_path}")
    return output_path

if __name__ == "__main__":
    create_excel_template()
