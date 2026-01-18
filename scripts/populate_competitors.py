"""
Certify Intel - Competitor Data Population
Populates the Excel template with initial competitor data from existing research.
AUTO-FIT COLUMNS: All columns are sized to fit the widest content.
WHITE BACKGROUND: All cells have white fill (no empty/no-fill cells).
"""
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

# White fill for all data cells
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

def auto_fit_columns(worksheet):
    """Auto-fit all columns to the widest content."""
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        # Add padding and set minimum/maximum widths
        adjusted_width = max_length + 3
        if adjusted_width < 12:
            adjusted_width = 12
        if adjusted_width > 60:
            adjusted_width = 60
        worksheet.column_dimensions[column_letter].width = adjusted_width

def apply_white_background(worksheet, max_row=100, max_col=50):
    """Apply white background to all cells in the worksheet."""
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = worksheet.cell(row=row, column=col)
            # Only apply white fill if cell doesn't already have a colored fill (headers)
            if cell.fill.fgColor.rgb in (None, '00000000', 'FFFFFFFF') or cell.fill.fill_type is None:
                cell.fill = WHITE_FILL


def populate_competitors():
    """Populate the Excel template with initial competitor data."""
    
    # Load the template
    wb = load_workbook(r"C:\Users\conno\Downloads\Certify_Health_Intelv1\Certify_Intel_Dashboard_Template.xlsx")
    ws = wb["Competitors"]
    
    # Initial competitor data (15 priority competitors)
    competitors = [
        {
            "name": "Phreesia",
            "website": "https://www.phreesia.com",
            "status": "Active",
            "threat": "High",
            "pricing_model": "Per Visit",
            "base_price": "$3.00",
            "price_unit": "per visit",
            "products": "Intake; Payments; Scheduling",
            "features": "Digital intake, Eligibility verification, Patient payments, Appointment reminders",
            "integrations": "Epic; Cerner; Allscripts; athenahealth",
            "certifications": "HIPAA; SOC2; HITRUST",
            "segments": "Health Systems; Large Practices",
            "size_focus": "Large (50+)",
            "geo_focus": "US National",
            "customer_count": "3000+",
            "acquisition_rate": "50+ new logos/year",
            "key_customers": "Listed on website",
            "g2_rating": "4.5",
            "employees": "1500+",
            "growth_rate": "15% YoY",
            "founded": "2005",
            "hq": "Raleigh, NC",
            "funding": "$300M+",
            "round": "Public (NYSE: PHR)",
            "investors": "Public",
            "traffic": "500K+",
            "social": "50K+",
            "launches": "AI-powered intake (2024)",
            "news": "25",
        },
        {
            "name": "Clearwave",
            "website": "https://www.clearwaveinc.com",
            "status": "Active",
            "threat": "High",
            "pricing_model": "Per Provider",
            "base_price": "$200",
            "price_unit": "per provider/month",
            "products": "Intake; Eligibility; Payments",
            "features": "Patient self-service, Insurance verification, Co-pay collection, Check-in kiosks",
            "integrations": "Multiple EHR vendors",
            "certifications": "HIPAA",
            "segments": "Specialty Practices; Ambulatory",
            "size_focus": "Medium (15-50)",
            "geo_focus": "US National",
            "customer_count": "500+",
            "acquisition_rate": "Monthly announcements",
            "key_customers": "Posted on LinkedIn",
            "g2_rating": "4.2",
            "employees": "200+",
            "growth_rate": "20% YoY",
            "founded": "2003",
            "hq": "Atlanta, GA",
            "funding": "$50M+",
            "round": "Series C",
            "investors": "Private equity backed",
            "traffic": "50K+",
            "social": "5K+",
            "launches": "Clearwave AI launch",
            "news": "15",
        },
        {
            "name": "Epion Health",
            "website": "https://epionhealth.com",
            "status": "Active",
            "threat": "Medium",
            "pricing_model": "Per Location",
            "base_price": "Unknown",
            "price_unit": "per month",
            "products": "Intake; Check-in",
            "features": "Patient check-in, Digital forms, Wait time management",
            "integrations": "Major EHR vendors",
            "certifications": "HIPAA",
            "segments": "Health Systems; Enterprise",
            "size_focus": "Large (50+)",
            "geo_focus": "US National",
            "customer_count": "300+",
            "acquisition_rate": "Unknown",
            "key_customers": "Health systems",
            "g2_rating": "4.0",
            "employees": "300+",
            "growth_rate": "Acquired",
            "founded": "2011",
            "hq": "Acquired by Kyruus",
            "funding": "$40M",
            "round": "Acquisition",
            "investors": "Kyruus",
            "traffic": "25K+",
            "social": "5K",
            "launches": "Kyruus integration",
            "news": "5",
        },
        {
            "name": "Tonic Health",
            "website": "https://tonicforhealth.com",
            "status": "Active",
            "threat": "Medium",
            "pricing_model": "Subscription",
            "base_price": "Unknown",
            "price_unit": "per month",
            "products": "Intake Forms; Patient Engagement",
            "features": "Smart forms, Conditional logic, EHR integration",
            "integrations": "Limited",
            "certifications": "HIPAA",
            "segments": "SMB Practices",
            "size_focus": "Small (1-15)",
            "geo_focus": "US",
            "customer_count": "Unknown",
            "acquisition_rate": "Unknown",
            "key_customers": "Unknown",
            "g2_rating": "4.0",
            "employees": "50+",
            "growth_rate": "Unknown",
            "founded": "2015",
            "hq": "Unknown",
            "funding": "Unknown",
            "round": "Seed/Series A",
            "investors": "Unknown",
            "traffic": "10K",
            "social": "1K",
            "launches": "Unknown",
            "news": "2",
        },
        {
            "name": "AdvancedMD",
            "website": "https://www.advancedmd.com",
            "status": "Active",
            "threat": "Medium",
            "pricing_model": "Per Provider",
            "base_price": "$400",
            "price_unit": "per provider/month",
            "products": "PM; EHR; Intake; RCM",
            "features": "Full practice management, EHR, Billing, Patient portal",
            "integrations": "Own EHR ecosystem",
            "certifications": "HIPAA; SOC2",
            "segments": "Independent Practices",
            "size_focus": "Medium (15-50)",
            "geo_focus": "US National",
            "customer_count": "40000+",
            "acquisition_rate": "Strong",
            "key_customers": "Many case studies",
            "g2_rating": "4.3",
            "employees": "800+",
            "growth_rate": "10% YoY",
            "founded": "1999",
            "hq": "Salt Lake City, UT",
            "funding": "$100M+",
            "round": "PE Owned",
            "investors": "Global Payments",
            "traffic": "100K+",
            "social": "25K+",
            "launches": "Telehealth updates",
            "news": "20",
        },
        {
            "name": "Kareo (Tebra)",
            "website": "https://www.kareo.com",
            "status": "Active",
            "threat": "Medium",
            "pricing_model": "Tiered",
            "base_price": "$150",
            "price_unit": "starting price",
            "products": "PM; EHR; Billing",
            "features": "Practice management, Billing, EHR, Patient engagement",
            "integrations": "Many third-party integrations",
            "certifications": "HIPAA",
            "segments": "Independent Practices",
            "size_focus": "Small (1-15)",
            "geo_focus": "US National",
            "customer_count": "75000+",
            "acquisition_rate": "High volume",
            "key_customers": "Case studies available",
            "g2_rating": "4.1",
            "employees": "1000+",
            "growth_rate": "Merged",
            "founded": "2004",
            "hq": "Irvine, CA",
            "funding": "$100M+",
            "round": "Merged with PatientPop",
            "investors": "PE Backed",
            "traffic": "200K+",
            "social": "20K+",
            "launches": "Tebra rebrand",
            "news": "15",
        },
        {
            "name": "SimplePractice",
            "website": "https://www.simplepractice.com",
            "status": "Active",
            "threat": "Low",
            "pricing_model": "Subscription",
            "base_price": "$29",
            "price_unit": "per month starting",
            "products": "PM; Intake; Telehealth",
            "features": "Practice management, Telehealth, Client portal, Billing",
            "integrations": "Limited to behavioral health",
            "certifications": "HIPAA",
            "segments": "Behavioral Health; Solo",
            "size_focus": "Solo",
            "geo_focus": "US",
            "customer_count": "100000+",
            "acquisition_rate": "Very high",
            "key_customers": "Testimonials",
            "g2_rating": "4.6",
            "employees": "500+",
            "growth_rate": "20% YoY",
            "founded": "2012",
            "hq": "Los Angeles, CA",
            "funding": "$50M+",
            "round": "Series B",
            "investors": "Insight Partners",
            "traffic": "400K+",
            "social": "100K+",
            "launches": "AI notes",
            "news": "10",
        },
        {
            "name": "NextGen Healthcare",
            "website": "https://www.nextgen.com",
            "status": "Active",
            "threat": "Low",
            "pricing_model": "Enterprise",
            "base_price": "Custom",
            "price_unit": "custom pricing",
            "products": "EHR; PM; RCM",
            "features": "Enterprise EHR, Population health, RCM, Interoperability",
            "integrations": "Own ecosystem",
            "certifications": "All major",
            "segments": "Ambulatory; Specialty",
            "size_focus": "Large (50+)",
            "geo_focus": "US National",
            "customer_count": "100000+",
            "acquisition_rate": "Enterprise sales",
            "key_customers": "Health systems",
            "g2_rating": "3.8",
            "employees": "3000+",
            "growth_rate": "5% YoY",
            "founded": "1974",
            "hq": "Atlanta, GA",
            "funding": "$500M+",
            "round": "Public (NASDAQ: NXGN)",
            "investors": "Public",
            "traffic": "300K+",
            "social": "20K+",
            "launches": "Platform updates",
            "news": "30",
        },
        {
            "name": "Solutionreach",
            "website": "https://www.solutionreach.com",
            "status": "Active",
            "threat": "Medium",
            "pricing_model": "Per Location",
            "base_price": "$299",
            "price_unit": "per location/month",
            "products": "Patient Engagement; Reminders",
            "features": "Appointment reminders, Reviews, Recalls, Two-way texting",
            "integrations": "Many EHR/PM systems",
            "certifications": "HIPAA",
            "segments": "Multi-specialty; Dental",
            "size_focus": "Medium (15-50)",
            "geo_focus": "US National",
            "customer_count": "25000+",
            "acquisition_rate": "Good",
            "key_customers": "Many logos",
            "g2_rating": "4.2",
            "employees": "400+",
            "growth_rate": "10% YoY",
            "founded": "2000",
            "hq": "Lehi, UT",
            "funding": "$50M+",
            "round": "PE Owned",
            "investors": "Marlin Equity",
            "traffic": "75K+",
            "social": "10K+",
            "launches": "AI messaging",
            "news": "12",
        },
        {
            "name": "Doctible",
            "website": "https://www.doctible.com",
            "status": "Active",
            "threat": "Medium",
            "pricing_model": "Per Location",
            "base_price": "Unknown",
            "price_unit": "per location",
            "products": "Patient Engagement; Reviews",
            "features": "Online scheduling, Reviews, Reminders, Patient communication",
            "integrations": "Multiple dental/medical",
            "certifications": "HIPAA",
            "segments": "Dental; Medical",
            "size_focus": "Small (1-15)",
            "geo_focus": "US",
            "customer_count": "Unknown",
            "acquisition_rate": "Growing",
            "key_customers": "On website",
            "g2_rating": "4.7",
            "employees": "100+",
            "growth_rate": "Unknown",
            "founded": "2015",
            "hq": "San Diego, CA",
            "funding": "Unknown",
            "round": "Unknown",
            "investors": "Unknown",
            "traffic": "50K+",
            "social": "5K",
            "launches": "Unknown",
            "news": "5",
        },
        {
            "name": "Luma Health",
            "website": "https://www.lumahealth.io",
            "status": "Active",
            "threat": "Medium",
            "pricing_model": "Per Location",
            "base_price": "Unknown",
            "price_unit": "per location",
            "products": "Patient Engagement; Scheduling",
            "features": "Patient scheduling, Referral management, Waitlist, Communication",
            "integrations": "Major EHRs",
            "certifications": "HIPAA",
            "segments": "Health Systems; Specialty",
            "size_focus": "Medium (15-50)",
            "geo_focus": "US National",
            "customer_count": "300+",
            "acquisition_rate": "Growing",
            "key_customers": "Health systems",
            "g2_rating": "4.8",
            "employees": "200+",
            "growth_rate": "Growing",
            "founded": "2015",
            "hq": "San Francisco, CA",
            "funding": "$50M+",
            "round": "Series C",
            "investors": "VC Backed",
            "traffic": "30K+",
            "social": "5K",
            "launches": "AI scheduling",
            "news": "8",
        },
        {
            "name": "Relatient",
            "website": "https://www.relatient.com",
            "status": "Active",
            "threat": "Medium",
            "pricing_model": "Per Location",
            "base_price": "Unknown",
            "price_unit": "per location",
            "products": "Patient Engagement; Self-Service",
            "features": "Self-scheduling, Reminders, Recalls, Digital registration",
            "integrations": "Multiple EHRs",
            "certifications": "HIPAA",
            "segments": "Health Systems; Multi-specialty",
            "size_focus": "Medium (15-50)",
            "geo_focus": "US National",
            "customer_count": "500+",
            "acquisition_rate": "Good",
            "key_customers": "On website",
            "g2_rating": "4.3",
            "employees": "200+",
            "growth_rate": "10% YoY",
            "founded": "2014",
            "hq": "Franklin, TN",
            "funding": "$30M+",
            "round": "PE Backed",
            "investors": "Cressey & Company",
            "traffic": "20K+",
            "social": "3K",
            "launches": "Unknown",
            "news": "5",
        },
        {
            "name": "IntakeQ",
            "website": "https://intakeq.com",
            "status": "Active",
            "threat": "Low",
            "pricing_model": "Subscription",
            "base_price": "$49",
            "price_unit": "per month",
            "products": "Intake Forms; Scheduling",
            "features": "Online forms, Appointment scheduling, Client portal, Notes",
            "integrations": "Limited",
            "certifications": "HIPAA",
            "segments": "Wellness; Behavioral",
            "size_focus": "Solo",
            "geo_focus": "US; Canada",
            "customer_count": "10000+",
            "acquisition_rate": "Self-service",
            "key_customers": "Testimonials",
            "g2_rating": "4.8",
            "employees": "50+",
            "growth_rate": "Growing",
            "founded": "2014",
            "hq": "Canada",
            "funding": "Bootstrapped",
            "round": "None",
            "investors": "None",
            "traffic": "100K+",
            "social": "5K",
            "launches": "Unknown",
            "news": "3",
        },
        {
            "name": "DrChrono",
            "website": "https://www.drchrono.com",
            "status": "Active",
            "threat": "Low",
            "pricing_model": "Per Provider",
            "base_price": "$199",
            "price_unit": "per provider/month",
            "products": "EHR; PM; RCM",
            "features": "Mobile EHR, Practice management, Medical billing, iPad first",
            "integrations": "Many third-party",
            "certifications": "HIPAA",
            "segments": "Independent Practices",
            "size_focus": "Small (1-15)",
            "geo_focus": "US",
            "customer_count": "Unknown",
            "acquisition_rate": "Unknown",
            "key_customers": "Listed",
            "g2_rating": "4.0",
            "employees": "200+",
            "growth_rate": "Acquired",
            "founded": "2009",
            "hq": "Sunnyvale, CA",
            "funding": "$50M+",
            "round": "Acquired by EverHealth",
            "investors": "EverHealth",
            "traffic": "50K+",
            "social": "15K+",
            "launches": "EverHealth integration",
            "news": "8",
        },
        {
            "name": "CareCloud",
            "website": "https://www.carecloud.com",
            "status": "Active",
            "threat": "Medium",
            "pricing_model": "Per Provider",
            "base_price": "$300",
            "price_unit": "per provider/month",
            "products": "EHR; PM; RCM",
            "features": "Cloud EHR, Practice management, Revenue cycle, Analytics",
            "integrations": "Own platform",
            "certifications": "HIPAA; SOC2",
            "segments": "Multi-specialty",
            "size_focus": "Medium (15-50)",
            "geo_focus": "US National",
            "customer_count": "5000+",
            "acquisition_rate": "Moderate",
            "key_customers": "On website",
            "g2_rating": "3.9",
            "employees": "500+",
            "growth_rate": "5% YoY",
            "founded": "2009",
            "hq": "Somerset, NJ",
            "funding": "$100M+",
            "round": "Public (NASDAQ: CCLD)",
            "investors": "Public",
            "traffic": "30K+",
            "social": "5K",
            "launches": "Platform updates",
            "news": "10",
        },
    ]
    
    # Populate data starting from row 2
    today = datetime.now().strftime("%Y-%m-%d")
    
    for row_idx, comp in enumerate(competitors, start=2):
        ws.cell(row=row_idx, column=1, value=comp["name"])
        ws.cell(row=row_idx, column=2, value=comp["website"])
        ws.cell(row=row_idx, column=3, value=comp["status"])
        ws.cell(row=row_idx, column=4, value=comp["threat"])
        ws.cell(row=row_idx, column=5, value=today)
        ws.cell(row=row_idx, column=6, value="")  # Notes
        ws.cell(row=row_idx, column=7, value="")  # Data Quality Score
        ws.cell(row=row_idx, column=8, value=comp["pricing_model"])
        ws.cell(row=row_idx, column=9, value=comp["base_price"])
        ws.cell(row=row_idx, column=10, value=comp["price_unit"])
        ws.cell(row=row_idx, column=11, value=comp["products"])
        ws.cell(row=row_idx, column=12, value=comp["features"])
        ws.cell(row=row_idx, column=13, value=comp["integrations"])
        ws.cell(row=row_idx, column=14, value=comp["certifications"])
        ws.cell(row=row_idx, column=15, value=comp["segments"])
        ws.cell(row=row_idx, column=16, value=comp["size_focus"])
        ws.cell(row=row_idx, column=17, value=comp["geo_focus"])
        ws.cell(row=row_idx, column=18, value=comp["customer_count"])
        ws.cell(row=row_idx, column=19, value=comp["acquisition_rate"])
        ws.cell(row=row_idx, column=20, value=comp["key_customers"])
        ws.cell(row=row_idx, column=21, value=comp["g2_rating"])
        ws.cell(row=row_idx, column=22, value=comp["employees"])
        ws.cell(row=row_idx, column=23, value=comp["growth_rate"])
        ws.cell(row=row_idx, column=24, value=comp["founded"])
        ws.cell(row=row_idx, column=25, value=comp["hq"])
        ws.cell(row=row_idx, column=26, value=comp["funding"])
        ws.cell(row=row_idx, column=27, value=comp["round"])
        ws.cell(row=row_idx, column=28, value=comp["investors"])
        ws.cell(row=row_idx, column=29, value=comp["traffic"])
        ws.cell(row=row_idx, column=30, value=comp["social"])
        ws.cell(row=row_idx, column=31, value=comp["launches"])
        ws.cell(row=row_idx, column=32, value=comp["news"])
    # Auto-fit all columns in all sheets
    for sheet_name in wb.sheetnames:
        auto_fit_columns(wb[sheet_name])
    
    # Apply white background to all cells in all sheets
    for sheet_name in wb.sheetnames:
        apply_white_background(wb[sheet_name])
    
    # Save the populated workbook
    output_path = r"C:\Users\conno\Downloads\Certify_Health_Intelv1\Certify_Intel_Dashboard_MWP.xlsx"
    wb.save(output_path)
    print(f"MWP Dashboard created with {len(competitors)} competitors: {output_path}")
    return output_path

if __name__ == "__main__":
    populate_competitors()
